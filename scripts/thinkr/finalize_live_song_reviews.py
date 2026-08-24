"""Apply the final human song-resolution decisions to both local workbooks.

No database or network access is performed. Both workbooks are backed up and
validated before replacement.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT: Final = Path(__file__).resolve().parents[2]
WORKING: Final = ROOT / "private-data" / "imports" / "thinkr" / "working"
SONG_REVIEW: Final = WORKING / "live_setlist_song_review.xlsx"
REGISTRATION_REVIEW: Final = WORKING / "live_setlist_song_registration_review.xlsx"
BACKUPS: Final = WORKING / "backups"

EXPECTED: Final = {
    "LINK_EXISTING": 212,
    "NEW_VERSION": 21,
    "NEW_SONG": 18,
    "SPECIAL_UNRESOLVED": 3,
}
LAGTRAIN: Final = ("ラグトレイン", "ヰ世界情緒", 126)
SAME_SINGER: Final = ("生きていく光は", "ヰ世界情緒", 69)
TO_NEW_SONG: Final = {
    ("ロマンチック願望(アーカイブ限定)", "V.W.P"),
    ("自由に捕らわれる(アーカイブ限定)", "V.W.P"),
}

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
SAFE_FILL: Final = PatternFill("solid", fgColor="E8F3EB")
THIN: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def header(sheet) -> list[str]:
    return [str(cell.value or "") for cell in sheet[1]]


def rows(sheet) -> list[dict[str, object]]:
    fields = header(sheet)
    return [dict(zip(fields, values, strict=True)) for values in sheet.iter_rows(min_row=2, values_only=True)]


def target_song(songs: list[dict[str, object]], song_id: int) -> dict[str, object]:
    matches = [row for row in songs if int(row["id"]) == song_id]
    if len(matches) != 1:
        raise RuntimeError(f"songs_reference #{song_id}を一意に取得できません。")
    return matches[0]


def link_resolution(row: dict[str, object], song: dict[str, object], note: str) -> None:
    row.update(
        resolution_status="LINK_EXISTING",
        song_id=song["id"],
        reference_song_id=song["id"],
        reference_song_title=song["title"],
        reference_artist_credit=song["artist_credit"],
        reference_version_type=song["version_type"],
        reference_version_name=song["version_name"],
        song_group_id=song["song_group_id"],
        decision_note=note,
    )


def finalize_resolutions(review_rows: list[dict[str, object]], songs: list[dict[str, object]]) -> None:
    for title, artist, song_id in (LAGTRAIN, SAME_SINGER):
        matches = [row for row in review_rows if (row["song_title_raw"], row["artist_credit_raw"]) == (title, artist)]
        if len(matches) != 1:
            raise RuntimeError(f"最終LINK対象を一意に取得できません: {title} / {artist}")
        note = (
            "夜河世界名義とヰ世界情緒名義は実質的歌唱者が同一。raw artistはlive行に保持。"
            if song_id == 69
            else "人間判断によりCANDY LIVEで歌唱した通常版へリンク確定。"
        )
        link_resolution(matches[0], target_song(songs, song_id), note)

    for row in review_rows:
        if (row["song_title_raw"], row["artist_credit_raw"]) in TO_NEW_SONG:
            row.update(
                resolution_status="NEW_SONG",
                song_id=None,
                reference_song_id=None,
                reference_song_title=None,
                reference_artist_credit=None,
                reference_version_type=None,
                reference_version_name=None,
                song_group_id=None,
                decision_note=(
                    "既存song/group候補なし。アーカイブ限定は行annotationに残し、"
                    "V.W.P standardの新規song/group候補へ再分類。"
                ),
            )
    counts = Counter(str(row["resolution_status"]) for row in review_rows)
    if dict(counts) != EXPECTED:
        raise RuntimeError(f"最終分類件数が想定外です: {dict(counts)}")


def style_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.border = THIN
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def rebuild_resolution_sheet(workbook, name: str, fields: list[str], data: list[dict[str, object]], index: int) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])
    sheet = workbook.create_sheet(name, index)
    sheet.append(fields)
    for row in data:
        sheet.append([row.get(field) for field in fields])
    style_sheet(sheet)


def update_song_review(workbook) -> list[dict[str, object]]:
    review_fields = header(workbook["song_id_review"])
    review_rows = rows(workbook["song_id_review"])
    songs = rows(workbook["songs_reference"])
    finalize_resolutions(review_rows, songs)
    ordered = sorted(review_rows, key=lambda row: (str(row["resolution_status"]), str(row["song_title_raw"]), str(row["artist_credit_raw"])))
    rebuild_resolution_sheet(workbook, "song_id_review", review_fields, ordered, 0)
    rebuild_resolution_sheet(workbook, "new_versions_review", review_fields, [r for r in ordered if r["resolution_status"] == "NEW_VERSION"], 1)
    rebuild_resolution_sheet(workbook, "new_songs_review", review_fields, [r for r in ordered if r["resolution_status"] == "NEW_SONG"], 2)
    rebuild_resolution_sheet(workbook, "unresolved_review", review_fields, [r for r in ordered if r["resolution_status"] == "SPECIAL_UNRESOLVED"], 3)
    return ordered


def strip_archive(title: str) -> str:
    return title.replace("(アーカイブ限定)", "").strip()


def update_registration_review(workbook, final_rows: list[dict[str, object]]) -> None:
    old_versions = {(r["song_title_raw"], r["artist_credit_raw"]): r for r in rows(workbook["new_versions"])}
    old_songs = {(r["song_title_raw"], r["artist_credit_raw"]): r for r in rows(workbook["new_songs"])}
    version_fields = header(workbook["new_versions"])
    song_fields = header(workbook["new_songs"])
    link_fields = header(workbook["link_existing"])
    special_fields = header(workbook["special_unresolved"])

    links = []
    versions = []
    new_songs = []
    special = []
    for resolution in final_rows:
        key = (resolution["song_title_raw"], resolution["artist_credit_raw"])
        status = resolution["resolution_status"]
        if status == "LINK_EXISTING":
            links.append({
                "song_title_raw": resolution["song_title_raw"],
                "artist_credit_raw": resolution["artist_credit_raw"],
                "occurrence_count": resolution["occurrence_count"],
                "example_performances": resolution["example_performances"],
                "song_id": resolution["song_id"],
                "reference_song_title": resolution["reference_song_title"],
                "reference_artist_credit": resolution["reference_artist_credit"],
                "reference_version_type": resolution["reference_version_type"],
                "reference_version_name": resolution["reference_version_name"],
                "song_group_id": resolution["song_group_id"],
                "annotation_raw": resolution["annotation_raw"],
                "note": resolution["decision_note"],
            })
        elif status == "NEW_VERSION":
            row = dict(old_versions[key])
            row.update(
                review_decision="ACCEPT",
                confirmed_song_group_id=row["proposed_song_group_id"],
                confirmed_version_type=row["proposed_version_type"],
                confirmed_version_name=row["proposed_version_name"],
                approval_class="SAFE_TO_APPROVE",
                remaining_human_question=None,
            )
            versions.append(row)
        elif status == "NEW_SONG":
            if key in old_songs:
                row = dict(old_songs[key])
            else:
                title = strip_archive(str(resolution["song_title_raw"]))
                row = {field: None for field in song_fields}
                row.update(
                    title=title,
                    artist_credit=resolution["artist_credit_raw"],
                    occurrence_count=resolution["occurrence_count"],
                    example_performances=resolution["example_performances"],
                    song_title_raw=resolution["song_title_raw"],
                    artist_credit_raw=resolution["artist_credit_raw"],
                    existing_title_check="NO_EXACT_OR_NORMALIZED_MATCH",
                    song_group_creation_candidate=True,
                    note=resolution["decision_note"],
                    proposed_title=title,
                    proposed_artist_credit=resolution["artist_credit_raw"],
                    proposed_create_song_group=True,
                    proposed_version_type="standard",
                    proposed_version_name=None,
                    registration_reason="既存song/group候補なし。新規group＋standard songとして登録。",
                    confidence="HIGH",
                    approval_class="SAFE_TO_APPROVE",
                )
            row.update(review_decision="ACCEPT", remaining_human_question=None)
            new_songs.append(row)
        else:
            special.append({
                "song_title_raw": resolution["song_title_raw"],
                "artist_credit_raw": resolution["artist_credit_raw"],
                "occurrence_count": resolution["occurrence_count"],
                "example_performances": resolution["example_performances"],
                "song_id": None,
                "resolution_status": "SPECIAL_UNRESOLVED",
                "note": resolution["decision_note"],
            })

    summary = workbook["summary"]
    summary.delete_rows(2, summary.max_row - 1)
    for status in ("LINK_EXISTING", "NEW_VERSION", "NEW_SONG", "SPECIAL_UNRESOLVED"):
        summary.append([status, EXPECTED[status], 0, "最終人間判断確定。DB未反映。"])
    style_sheet(summary)
    rebuild_resolution_sheet(workbook, "link_existing", link_fields, sorted(links, key=lambda r: (str(r["song_title_raw"]), str(r["artist_credit_raw"]))), 1)
    rebuild_resolution_sheet(workbook, "new_versions", version_fields, sorted(versions, key=lambda r: (str(r["song_title"]), str(r["artist_credit"]))), 2)
    rebuild_resolution_sheet(workbook, "new_songs", song_fields, sorted(new_songs, key=lambda r: (str(r["title"]), str(r["artist_credit"]))), 3)
    rebuild_resolution_sheet(workbook, "special_unresolved", special_fields, sorted(special, key=lambda r: (str(r["song_title_raw"]), str(r["artist_credit_raw"]))), 4)


def timestamped_backup(path: Path) -> Path:
    stamp = datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y%m%d-%H%M%S")
    candidate = BACKUPS / f"{path.stem}.before-final-{stamp}{path.suffix}"
    counter = 1
    while candidate.exists():
        candidate = BACKUPS / f"{path.stem}.before-final-{stamp}-{counter}{path.suffix}"
        counter += 1
    return candidate


def validate(path: Path, registration: bool) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if registration:
            expected = {"link_existing": 212, "new_versions": 21, "new_songs": 18, "special_unresolved": 3}
            for name, count in expected.items():
                if workbook[name].max_row - 1 != count:
                    raise RuntimeError(f"{name}の保存後件数が不正です。")
            if any(r["review_decision"] != "ACCEPT" for r in rows(workbook["new_versions"]) + rows(workbook["new_songs"])):
                raise RuntimeError("登録候補39件がACCEPTではありません。")
        else:
            review = rows(workbook["song_id_review"])
            if Counter(str(r["resolution_status"]) for r in review) != Counter(EXPECTED):
                raise RuntimeError("song reviewの保存後分類が不正です。")
            if workbook["_true_setlist_source"].max_row - 1 != 378:
                raise RuntimeError("TRUE source 378行が維持されていません。")
    finally:
        workbook.close()


def run() -> tuple[Path, Path]:
    for path in (SONG_REVIEW, REGISTRATION_REVIEW):
        if not path.exists():
            raise RuntimeError(f"workbookがありません: {path}")
    song_wb = load_workbook(SONG_REVIEW, read_only=False, data_only=False)
    reg_wb = load_workbook(REGISTRATION_REVIEW, read_only=False, data_only=False)
    handles_and_names = [tempfile.mkstemp(prefix=".final-review-", suffix=".xlsx", dir=WORKING) for _ in range(2)]
    for handle, _ in handles_and_names:
        os.close(handle)
    song_temp, reg_temp = (Path(item[1]) for item in handles_and_names)
    try:
        final_rows = update_song_review(song_wb)
        update_registration_review(reg_wb, final_rows)
        song_wb.save(song_temp)
        reg_wb.save(reg_temp)
        song_wb.close()
        reg_wb.close()
        validate(song_temp, False)
        validate(reg_temp, True)
        BACKUPS.mkdir(parents=True, exist_ok=True)
        song_backup = timestamped_backup(SONG_REVIEW)
        reg_backup = timestamped_backup(REGISTRATION_REVIEW)
        shutil.copy2(SONG_REVIEW, song_backup)
        shutil.copy2(REGISTRATION_REVIEW, reg_backup)
        os.replace(song_temp, SONG_REVIEW)
        os.replace(reg_temp, REGISTRATION_REVIEW)
        return song_backup, reg_backup
    finally:
        try:
            song_wb.close()
            reg_wb.close()
        except Exception:
            pass
        song_temp.unlink(missing_ok=True)
        reg_temp.unlink(missing_ok=True)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        backups = run()
        for status, count in EXPECTED.items():
            print(f"{status}: {count}")
        print(f"backups: {backups[0].name}, {backups[1].name}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError, KeyError) as error:
        print(f"最終review反映を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

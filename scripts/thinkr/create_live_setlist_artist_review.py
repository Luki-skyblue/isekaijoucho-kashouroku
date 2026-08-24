"""Create a local review workbook for live artist participation decisions.

The script performs SELECT requests only. It never updates Supabase and refuses
to overwrite an existing workbook. A hidden source sheet retains stable setlist
entry IDs so a later step can group only human-confirmed TRUE rows by raw title.
"""

from __future__ import annotations

import os
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from import_live_ready import create_anon_client


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
OUTPUT_PATH: Final = (
    REPOSITORY_ROOT
    / "private-data"
    / "imports"
    / "thinkr"
    / "working"
    / "live_setlist_artist_review.xlsx"
)
EXPECTED_PERFORMANCES: Final = 29
EXPECTED_SETLIST_ROWS: Final = 499

ARTIST_FIELDS: Final = [
    "artist_credit_raw",
    "occurrence_count",
    "example_performances",
    "suggested_participation",
    "decision",
    "note",
]
SONG_FIELDS: Final = [
    "id",
    "title",
    "title_kana",
    "sort_title",
    "artist_credit",
    "song_type",
    "version_name",
    "version_type",
    "is_primary_version",
    "first_date",
    "song_group_id",
]
SOURCE_FIELDS: Final = [
    "setlist_entry_id",
    "live_performance_id",
    "performance_title",
    "performance_date",
    "sort_order",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "current_song_id",
    "current_joucho_participation",
]

# Only exact, unambiguously individual non-Joucho credits are suggested FALSE.
# Every group/collaboration spelling remains undecided even when its likely
# participants seem obvious.
EXPLICIT_OTHER_SOLO_ARTISTS: Final = {
    "花譜",
    "理芽",
    "春猿火",
    "幸祜",
    "CIEL",
    "狐子",
}

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
RAW_FILL: Final = PatternFill("solid", fgColor="F5F4F1")
INPUT_FILL: Final = PatternFill("solid", fgColor="FFF6D8")
TRUE_FILL: Final = PatternFill("solid", fgColor="E8F3EB")
FALSE_FILL: Final = PatternFill("solid", fgColor="F2F2F2")
UNKNOWN_FILL: Final = PatternFill("solid", fgColor="FFF0D9")
THIN_BORDER: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def suggested_participation(artist_credit_raw: str | None) -> str:
    if artist_credit_raw and "ヰ世界情緒" in artist_credit_raw:
        return "TRUE"
    if artist_credit_raw in EXPLICIT_OTHER_SOLO_ARTISTS:
        return "FALSE"
    return "?"


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def fetch_data() -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    client = create_anon_client()
    performances = client.select(
        "live_performances",
        [
            ("select", "id,title,performance_date"),
            ("order", "performance_date.asc,id.asc"),
        ],
    )
    setlists = client.select(
        "live_setlist_entries",
        [
            (
                "select",
                "id,live_performance_id,sort_order,setlist_no_raw,song_title_raw,"
                "artist_credit_raw,song_id,joucho_participation",
            ),
            ("order", "live_performance_id.asc,sort_order.asc,id.asc"),
        ],
    )
    songs = client.select(
        "songs",
        [
            ("select", ",".join(SONG_FIELDS)),
            ("order", "title_kana.asc.nullslast,title.asc,id.asc"),
        ],
    )
    if len(performances) != EXPECTED_PERFORMANCES:
        raise RuntimeError(
            f"公開performanceが{EXPECTED_PERFORMANCES}件ではありません: "
            f"{len(performances)}"
        )
    if len(setlists) != EXPECTED_SETLIST_ROWS:
        raise RuntimeError(
            f"公開setlistが{EXPECTED_SETLIST_ROWS}行ではありません: {len(setlists)}"
        )
    performance_ids = {row.get("id") for row in performances}
    if any(row.get("live_performance_id") not in performance_ids for row in setlists):
        raise RuntimeError("公開performanceへ解決できないsetlist rowがあります。")
    return performances, setlists, songs


def build_artist_rows(
    performances: list[dict[str, object]], setlists: list[dict[str, object]]
) -> list[dict[str, object]]:
    title_by_id = {row["id"]: str(row["title"]) for row in performances}
    counts: Counter[str | None] = Counter()
    examples: dict[str | None, set[str]] = defaultdict(set)
    for row in setlists:
        credit = row.get("artist_credit_raw")
        if credit is not None and not isinstance(credit, str):
            raise RuntimeError("artist_credit_rawが文字列またはNULLではありません。")
        counts[credit] += 1
        examples[credit].add(title_by_id[row["live_performance_id"]])

    rows = [
        {
            "artist_credit_raw": credit,
            "occurrence_count": count,
            "example_performances": " / ".join(sorted(examples[credit])[:5]),
            "suggested_participation": suggested_participation(credit),
            "decision": None,
            "note": None,
        }
        for credit, count in counts.items()
    ]
    suggestion_order = {"?": 0, "TRUE": 1, "FALSE": 2}
    return sorted(
        rows,
        key=lambda row: (
            suggestion_order[str(row["suggested_participation"])],
            -int(row["occurrence_count"]),
            str(row["artist_credit_raw"] or ""),
        ),
    )


def build_source_rows(
    performances: list[dict[str, object]], setlists: list[dict[str, object]]
) -> list[dict[str, object]]:
    performance_by_id = {row["id"]: row for row in performances}
    return [
        {
            "setlist_entry_id": row["id"],
            "live_performance_id": row["live_performance_id"],
            "performance_title": performance_by_id[row["live_performance_id"]]["title"],
            "performance_date": performance_by_id[row["live_performance_id"]][
                "performance_date"
            ],
            "sort_order": row["sort_order"],
            "setlist_no_raw": row.get("setlist_no_raw"),
            "song_title_raw": row.get("song_title_raw"),
            "artist_credit_raw": row.get("artist_credit_raw"),
            "current_song_id": row.get("song_id"),
            "current_joucho_participation": row.get("joucho_participation"),
        }
        for row in setlists
    ]


def append_table(
    sheet,
    fields: list[str],
    rows: list[dict[str, object]],
    *,
    text_fields: set[str],
) -> None:
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field) for field in fields])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if fields[cell.column - 1] in text_fields:
                cell.number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
    sheet.row_dimensions[1].height = 28


def style_artist_sheet(sheet, row_count: int) -> None:
    widths = {"A": 34, "B": 16, "C": 62, "D": 24, "E": 16, "F": 48}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, row_count + 2):
        for column in (1, 2, 3, 4):
            sheet.cell(row, column).fill = RAW_FILL
        for column in (5, 6):
            sheet.cell(row, column).fill = INPUT_FILL
    decision_validation = DataValidation(
        type="list", formula1='"TRUE,FALSE,?"', allow_blank=True
    )
    decision_validation.error = "TRUE / FALSE / ? のいずれかを選択してください。"
    decision_validation.errorTitle = "入力値を確認してください"
    decision_validation.prompt = "人間の確定判断を選択します。"
    decision_validation.promptTitle = "joucho participation"
    decision_validation.showErrorMessage = True
    decision_validation.showInputMessage = True
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"E2:E{row_count + 1}")
    sheet.conditional_formatting.add(
        f"E2:E{row_count + 1}",
        FormulaRule(formula=['E2="TRUE"'], fill=TRUE_FILL),
    )
    sheet.conditional_formatting.add(
        f"E2:E{row_count + 1}",
        FormulaRule(formula=['E2="FALSE"'], fill=FALSE_FILL),
    )
    sheet.conditional_formatting.add(
        f"E2:E{row_count + 1}",
        FormulaRule(formula=['E2="?"'], fill=UNKNOWN_FILL),
    )
    sheet["D1"].comment = Comment(
        "参考候補です。DBへ反映されず、decisionを自動入力しません。", "Codex"
    )
    sheet["E1"].comment = Comment(
        "人間が確定する列です。次工程はdecision=TRUEだけをsong_title_raw単位に集計します。",
        "Codex",
    )


def style_songs_sheet(sheet) -> None:
    widths = {
        "A": 10,
        "B": 38,
        "C": 30,
        "D": 30,
        "E": 24,
        "F": 18,
        "G": 22,
        "H": 18,
        "I": 18,
        "J": 16,
        "K": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def style_source_sheet(sheet) -> None:
    widths = {
        "A": 18,
        "B": 20,
        "C": 40,
        "D": 18,
        "E": 14,
        "F": 18,
        "G": 42,
        "H": 34,
        "I": 18,
        "J": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def write_workbook(
    artist_rows: list[dict[str, object]],
    songs: list[dict[str, object]],
    source_rows: list[dict[str, object]],
) -> None:
    if OUTPUT_PATH.exists():
        raise RuntimeError(f"既存review workbookを上書きしません: {OUTPUT_PATH}")
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    artist_sheet = workbook.active
    artist_sheet.title = "artist_participation"
    append_table(
        artist_sheet,
        ARTIST_FIELDS,
        artist_rows,
        text_fields={
            "artist_credit_raw",
            "example_performances",
            "suggested_participation",
            "decision",
            "note",
        },
    )
    style_artist_sheet(artist_sheet, len(artist_rows))

    songs_sheet = workbook.create_sheet("songs_reference")
    append_table(
        songs_sheet,
        SONG_FIELDS,
        songs,
        text_fields=set(SONG_FIELDS) - {"id", "song_group_id"},
    )
    style_songs_sheet(songs_sheet)

    source_sheet = workbook.create_sheet("_setlist_source")
    append_table(
        source_sheet,
        SOURCE_FIELDS,
        source_rows,
        text_fields={
            "performance_title",
            "performance_date",
            "setlist_no_raw",
            "song_title_raw",
            "artist_credit_raw",
            "current_joucho_participation",
        },
    )
    style_source_sheet(source_sheet)
    source_sheet.sheet_state = "hidden"

    workbook.properties.title = "Live setlist artist participation review"
    workbook.properties.description = (
        "Human decisions remain blank. The hidden _setlist_source sheet enables "
        "a later TRUE-only song_title_raw review without refetching or guessing."
    )
    temporary_handle, temporary_name = tempfile.mkstemp(
        prefix=f".{OUTPUT_PATH.stem}-", suffix=".xlsx", dir=OUTPUT_PATH.parent
    )
    os.close(temporary_handle)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        workbook.close()
        os.replace(temporary_path, OUTPUT_PATH)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def validate_workbook(
    artist_rows: list[dict[str, object]], songs_count: int, source_count: int
) -> None:
    workbook = load_workbook(OUTPUT_PATH, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != [
            "artist_participation",
            "songs_reference",
            "_setlist_source",
        ]:
            raise RuntimeError("workbookシート構成が想定外です。")
        artist_sheet = workbook["artist_participation"]
        songs_sheet = workbook["songs_reference"]
        source_sheet = workbook["_setlist_source"]
        if artist_sheet.max_row - 1 != len(artist_rows):
            raise RuntimeError("artist種類数がworkbookと一致しません。")
        if songs_sheet.max_row - 1 != songs_count:
            raise RuntimeError("songs件数がworkbookと一致しません。")
        if source_sheet.max_row - 1 != source_count:
            raise RuntimeError("setlist source件数がworkbookと一致しません。")
        if any(artist_sheet.cell(row, 5).value is not None for row in range(2, artist_sheet.max_row + 1)):
            raise RuntimeError("decision列に初期値が混入しています。")
        if len(artist_sheet.data_validations.dataValidation) != 1:
            raise RuntimeError("decision dropdownを確認できません。")
        if source_sheet.sheet_state != "hidden":
            raise RuntimeError("補助setlist sourceシートがhiddenではありません。")
    finally:
        workbook.close()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        performances, setlists, songs = fetch_data()
        artist_rows = build_artist_rows(performances, setlists)
        source_rows = build_source_rows(performances, setlists)
        write_workbook(artist_rows, songs, source_rows)
        validate_workbook(artist_rows, len(songs), len(source_rows))
        suggestion_types = Counter(
            str(row["suggested_participation"]) for row in artist_rows
        )
        suggestion_rows: Counter[str] = Counter()
        for row in artist_rows:
            suggestion_rows[str(row["suggested_participation"])] += int(
                row["occurrence_count"]
            )
        print(f"workbook created: {OUTPUT_PATH}")
        print(f"artist credit types: {len(artist_rows)}")
        for value in ("TRUE", "FALSE", "?"):
            print(
                f"suggested {value}: types={suggestion_types[value]}, "
                f"rows={suggestion_rows[value]}"
            )
        print(f"songs reference: {len(songs)}")
        print(f"hidden setlist source: {len(source_rows)}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError) as error:
        print(f"artist review生成を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

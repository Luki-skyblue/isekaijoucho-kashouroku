"""Apply approved song resolution decisions to the local review workbook.

The operation is local-only. It never contacts or updates Supabase, preserves
the raw title/artist source sheet, backs up the current workbook, and refuses to
continue if existing human song_id/note input would be overwritten.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
WORKING_ROOT: Final = (
    REPOSITORY_ROOT / "private-data" / "imports" / "thinkr" / "working"
)
WORKBOOK_PATH: Final = WORKING_ROOT / "live_setlist_song_review.xlsx"
BACKUP_ROOT: Final = WORKING_ROOT / "backups"

EXPECTED_SOURCE_ROWS: Final = 378
EXPECTED_PAIR_COUNT: Final = 254
EXPECTED_STATUS_COUNTS: Final = {
    "LINK_EXISTING": 210,
    "NEW_VERSION": 24,
    "NEW_SONG": 16,
    "UNRESOLVED": 1,
    "SPECIAL_UNRESOLVED": 3,
}

SOURCE_HEADERS: Final = [
    "setlist_entry_id",
    "live_performance_id",
    "performance_title",
    "performance_date",
    "sort_order",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "current_song_id",
    "current_joucho_participation",
]
RESOLUTION_HEADERS: Final = [
    "song_title_raw",
    "artist_credit_raw",
    "occurrence_count",
    "example_performances",
    "resolution_status",
    "song_id",
    "reference_song_id",
    "reference_song_title",
    "reference_artist_credit",
    "reference_version_type",
    "reference_version_name",
    "song_group_id",
    "annotation_raw",
    "decision_note",
]

DIFF_LINK_PAIRS: Final = {
    ("new world", "ヰ世界情緒 with Aiobahn"),
    ("ぼくらの逃避行", "ヰ世界情緒 with VALIS"),
    ("エリカの憂い", "ヰ世界情緒 & 星界"),
    ("ノンブレス・オブリージュ", "ヰ世界情緒 feat. 星界"),
    ("ワールド・コーリング", "ヰ世界情緒 feat. 春猿火"),
    ("刻印", "ヰ世界情緒 feat. 幸祜"),
    ("暗闇", "ヰ世界情緒 feat. 花譜"),
    ("機械の声", "V.W.P & V.I.P"),
    ("機械の声", "V.W.P feat. V.I.P"),
    ("泡沫", "ヰ世界情緒 feat. 理芽"),
    ("泡沫", "理芽×ヰ世界情緒"),
    ("牢獄", "ヰ世界情緒 feat. 春猿火"),
    ("生存", "春猿火×ヰ世界情緒"),
    ("異星にいこうね", "ヰ世界情緒 & 星界"),
    ("雛鳥", "花譜 feat. ヰ世界情緒"),
}

AMBIGUOUS_LINK_IDS: Final = {
    ("JANE DOE", "ヰ世界情緒"): 325,
    ("あわく心模様", "ヰ世界情緒"): 164,
    ("祭壇", "V.W.P"): 59,
    ("言霊", "V.W.P"): 49,
    ("言霊", "花譜 feat. 理芽&春猿火&ヰ世界情緒"): 27,
    ("輪廻", "V.W.P"): 68,
    ("鏡面の波", "ヰ世界情緒"): 309,
    ("電脳", "V.W.P"): 50,
    ("魔女", "春猿火×ヰ世界情緒"): 321,
}

AMBIGUOUS_NEW_VERSION_REFERENCE: Final = {
    ("輪廻", "V.W.P & 狐子"): 68,
}
UNRESOLVED_PAIRS: Final = {
    ("ラグトレイン", "ヰ世界情緒"),
}

ANNOTATION_RULES: Final = {
    ("ANEMONE(アーカイブ限定)", "ヰ世界情緒"): (85, "アーカイブ限定"),
    ("ロマンチック願望(アーカイブ限定)", "V.W.P"): (None, "アーカイブ限定"),
    ("変身(アーカイブ限定)", "理芽×ヰ世界情緒"): (None, "アーカイブ限定"),
    ("深淵(アーカイブ限定)", "ヰ世界情緒"): (84, "アーカイブ限定"),
    ("玩具(アーカイブ限定)", "理芽×ヰ世界情緒"): (None, "アーカイブ限定"),
    ("自由に捕らわれる(アーカイブ限定)", "V.W.P"): (None, "アーカイブ限定"),
    ("輪廻(アーカイブ限定)", "ヰ世界情緒"): (86, "アーカイブ限定"),
}

ANNOTATION_NEW_VERSION_REFERENCES: Final = {
    ("変身(アーカイブ限定)", "理芽×ヰ世界情緒"): 81,
    ("玩具(アーカイブ限定)", "理芽×ヰ世界情緒"): 101,
}

ALIAS_LINK_TITLES: Final = {
    ("言霊 (multilingual ver.)", "V.W.P feat. V.I.P"): "言霊 multilingual ver.",
    ("電脳 (multilingual ver.)", "V.W.P feat. V.I.P"): "電脳 multilingual ver.",
    ("Chaining Intentoin", "ヰ世界情緒"): "Chaining Intention",
    ("God knows...", "ヰ世界情緒"): "God knows…",
    ("Hello, Worker", "ヰ世界情緒"): "Hello,Worker",
    ("ODD & ENDS", "ヰ世界情緒"): "ODDS ＆ ENDS",
    ("Paradisus-Paradoxum", "ヰ世界情緒"): "Paradisus-Paradoxm",
    ("ラブカ？", "ヰ世界情緒"): "ラブカ?",
    ("奏（かなで）", "ヰ世界情緒"): "奏",
    ("永遠に枯れぬ花", "ヰ世界情緒"): "永久に枯れぬ花",
}

SPECIAL_UNRESOLVED_PAIRS: Final = {
    ("異世界転調リクヱスト", "ヰ世界情緒 with VALIS"),
    ("異世界転調リクヱスト", "ヰ世界情緒×VALIS"),
    ("祭霊(祭壇+言霊)", "V.W.P"),
}

NEW_SONG_PAIRS: Final = {
    ("Monster", "V.W.P"),
    ("One Last Kiss", "V.W.P"),
    ("See You Again", "V.W.P"),
    ("おどるポンポコリン", "V.W.P"),
    ("ないものねだり", "V.W.P"),
    ("イケナイ太陽", "V.W.P"),
    ("カレンの清掃", "ヰ世界情緒 & 星界"),
    ("ブリキノダンス", "V.W.P"),
    ("ヘビーローテーション", "V.W.P"),
    ("マカロン", "ヰ世界情緒"),
    ("丸の内サディスティック", "V.W.P"),
    ("唱", "V.W.P"),
    ("夢をかなえてドラえもん", "V.W.P"),
    ("女々しくて", "V.W.P"),
    ("新宝島", "V.W.P"),
    ("現象", "V.W.P"),
}

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
RAW_FILL: Final = PatternFill("solid", fgColor="F5F4F1")
LINK_FILL: Final = PatternFill("solid", fgColor="E8F3EB")
VERSION_FILL: Final = PatternFill("solid", fgColor="FFF0D9")
NEW_FILL: Final = PatternFill("solid", fgColor="E9F0FA")
UNRESOLVED_FILL: Final = PatternFill("solid", fgColor="F8E2E2")
THIN_BORDER: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def header(sheet) -> list[str]:
    return [str(cell.value or "") for cell in sheet[1]]


def credit_key(value: object) -> str:
    # Human policy allows whitespace-only differences. Operators, order, feat.,
    # with, and group spellings remain distinct.
    return "".join(str(value or "").split())


def source_rows(workbook) -> list[dict[str, object]]:
    sheet = workbook["_true_setlist_source"]
    fields = header(sheet)
    if fields != SOURCE_HEADERS:
        raise RuntimeError("_true_setlist_sourceの列構成が想定外です。")
    rows = [
        dict(zip(fields, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]
    if len(rows) != EXPECTED_SOURCE_ROWS:
        raise RuntimeError("TRUE sourceが378行ではありません。")
    if any("DISCOTHEQUE" in str(row["song_title_raw"] or "") for row in rows):
        raise RuntimeError(
            "V.W.P DISCOTHEQUEが現在sourceに存在します。composite判断を追加してください。"
        )
    return rows


def songs_reference(workbook) -> list[dict[str, object]]:
    sheet = workbook["songs_reference"]
    fields = header(sheet)
    rows = [
        dict(zip(fields, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]
    ids = [row["id"] for row in rows]
    if len(ids) != len(set(ids)):
        raise RuntimeError("songs_referenceのidがuniqueではありません。")
    return rows


def grouped_pairs(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    counts: Counter[tuple[str, str]] = Counter()
    performances: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in rows:
        title = str(row["song_title_raw"] or "")
        artist = str(row["artist_credit_raw"] or "")
        if not title:
            raise RuntimeError("空のsong_title_rawがあります。")
        pair = (title, artist)
        counts[pair] += 1
        performances[pair].add(str(row["performance_title"]))
    result = [
        {
            "pair": pair,
            "song_title_raw": pair[0],
            "artist_credit_raw": pair[1],
            "occurrence_count": count,
            "example_performances": " / ".join(sorted(performances[pair])[:5]),
        }
        for pair, count in counts.items()
    ]
    if len(result) != EXPECTED_PAIR_COUNT:
        raise RuntimeError(f"title×artist組が254ではありません: {len(result)}")
    return result


def resolve_by_title(songs: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    result: dict[str, list[dict[str, object]]] = defaultdict(list)
    for song in songs:
        result[str(song["title"])].append(song)
    return result


def reference_fields(song: dict[str, object] | None) -> dict[str, object]:
    if song is None:
        return {
            "reference_song_id": None,
            "reference_song_title": None,
            "reference_artist_credit": None,
            "reference_version_type": None,
            "reference_version_name": None,
            "song_group_id": None,
        }
    return {
        "reference_song_id": song["id"],
        "reference_song_title": song["title"],
        "reference_artist_credit": song.get("artist_credit"),
        "reference_version_type": song.get("version_type"),
        "reference_version_name": song.get("version_name"),
        "song_group_id": song.get("song_group_id"),
    }


def resolution_row(
    source: dict[str, object],
    status: str,
    reference: dict[str, object] | None,
    *,
    annotation: str | None = None,
    note: str | None = None,
) -> dict[str, object]:
    row = {
        "song_title_raw": source["song_title_raw"],
        "artist_credit_raw": source["artist_credit_raw"],
        "occurrence_count": source["occurrence_count"],
        "example_performances": source["example_performances"],
        "resolution_status": status,
        "song_id": reference["id"] if status == "LINK_EXISTING" and reference else None,
        "annotation_raw": annotation,
        "decision_note": note,
    }
    row.update(reference_fields(reference))
    return row


def create_resolutions(
    pairs: list[dict[str, object]], songs: list[dict[str, object]]
) -> list[dict[str, object]]:
    by_title = resolve_by_title(songs)
    by_id = {int(song["id"]): song for song in songs}
    resolutions: list[dict[str, object]] = []
    for source in pairs:
        pair = source["pair"]
        title, artist = pair
        exact = by_title.get(title, [])
        if len(exact) == 1 and credit_key(exact[0].get("artist_credit")) == credit_key(artist):
            resolutions.append(resolution_row(source, "LINK_EXISTING", exact[0]))
            continue

        if len(exact) == 1:
            if pair in DIFF_LINK_PAIRS:
                resolutions.append(
                    resolution_row(
                        source,
                        "LINK_EXISTING",
                        exact[0],
                        note="歌唱者構成同一。feat./×/with/&または名義順のみの表記差。",
                    )
                )
            else:
                resolutions.append(
                    resolution_row(
                        source,
                        "NEW_VERSION",
                        exact[0],
                        note="同タイトルの既存songと歌唱者構成が異なる。",
                    )
                )
            continue

        if len(exact) > 1:
            if pair in AMBIGUOUS_LINK_IDS:
                target = by_id[AMBIGUOUS_LINK_IDS[pair]]
                if target not in exact:
                    raise RuntimeError(f"ambiguous選択songが同タイトル候補ではありません: {pair}")
                resolutions.append(
                    resolution_row(source, "LINK_EXISTING", target, note="人間が既存versionを指定。")
                )
            elif pair in AMBIGUOUS_NEW_VERSION_REFERENCE:
                target = by_id[AMBIGUOUS_NEW_VERSION_REFERENCE[pair]]
                resolutions.append(
                    resolution_row(
                        source,
                        "NEW_VERSION",
                        target,
                        note="既存候補に同じ歌唱者構成がない。",
                    )
                )
            elif pair in UNRESOLVED_PAIRS:
                resolutions.append(
                    resolution_row(
                        source,
                        "UNRESOLVED",
                        None,
                        note="Twitter cover / standard のどちらか要確認。",
                    )
                )
            else:
                raise RuntimeError(f"未処理のambiguous pairです: {pair}")
            continue

        if pair in ANNOTATION_RULES:
            linked_id, annotation = ANNOTATION_RULES[pair]
            if linked_id is not None:
                target = by_id[linked_id]
                resolutions.append(
                    resolution_row(
                        source,
                        "LINK_EXISTING",
                        target,
                        annotation=annotation,
                        note="raw titleは保持し、アーカイブ限定を行annotationとして分離。",
                    )
                )
            else:
                reference_id = ANNOTATION_NEW_VERSION_REFERENCES.get(pair)
                target = by_id[reference_id] if reference_id is not None else None
                resolutions.append(
                    resolution_row(
                        source,
                        "NEW_VERSION",
                        target,
                        annotation=annotation,
                        note="アーカイブ限定はversion理由にせず、歌唱者構成に合う既存versionなし。",
                    )
                )
            continue

        if pair in ALIAS_LINK_TITLES:
            target_title = ALIAS_LINK_TITLES[pair]
            targets = by_title.get(target_title, [])
            if len(targets) != 1:
                raise RuntimeError(f"title variation候補を一意解決できません: {pair}")
            resolutions.append(
                resolution_row(
                    source,
                    "LINK_EXISTING",
                    targets[0],
                    note=f"raw titleを保持し、照合時のみ {target_title!r} と対応。",
                )
            )
            continue

        if pair in SPECIAL_UNRESOLVED_PAIRS:
            if title == "祭霊(祭壇+言霊)":
                note = "祭壇＋言霊の複合歌唱候補。composite/mashupモデル待ち。"
            else:
                note = (
                    "Anima IIIの『異世界転調リクヱスト / ぼくらの逃避行 with. VALIS』"
                    "複合トラック候補。通常songへ単純リンクしない。"
                )
            resolutions.append(
                resolution_row(source, "SPECIAL_UNRESOLVED", None, note=note)
            )
            continue

        if pair in NEW_SONG_PAIRS:
            resolutions.append(
                resolution_row(
                    source,
                    "NEW_SONG",
                    None,
                    note="既存songsにtitle候補なし。新規song登録候補。",
                )
            )
            continue
        raise RuntimeError(f"未処理のno-title-match pairです: {pair}")

    counts = Counter(str(row["resolution_status"]) for row in resolutions)
    if dict(counts) != EXPECTED_STATUS_COUNTS:
        raise RuntimeError(f"resolution件数が想定外です: {dict(counts)}")
    if any(
        (row["resolution_status"] == "LINK_EXISTING") != (row["song_id"] is not None)
        for row in resolutions
    ):
        raise RuntimeError("LINK_EXISTINGとsong_id NULL状態が一致しません。")
    return sorted(
        resolutions,
        key=lambda row: (
            str(row["resolution_status"]),
            str(row["song_title_raw"]),
            str(row["artist_credit_raw"]),
        ),
    )


def append_table(sheet, rows: list[dict[str, object]]) -> None:
    sheet.append(RESOLUTION_HEADERS)
    for row in rows:
        sheet.append([row.get(field) for field in RESOLUTION_HEADERS])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.border = THIN_BORDER
    status_fill = {
        "LINK_EXISTING": LINK_FILL,
        "NEW_VERSION": VERSION_FILL,
        "NEW_SONG": NEW_FILL,
        "UNRESOLVED": UNRESOLVED_FILL,
        "SPECIAL_UNRESOLVED": UNRESOLVED_FILL,
    }
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
        for column in (1, 2, 4, 5, 8, 9, 10, 11, 13, 14):
            row[column - 1].number_format = "@"
        fill = status_fill[str(row[4].value)]
        row[4].fill = fill
        if row[12].value:
            row[12].fill = VERSION_FILL
    widths = {
        "A": 42,
        "B": 36,
        "C": 14,
        "D": 58,
        "E": 24,
        "F": 12,
        "G": 18,
        "H": 36,
        "I": 36,
        "J": 20,
        "K": 24,
        "L": 16,
        "M": 22,
        "N": 62,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{len(rows) + 1}"
    sheet["F1"].comment = Comment(
        "LINK_EXISTINGだけ値を持ちます。DBにはまだ反映していません。", "Codex"
    )
    sheet["M1"].comment = Comment(
        "アーカイブ限定等、元のsetlist行固有annotationです。raw titleはA列に保持します。",
        "Codex",
    )


def replace_review_sheets(workbook, resolutions: list[dict[str, object]]) -> None:
    old_review = workbook["song_id_review"]
    old_header = header(old_review)
    if "song_id" not in old_header or "note" not in old_header:
        raise RuntimeError("既存song_id_reviewの列構成が想定外です。")
    song_id_column = old_header.index("song_id") + 1
    note_column = old_header.index("note") + 1
    if any(
        old_review.cell(row, song_id_column).value is not None
        or old_review.cell(row, note_column).value is not None
        for row in range(2, old_review.max_row + 1)
    ):
        raise RuntimeError("既存の人間song_id/note入力があるため上書きしません。")

    for sheet_name in (
        "song_id_review",
        "new_versions_review",
        "new_songs_review",
        "unresolved_review",
    ):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    review = workbook.create_sheet("song_id_review", 0)
    append_table(review, resolutions)
    versions = workbook.create_sheet("new_versions_review", 1)
    append_table(
        versions,
        [row for row in resolutions if row["resolution_status"] == "NEW_VERSION"],
    )
    new_songs = workbook.create_sheet("new_songs_review", 2)
    append_table(
        new_songs,
        [row for row in resolutions if row["resolution_status"] == "NEW_SONG"],
    )
    unresolved = workbook.create_sheet("unresolved_review", 3)
    append_table(
        unresolved,
        [
            row
            for row in resolutions
            if row["resolution_status"] in {"UNRESOLVED", "SPECIAL_UNRESOLVED"}
        ],
    )
    workbook.properties.description = (
        "Resolution is keyed by exact (song_title_raw, artist_credit_raw). "
        "Operators/order are ignored only where explicitly approved by human policy."
    )


def backup_path() -> Path:
    timestamp = datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y%m%d-%H%M%S")
    candidate = BACKUP_ROOT / f"live_setlist_song_review.before-resolution-{timestamp}.xlsx"
    suffix = 1
    while candidate.exists():
        candidate = BACKUP_ROOT / (
            f"live_setlist_song_review.before-resolution-{timestamp}-{suffix}.xlsx"
        )
        suffix += 1
    return candidate


def validate_saved(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        expected_sheets = {
            "song_id_review",
            "new_versions_review",
            "new_songs_review",
            "unresolved_review",
            "songs_reference",
            "_true_setlist_source",
        }
        if set(workbook.sheetnames) != expected_sheets:
            raise RuntimeError(f"保存後シート構成が想定外です: {workbook.sheetnames}")
        review = workbook["song_id_review"]
        if header(review) != RESOLUTION_HEADERS or review.max_row - 1 != 254:
            raise RuntimeError("保存後resolution reviewが254組ではありません。")
        fields = header(review)
        rows = [
            dict(zip(fields, values, strict=True))
            for values in review.iter_rows(min_row=2, values_only=True)
        ]
        counts = Counter(str(row["resolution_status"]) for row in rows)
        if dict(counts) != EXPECTED_STATUS_COUNTS:
            raise RuntimeError("保存後status件数が一致しません。")
        if workbook["new_versions_review"].max_row - 1 != 24:
            raise RuntimeError("NEW_VERSION補助sheetが24組ではありません。")
        if workbook["new_songs_review"].max_row - 1 != 16:
            raise RuntimeError("NEW_SONG補助sheetが16組ではありません。")
        if workbook["unresolved_review"].max_row - 1 != 4:
            raise RuntimeError("未解決補助sheetが4組ではありません。")
        if workbook["_true_setlist_source"].max_row - 1 != 378:
            raise RuntimeError("元のTRUE source 378行が維持されていません。")
        if workbook["_true_setlist_source"].sheet_state != "hidden":
            raise RuntimeError("TRUE source sheetがhiddenではありません。")
    finally:
        workbook.close()


def run() -> Path:
    if not WORKBOOK_PATH.exists():
        raise RuntimeError(f"song reviewがありません: {WORKBOOK_PATH}")
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    handle, temporary_name = tempfile.mkstemp(
        prefix=".live-song-resolution-", suffix=".xlsx", dir=WORKING_ROOT
    )
    os.close(handle)
    temporary = Path(temporary_name)
    backup: Path | None = None
    try:
        original_songs = tuple(
            tuple(cell.value for cell in row)
            for row in workbook["songs_reference"].iter_rows()
        )
        original_source = tuple(
            tuple(cell.value for cell in row)
            for row in workbook["_true_setlist_source"].iter_rows()
        )
        source = source_rows(workbook)
        songs = songs_reference(workbook)
        resolutions = create_resolutions(grouped_pairs(source), songs)
        replace_review_sheets(workbook, resolutions)
        workbook.save(temporary)
        workbook.close()
        validate_saved(temporary)

        check = load_workbook(temporary, read_only=True, data_only=False)
        try:
            if original_songs != tuple(
                tuple(values)
                for values in check["songs_reference"].iter_rows(values_only=True)
            ):
                raise RuntimeError("songs_referenceが意図せず変化しています。")
            if original_source != tuple(
                tuple(values)
                for values in check["_true_setlist_source"].iter_rows(values_only=True)
            ):
                raise RuntimeError("TRUE setlist sourceが意図せず変化しています。")
        finally:
            check.close()

        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
        backup = backup_path()
        shutil.copy2(WORKBOOK_PATH, backup)
        os.replace(temporary, WORKBOOK_PATH)
        return backup
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        temporary.unlink(missing_ok=True)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        backup = run()
        for status, count in EXPECTED_STATUS_COUNTS.items():
            print(f"{status}: {count}")
        print(f"workbook updated: {WORKBOOK_PATH}")
        print(f"backup created: {backup}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError, KeyError) as error:
        print(f"song resolution反映を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

"""Apply approved artist decisions and create the manual song ID review.

This is a local-only workbook operation. It backs up the existing artist
review, changes only its decision column, and creates a separate song review.
No database or network access is performed.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
WORKING_ROOT: Final = (
    REPOSITORY_ROOT / "private-data" / "imports" / "thinkr" / "working"
)
ARTIST_REVIEW_PATH: Final = WORKING_ROOT / "live_setlist_artist_review.xlsx"
SONG_REVIEW_PATH: Final = WORKING_ROOT / "live_setlist_song_review.xlsx"
BACKUP_ROOT: Final = WORKING_ROOT / "backups"

EXPECTED_ARTIST_TYPES: Final = 61
EXPECTED_SETLIST_ROWS: Final = 499
EXPECTED_TRUE_TYPES: Final = 27
EXPECTED_TRUE_ROWS: Final = 378
EXPECTED_FALSE_TYPES: Final = 34
EXPECTED_FALSE_ROWS: Final = 121

VWP_TRUE_CREDITS: Final = {
    "V.W.P",
    "V.W.P & 狐子",
    "V.W.P feat. V.I.P",
    "V.W.P & V.I.P",
    "V.W.P & V.I.P with VALIS",
}

ARTIST_HEADERS: Final = [
    "artist_credit_raw",
    "occurrence_count",
    "example_performances",
    "suggested_participation",
    "decision",
    "note",
]
SONG_REVIEW_HEADERS: Final = [
    "song_title_raw",
    "occurrence_count",
    "example_performances",
    "artist_credit_examples",
    "song_id",
    "note",
]
SOURCE_HEADERS: Final = [
    "setlist_entry_id",
    "live_performance_id",
    "performance_title",
    "performance_date",
    "sort_order",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "current_song_id",
    "current_joucho_participation",
]

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
RAW_FILL: Final = PatternFill("solid", fgColor="F5F4F1")
INPUT_FILL: Final = PatternFill("solid", fgColor="FFF6D8")
THIN_BORDER: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def header(sheet) -> list[str]:
    return [str(cell.value or "") for cell in sheet[1]]


def sheet_values(workbook) -> dict[str, tuple[tuple[object, ...], ...]]:
    return {
        sheet_name: tuple(
            tuple(cell.value for cell in row)
            for row in workbook[sheet_name].iter_rows()
        )
        for sheet_name in workbook.sheetnames
    }


def expected_decision(suggested: str, artist_credit_raw: str) -> str:
    if suggested == "TRUE" or artist_credit_raw in VWP_TRUE_CREDITS:
        return "TRUE"
    if suggested in {"FALSE", "?"}:
        return "FALSE"
    raise RuntimeError(f"未知のsuggested participationです: {suggested!r}")


def apply_decisions(workbook) -> tuple[dict[str, str], list[dict[str, object]]]:
    artist_sheet = workbook["artist_participation"]
    if header(artist_sheet) != ARTIST_HEADERS:
        raise RuntimeError("artist_participationの列構成が想定外です。")
    if artist_sheet.max_row - 1 != EXPECTED_ARTIST_TYPES:
        raise RuntimeError("artist表記が61種類ではありません。")

    decisions: dict[str, str] = {}
    artist_rows: list[dict[str, object]] = []
    for row_number in range(2, artist_sheet.max_row + 1):
        credit_value = artist_sheet.cell(row_number, 1).value
        credit = "" if credit_value is None else str(credit_value)
        occurrence_count = int(artist_sheet.cell(row_number, 2).value)
        suggested = str(artist_sheet.cell(row_number, 4).value)
        current_decision = artist_sheet.cell(row_number, 5).value
        decision = expected_decision(suggested, credit)
        if current_decision not in {None, "", decision}:
            raise RuntimeError(
                f"既存decisionが今回の確定値と競合しています: {credit!r}"
            )
        if credit in decisions:
            raise RuntimeError(f"artist_credit_rawが重複しています: {credit!r}")
        decisions[credit] = decision
        artist_rows.append(
            {
                "artist_credit_raw": credit,
                "occurrence_count": occurrence_count,
                "decision": decision,
            }
        )
        artist_sheet.cell(row_number, 5).value = decision

    type_counts = Counter(str(row["decision"]) for row in artist_rows)
    row_counts: Counter[str] = Counter()
    for row in artist_rows:
        row_counts[str(row["decision"])] += int(row["occurrence_count"])
    if type_counts != Counter({"TRUE": EXPECTED_TRUE_TYPES, "FALSE": EXPECTED_FALSE_TYPES}):
        raise RuntimeError(f"artist decision種類数が想定外です: {dict(type_counts)}")
    if row_counts != Counter({"TRUE": EXPECTED_TRUE_ROWS, "FALSE": EXPECTED_FALSE_ROWS}):
        raise RuntimeError(f"artist decision行数が想定外です: {dict(row_counts)}")
    return decisions, artist_rows


def source_rows(workbook) -> list[dict[str, object]]:
    source_sheet = workbook["_setlist_source"]
    source_header = header(source_sheet)
    if source_header != SOURCE_HEADERS:
        raise RuntimeError("_setlist_sourceの列構成が想定外です。")
    rows = [
        dict(zip(source_header, values, strict=True))
        for values in source_sheet.iter_rows(min_row=2, values_only=True)
    ]
    if len(rows) != EXPECTED_SETLIST_ROWS:
        raise RuntimeError("setlist sourceが499行ではありません。")
    entry_ids = [row["setlist_entry_id"] for row in rows]
    if len(set(entry_ids)) != len(entry_ids):
        raise RuntimeError("setlist_entry_idがuniqueではありません。")
    return rows


def expand_decisions(
    decisions: dict[str, str], rows: list[dict[str, object]]
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    true_rows: list[dict[str, object]] = []
    false_rows: list[dict[str, object]] = []
    for row in rows:
        raw_value = row["artist_credit_raw"]
        credit = "" if raw_value is None else str(raw_value)
        if credit not in decisions:
            raise RuntimeError(
                f"decisionへ展開できないartist_credit_rawがあります: {credit!r}"
            )
        if decisions[credit] == "TRUE":
            true_rows.append(row)
        else:
            false_rows.append(row)
    if len(true_rows) != EXPECTED_TRUE_ROWS or len(false_rows) != EXPECTED_FALSE_ROWS:
        raise RuntimeError(
            "499 setlist rowsへのdecision展開件数が想定外です: "
            f"TRUE={len(true_rows)}, FALSE={len(false_rows)}"
        )
    return true_rows, false_rows


def build_song_rows(true_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    occurrences: Counter[str] = Counter()
    performances: dict[str, set[str]] = defaultdict(set)
    artists: dict[str, set[str]] = defaultdict(set)
    for row in true_rows:
        title_value = row["song_title_raw"]
        if title_value is None or not str(title_value).strip():
            raise RuntimeError("TRUE行に空のsong_title_rawがあります。")
        title = str(title_value)
        occurrences[title] += 1
        performances[title].add(str(row["performance_title"]))
        raw_artist = row["artist_credit_raw"]
        artists[title].add("" if raw_artist is None else str(raw_artist))
    return sorted(
        (
            {
                "song_title_raw": title,
                "occurrence_count": count,
                "example_performances": " / ".join(sorted(performances[title])[:5]),
                "artist_credit_examples": " / ".join(sorted(artists[title])[:5]),
                "song_id": None,
                "note": None,
            }
            for title, count in occurrences.items()
        ),
        key=lambda row: (-int(row["occurrence_count"]), str(row["song_title_raw"])),
    )


def append_table(
    sheet,
    fields: list[str],
    rows: list[dict[str, object]],
    *,
    text_fields: set[str],
) -> None:
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field) for field in fields])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if fields[cell.column - 1] in text_fields:
                cell.number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"


def copy_songs_reference(source_workbook, target_workbook) -> int:
    source = source_workbook["songs_reference"]
    target = target_workbook.create_sheet("songs_reference")
    rows = list(source.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("songs_referenceが空です。")
    for values in rows:
        target.append(list(values))
    fields = [str(value or "") for value in rows[0]]
    for cell in target[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.border = THIN_BORDER
    for row in target.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if fields[cell.column - 1] not in {"id", "song_group_id"}:
                cell.number_format = "@"
    target.freeze_panes = "A2"
    target.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)}"
    widths = [10, 38, 30, 30, 24, 18, 22, 18, 18, 16, 14]
    for index, width in enumerate(widths, start=1):
        target.column_dimensions[get_column_letter(index)].width = width
    return len(rows) - 1


def create_song_workbook(
    source_workbook, song_rows: list[dict[str, object]], true_rows: list[dict[str, object]]
):
    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "song_id_review"
    append_table(
        review_sheet,
        SONG_REVIEW_HEADERS,
        song_rows,
        text_fields={
            "song_title_raw",
            "example_performances",
            "artist_credit_examples",
            "note",
        },
    )
    widths = {"A": 44, "B": 16, "C": 64, "D": 50, "E": 14, "F": 48}
    for column, width in widths.items():
        review_sheet.column_dimensions[column].width = width
    for row_number in range(2, review_sheet.max_row + 1):
        for column in range(1, 5):
            review_sheet.cell(row_number, column).fill = RAW_FILL
        for column in (5, 6):
            review_sheet.cell(row_number, column).fill = INPUT_FILL
    song_id_validation = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True
    )
    song_id_validation.error = "song_idには1以上の整数を入力してください。"
    song_id_validation.showErrorMessage = True
    review_sheet.add_data_validation(song_id_validation)
    song_id_validation.add(f"E2:E{review_sheet.max_row}")
    review_sheet["E1"].comment = Comment(
        "自動確定されていません。songs_referenceを参照して人間が入力します。", "Codex"
    )

    songs_count = copy_songs_reference(source_workbook, workbook)

    source_sheet = workbook.create_sheet("_true_setlist_source")
    append_table(
        source_sheet,
        SOURCE_HEADERS,
        true_rows,
        text_fields={
            "performance_title",
            "performance_date",
            "setlist_no_raw",
            "song_title_raw",
            "artist_credit_raw",
            "current_joucho_participation",
        },
    )
    source_sheet.sheet_state = "hidden"
    workbook.properties.title = "Live setlist song ID review"
    workbook.properties.description = (
        "Only artist decision TRUE rows are grouped by exact song_title_raw. "
        "song_id remains a human input and can later be expanded through the hidden source."
    )
    return workbook, songs_count


def unique_backup_path() -> Path:
    timestamp = datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y%m%d-%H%M%S")
    candidate = BACKUP_ROOT / f"live_setlist_artist_review.before-decisions-{timestamp}.xlsx"
    suffix = 1
    while candidate.exists():
        candidate = BACKUP_ROOT / (
            f"live_setlist_artist_review.before-decisions-{timestamp}-{suffix}.xlsx"
        )
        suffix += 1
    return candidate


def temporary_xlsx(parent: Path, stem: str) -> Path:
    handle, name = tempfile.mkstemp(prefix=f".{stem}-", suffix=".xlsx", dir=parent)
    os.close(handle)
    return Path(name)


def validate_artist_temp(
    path: Path,
    original_values: dict[str, tuple[tuple[object, ...], ...]],
    decisions: dict[str, str],
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        current_values = sheet_values(workbook)
        if current_values.keys() != original_values.keys():
            raise RuntimeError("artist workbookのシート構成が変化しています。")
        for sheet_name in original_values:
            before = original_values[sheet_name]
            after = current_values[sheet_name]
            if len(before) != len(after):
                raise RuntimeError("artist workbookの行数が変化しています。")
            for row_index, (before_row, after_row) in enumerate(
                zip(before, after, strict=True), start=1
            ):
                for column_index, (before_value, after_value) in enumerate(
                    zip(before_row, after_row, strict=True), start=1
                ):
                    if sheet_name == "artist_participation" and row_index >= 2 and column_index == 5:
                        credit_value = after[row_index - 1][0]
                        credit = "" if credit_value is None else str(credit_value)
                        if after_value != decisions[credit]:
                            raise RuntimeError("decision保存値が確定値と一致しません。")
                    elif before_value != after_value:
                        raise RuntimeError(
                            f"decision以外が変化しています: {sheet_name}!R{row_index}C{column_index}"
                        )
    finally:
        workbook.close()


def validate_song_temp(
    path: Path, expected_song_rows: int, expected_songs_reference: int
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != [
            "song_id_review",
            "songs_reference",
            "_true_setlist_source",
        ]:
            raise RuntimeError("song reviewのシート構成が想定外です。")
        review = workbook["song_id_review"]
        songs = workbook["songs_reference"]
        source = workbook["_true_setlist_source"]
        if review.max_row - 1 != expected_song_rows:
            raise RuntimeError("unique song title件数が保存結果と一致しません。")
        if songs.max_row - 1 != expected_songs_reference:
            raise RuntimeError("songs_reference件数が一致しません。")
        if source.max_row - 1 != EXPECTED_TRUE_ROWS:
            raise RuntimeError("TRUE setlist sourceが378行ではありません。")
        if any(review.cell(row, 5).value is not None for row in range(2, review.max_row + 1)):
            raise RuntimeError("song_id人間入力欄に初期値が混入しています。")
        if source.sheet_state != "hidden":
            raise RuntimeError("TRUE source sheetがhiddenではありません。")
    finally:
        workbook.close()


def run() -> tuple[int, Path]:
    if not ARTIST_REVIEW_PATH.exists():
        raise RuntimeError(f"artist reviewがありません: {ARTIST_REVIEW_PATH}")
    if SONG_REVIEW_PATH.exists():
        raise RuntimeError(f"既存song reviewを上書きしません: {SONG_REVIEW_PATH}")

    workbook = load_workbook(ARTIST_REVIEW_PATH, read_only=False, data_only=False)
    artist_temp = temporary_xlsx(WORKING_ROOT, "artist-review-decisions")
    song_temp = temporary_xlsx(WORKING_ROOT, "song-review")
    backup_path: Path | None = None
    try:
        original_values = sheet_values(workbook)
        decisions, _ = apply_decisions(workbook)
        rows = source_rows(workbook)
        true_rows, _ = expand_decisions(decisions, rows)
        songs = build_song_rows(true_rows)
        song_workbook, songs_reference_count = create_song_workbook(
            workbook, songs, true_rows
        )

        workbook.save(artist_temp)
        song_workbook.save(song_temp)
        song_workbook.close()
        workbook.close()
        validate_artist_temp(artist_temp, original_values, decisions)
        validate_song_temp(song_temp, len(songs), songs_reference_count)

        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
        backup_path = unique_backup_path()
        shutil.copy2(ARTIST_REVIEW_PATH, backup_path)
        os.replace(artist_temp, ARTIST_REVIEW_PATH)
        try:
            os.replace(song_temp, SONG_REVIEW_PATH)
        except Exception:
            shutil.copy2(backup_path, ARTIST_REVIEW_PATH)
            raise
        return len(songs), backup_path
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        artist_temp.unlink(missing_ok=True)
        song_temp.unlink(missing_ok=True)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        unique_songs, backup_path = run()
        print("participation: TRUE types=27 rows=378")
        print("participation: FALSE types=34 rows=121")
        print("participation: ? types=0 rows=0")
        print(f"unique TRUE song_title_raw: {unique_songs}")
        print(f"artist review updated: {ARTIST_REVIEW_PATH}")
        print(f"song review created: {SONG_REVIEW_PATH}")
        print(f"backup created: {backup_path}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError) as error:
        print(f"review decision反映を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

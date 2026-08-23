"""Create a protected Excel workbook for human review of THINKR Wiki events."""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
CSV_PATH: Final = (
    REPOSITORY_ROOT
    / "private-data"
    / "imports"
    / "thinkr"
    / "working"
    / "events_review.csv"
)
XLSX_PATH: Final = CSV_PATH.with_suffix(".xlsx")
FIELDS: Final = [
    "source_category",
    "year",
    "page_key",
    "event_date_raw",
    "event_title_raw",
    "source_url",
    "import_target",
    "review_note",
]
COLUMN_WIDTHS: Final = {
    "A": 18,
    "B": 9,
    "C": 16,
    "D": 18,
    "E": 68,
    "F": 54,
    "G": 16,
    "H": 42,
}


def read_rows() -> list[dict[str, str]]:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if reader.fieldnames != FIELDS:
            raise RuntimeError(
                "events_review.csvの列構成が想定と異なるため、Excelを作成しません。"
            )

        rows = list(reader)
        if any(row["import_target"] or row["review_note"] for row in rows):
            raise RuntimeError(
                "確認列に入力済みの値があるため、空欄の初期Excelを作成しません。"
            )
        return rows


def write_workbook(rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "events_review"

    header_fill = PatternFill(fill_type="solid", fgColor="E7E5E1")
    for column_index, field_name in enumerate(FIELDS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=field_name)
        cell.data_type = "s"
        cell.number_format = "@"
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, field_name in enumerate(FIELDS, start=1):
            value = row[field_name]
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.data_type = "s"
            cell.number_format = "@"
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=field_name in {"event_title_raw", "source_url", "review_note"},
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:H{len(rows) + 1}"
    worksheet.row_dimensions[1].height = 24
    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width

    validation_sheet = workbook.create_sheet("_validation")
    for row_index, value in enumerate(("TRUE", "FALSE", "?"), start=1):
        cell = validation_sheet.cell(row=row_index, column=1, value=value)
        cell.data_type = "s"
        cell.number_format = "@"
    validation_sheet.sheet_state = "hidden"

    import_target_validation = DataValidation(
        type="list",
        formula1="'_validation'!$A$1:$A$3",
        allow_blank=True,
    )
    import_target_validation.error = "TRUE、FALSE、? のいずれかを選択してください。"
    import_target_validation.errorTitle = "入力値を確認してください"
    import_target_validation.prompt = "TRUE、FALSE、? から選択できます。"
    import_target_validation.promptTitle = "取り込み対象"
    import_target_validation.showErrorMessage = True
    import_target_validation.showInputMessage = True
    worksheet.add_data_validation(import_target_validation)
    import_target_validation.add(f"G2:G{len(rows) + 1}")

    workbook.save(XLSX_PATH)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    if XLSX_PATH.exists():
        print(
            f"既存のExcel選別結果を保護するため、上書きしません: {XLSX_PATH}",
            file=sys.stderr,
        )
        return 1

    try:
        rows = read_rows()
        write_workbook(rows)
    except (OSError, RuntimeError) as error:
        print(f"Excelの作成に失敗しました: {error}", file=sys.stderr)
        return 1

    print(f"作成件数: {len(rows)}")
    print(f"出力先: {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

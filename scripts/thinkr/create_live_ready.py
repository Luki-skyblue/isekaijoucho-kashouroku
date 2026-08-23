"""Create validated, local-only ready CSVs for the first live data pilot.

This script reads collected raw CSVs and the human-edited live review workbook.
It never connects to Supabase or WIKIWIKI. Output is created atomically and an
existing pilot directory is never overwritten.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import sys
import tempfile
import warnings
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore", message="Data Validation extension is not supported.*"
)


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
DATA_ROOT: Final = REPOSITORY_ROOT / "private-data" / "imports" / "thinkr"
EVENTS_RAW_PATH: Final = DATA_ROOT / "raw" / "events_raw.csv"
SETLISTS_RAW_PATH: Final = DATA_ROOT / "raw" / "setlists_raw.csv"
LIVE_REVIEW_PATH: Final = DATA_ROOT / "working" / "live_review.xlsx"
READY_ROOT: Final = DATA_ROOT / "ready"

PILOT_NAME: Final = "pilot-001"
OUTPUT_PATH: Final = READY_ROOT / PILOT_NAME
PILOT_PAGE_KEYS: Final = ("2021.1023", "2026.0501", "2026.0502")
SETLIST_COUNTS: Final = {
    "2021.1023": 21,
    "2026.0501": 23,
    "2026.0502": 21,
}

# These values are explicitly approved pilot review decisions. page_key is used
# only as a local performance_key; it is not a public or database identity ID.
PILOT_PERFORMANCES: Final = {
    "2021.1023": {
        "title": "Anima",
        "artist_credit": "ヰ世界情緒",
        "performance_date": "2021-10-23",
        "local_group_key": None,
        "group_sort_order": None,
    },
    "2026.0501": {
        "title": "Flower Closet",
        "artist_credit": "ヰ世界情緒",
        "performance_date": "2026-05-01",
        "local_group_key": "isekaijoucho-2days-live-2026",
        "group_sort_order": 1,
    },
    "2026.0502": {
        "title": "Anima Re:birth",
        "artist_credit": "ヰ世界情緒",
        "performance_date": "2026-05-02",
        "local_group_key": "isekaijoucho-2days-live-2026",
        "group_sort_order": 2,
    },
}

EVENT_GROUP_FIELDS: Final = [
    "local_group_key",
    "title",
    "title_kana",
    "sort_title",
    "notes",
]
PERFORMANCE_FIELDS: Final = [
    "performance_key",
    "local_group_key",
    "group_sort_order",
    "title",
    "title_kana",
    "sort_title",
    "artist_credit",
    "performance_date",
    "format_label",
    "image_url",
    "venue",
    "streaming_platforms",
    "notes",
    "published_at",
    "is_listed",
]
SOURCE_FIELDS: Final = [
    "performance_key",
    "source_type",
    "source_url",
    "source_key",
    "event_title_raw",
    "event_date_raw",
    "source_category_raw",
    "notes",
]
SETLIST_FIELDS: Final = [
    "performance_key",
    "sort_order",
    "entry_type",
    "setlist_no_raw",
    "song_id",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "marker_label",
    "notes",
    "joucho_participation",
]

EVENTS_RAW_FIELDS: Final = [
    "source_category",
    "year",
    "page_key",
    "event_date_raw",
    "event_title_raw",
    "source_url",
    "import_target",
]
SETLISTS_RAW_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "setlist_table_index",
    "setlist_section_raw",
    "setlist_table_caption_raw",
    "setlist_table_headers_raw",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "setlist_row_cells_raw",
    "source_url",
]
EVENT_REVIEW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "source_url",
    "import_target",
    "title",
    "title_kana",
    "artist_credit",
    "event_type",
    "performance_name",
    "performance_date",
    "venue",
    "official_url",
    "review_note",
]
SETLIST_REVIEW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "import_target",
    "setlist_table_index",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "source_url",
    "song_id",
    "joucho_participation",
    "review_note",
]

EVENT_RAW_REVIEW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "source_url",
]
SETLIST_RAW_REVIEW_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "setlist_table_index",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "source_url",
]


def read_csv(path: Path, expected_fields: list[str]) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if reader.fieldnames != expected_fields:
            raise RuntimeError(f"{path.name}の列構成が想定と異なります。")
        return list(reader)


def workbook_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def read_review_sheet(sheet_name: str, expected_fields: list[str]) -> list[dict[str, str]]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Data Validation extension is not supported.*"
        )
        workbook = load_workbook(LIVE_REVIEW_PATH, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"live_review.xlsxに{sheet_name}シートがありません。")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows, ()))
        if headers != expected_fields:
            raise RuntimeError(
                f"live_review.xlsxの{sheet_name}シートの列構成が想定と異なります。"
            )
        return [
            {
                field: workbook_cell_text(value)
                for field, value in zip(expected_fields, row, strict=True)
            }
            for row in rows
        ]
    finally:
        workbook.close()


def unique_rows_by_key(
    rows: list[dict[str, str]], key_field: str, source_name: str
) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        key = row[key_field]
        if key in result:
            raise RuntimeError(f"{source_name}に重複する{key_field}があります: {key}")
        result[key] = row
    return result


def optional_text(value: str) -> str | None:
    return None if value == "" else value


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as input_file:
        for chunk in iter(lambda: input_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_ready_rows() -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    raw_events = read_csv(EVENTS_RAW_PATH, EVENTS_RAW_FIELDS)
    raw_setlists = read_csv(SETLISTS_RAW_PATH, SETLISTS_RAW_FIELDS)
    review_events = read_review_sheet("events", EVENT_REVIEW_FIELDS)
    review_setlists = read_review_sheet("setlists", SETLIST_REVIEW_FIELDS)

    raw_events_by_key = unique_rows_by_key(raw_events, "page_key", "events_raw.csv")
    review_events_by_key = unique_rows_by_key(
        review_events, "page_key", "live_review.xlsx events"
    )

    event_group_rows: list[dict[str, object]] = [
        {
            "local_group_key": "isekaijoucho-2days-live-2026",
            "title": "ヰ世界情緒 2DAYS LIVE",
            "title_kana": None,
            "sort_title": None,
            "notes": None,
        }
    ]
    performance_rows: list[dict[str, object]] = []
    source_rows: list[dict[str, object]] = []

    for page_key in PILOT_PAGE_KEYS:
        if page_key not in raw_events_by_key or page_key not in review_events_by_key:
            raise RuntimeError(f"pilot対象のイベントが入力にありません: {page_key}")
        raw = raw_events_by_key[page_key]
        review = review_events_by_key[page_key]
        approved = PILOT_PERFORMANCES[page_key]

        for field in EVENT_RAW_REVIEW_FIELDS:
            if review[field] != raw[field]:
                raise RuntimeError(
                    f"{page_key}の{field}がrawとreviewで一致しません。"
                )
        if review["import_target"] != "TRUE":
            raise RuntimeError(f"{page_key}のimport_targetがTRUEではありません。")
        for field in ("title", "artist_credit", "performance_date"):
            if review[field] != approved[field]:
                raise RuntimeError(
                    f"{page_key}の{field}が承認済みpilot値と一致しません。"
                )
        for field in (
            "title_kana",
            "event_type",
            "performance_name",
            "venue",
            "official_url",
            "review_note",
        ):
            if review[field] != "":
                raise RuntimeError(
                    f"{page_key}の未使用review列{field}に値があります。"
                )

        performance_rows.append(
            {
                "performance_key": page_key,
                "local_group_key": approved["local_group_key"],
                "group_sort_order": approved["group_sort_order"],
                "title": review["title"],
                "title_kana": optional_text(review["title_kana"]),
                "sort_title": None,
                "artist_credit": review["artist_credit"],
                "performance_date": review["performance_date"],
                "format_label": optional_text(review["event_type"]),
                "image_url": None,
                "venue": optional_text(review["venue"]),
                # JSON distinguishes an empty PostgreSQL text[] from CSV NULL.
                "streaming_platforms": json.dumps([], ensure_ascii=False),
                "notes": None,
                "published_at": None,
                "is_listed": False,
            }
        )
        source_rows.append(
            {
                "performance_key": page_key,
                "source_type": "thinkr_wiki",
                "source_url": raw["source_url"],
                # source_key identifies the source page; it is not a performance ID.
                "source_key": page_key,
                "event_title_raw": raw["event_title_raw"],
                "event_date_raw": raw["event_date_raw"],
                "source_category_raw": raw["source_category"],
                "notes": None,
            }
        )

    raw_setlists_by_key = {page_key: [] for page_key in PILOT_PAGE_KEYS}
    for row in raw_setlists:
        if row["page_key"] in raw_setlists_by_key:
            raw_setlists_by_key[row["page_key"]].append(row)

    review_setlists_by_key = {page_key: [] for page_key in PILOT_PAGE_KEYS}
    for row in review_setlists:
        if row["page_key"] in review_setlists_by_key:
            review_setlists_by_key[row["page_key"]].append(row)

    setlist_rows: list[dict[str, object]] = []
    for page_key in PILOT_PAGE_KEYS:
        raw_rows = raw_setlists_by_key[page_key]
        review_rows = review_setlists_by_key[page_key]
        if len(raw_rows) != SETLIST_COUNTS[page_key]:
            raise RuntimeError(
                f"{page_key}のrawセットリスト件数が想定外です: {len(raw_rows)}"
            )
        if len(review_rows) != len(raw_rows):
            raise RuntimeError(f"{page_key}のrawとreviewの行数が一致しません。")

        for row_number, (raw, review) in enumerate(
            zip(raw_rows, review_rows, strict=True), start=1
        ):
            for field in SETLIST_RAW_REVIEW_FIELDS:
                if review[field] != raw[field]:
                    raise RuntimeError(
                        f"{page_key}のセットリスト{row_number}行目の{field}が"
                        "rawとreviewで一致しません。"
                    )
            if review["source_category"] != raw_events_by_key[page_key]["source_category"]:
                raise RuntimeError(
                    f"{page_key}のセットリスト{row_number}行目の分類が一致しません。"
                )
            if review["import_target"] != "TRUE":
                raise RuntimeError(
                    f"{page_key}のセットリスト{row_number}行目がTRUEではありません。"
                )
            for field in ("song_id", "joucho_participation", "review_note"):
                if review[field] != "":
                    raise RuntimeError(
                        f"{page_key}のセットリスト{row_number}行目の{field}は"
                        "pilotでは空欄を想定しています。"
                    )
            if not raw["setlist_no_raw"].strip():
                raise RuntimeError(f"{page_key}の曲順rawに空欄があります。")
            if not raw["song_title_raw"].strip():
                raise RuntimeError(f"{page_key}の曲名rawに空欄があります。")
            if not raw["artist_credit_raw"].strip():
                raise RuntimeError(f"{page_key}のartist credit rawに空欄があります。")

            setlist_rows.append(
                {
                    "performance_key": page_key,
                    "sort_order": row_number * 100,
                    "entry_type": "song",
                    "setlist_no_raw": raw["setlist_no_raw"],
                    "song_id": None,
                    "song_title_raw": raw["song_title_raw"],
                    "artist_credit_raw": raw["artist_credit_raw"],
                    "note_raw": optional_text(raw["note_raw"]),
                    "marker_label": None,
                    "notes": None,
                    "joucho_participation": None,
                }
            )

    return event_group_rows, performance_rows, source_rows, setlist_rows


def validate_ready_rows(
    event_groups: list[dict[str, object]],
    performances: list[dict[str, object]],
    sources: list[dict[str, object]],
    setlists: list[dict[str, object]],
) -> None:
    if len(event_groups) != 1:
        raise RuntimeError("event groupは1件である必要があります。")
    group_key = event_groups[0]["local_group_key"]
    if event_groups[0]["title"] != "ヰ世界情緒 2DAYS LIVE":
        raise RuntimeError("2DAYS event group名が想定と異なります。")

    performance_keys = [str(row["performance_key"]) for row in performances]
    if len(performances) != 3 or len(set(performance_keys)) != 3:
        raise RuntimeError("performance_keyは3件すべてuniqueである必要があります。")
    if tuple(performance_keys) != PILOT_PAGE_KEYS:
        raise RuntimeError("performancesの対象または順序が想定と異なります。")

    for row in performances:
        page_key = str(row["performance_key"])
        if not str(row["title"]).strip():
            raise RuntimeError(f"{page_key}のtitleが空です。")
        parsed_date = date.fromisoformat(str(row["performance_date"]))
        if parsed_date.isoformat() != row["performance_date"]:
            raise RuntimeError(f"{page_key}のperformance_dateがISO dateではありません。")
        if row["published_at"] is not None or row["is_listed"] is not False:
            raise RuntimeError(f"{page_key}の初期公開状態が想定と異なります。")
        if row["streaming_platforms"] != "[]":
            raise RuntimeError(f"{page_key}のstreaming_platformsが空配列ではありません。")
        approved = PILOT_PERFORMANCES[page_key]
        if row["local_group_key"] != approved["local_group_key"]:
            raise RuntimeError(f"{page_key}のevent groupが想定と異なります。")
        if row["group_sort_order"] != approved["group_sort_order"]:
            raise RuntimeError(f"{page_key}のgroup_sort_orderが想定と異なります。")

    if performances[0]["local_group_key"] is not None:
        raise RuntimeError("2021.1023にはevent groupを設定しません。")
    if performances[1]["local_group_key"] != group_key:
        raise RuntimeError("2026.0501のevent groupが一致しません。")
    if performances[2]["local_group_key"] != group_key:
        raise RuntimeError("2026.0502のevent groupが一致しません。")

    source_counts = Counter(str(row["performance_key"]) for row in sources)
    if len(sources) != 3 or source_counts != Counter(PILOT_PAGE_KEYS):
        raise RuntimeError("各performanceにはsourceが1件必要です。")
    for row in sources:
        if not str(row["source_url"]).strip():
            raise RuntimeError("source_urlが空です。")
        if row["source_key"] != row["performance_key"]:
            raise RuntimeError("source_keyがpilot page_keyと一致しません。")
        if row["source_type"] != "thinkr_wiki":
            raise RuntimeError("source_typeがthinkr_wikiではありません。")

    setlist_counts = Counter(str(row["performance_key"]) for row in setlists)
    if len(setlists) != 65 or setlist_counts != Counter(SETLIST_COUNTS):
        raise RuntimeError(f"セットリスト件数が想定外です: {dict(setlist_counts)}")
    for page_key in PILOT_PAGE_KEYS:
        rows = [row for row in setlists if row["performance_key"] == page_key]
        expected_orders = [index * 100 for index in range(1, len(rows) + 1)]
        orders = [row["sort_order"] for row in rows]
        if orders != expected_orders or len(set(orders)) != len(orders):
            raise RuntimeError(f"{page_key}のsort_orderが想定と異なります。")
    if any(row["entry_type"] != "song" for row in setlists):
        raise RuntimeError("pilot setlistにはsong以外を含めません。")
    if any(row["song_id"] is not None for row in setlists):
        raise RuntimeError("pilotのsong_idは全件NULLです。")
    if any(row["joucho_participation"] is not None for row in setlists):
        raise RuntimeError("pilotのjoucho_participationは全件NULLです。")
    if any(row["marker_label"] is not None for row in setlists):
        raise RuntimeError("pilotにはmarkerを含めません。")


def csv_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def write_csv_file(
    path: Path, fields: list[str], rows: list[dict[str, object]]
) -> None:
    with path.open("x", encoding="utf-8-sig", newline="") as output_file:
        writer = csv.DictWriter(
            output_file, fieldnames=fields, lineterminator="\n", extrasaction="raise"
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({field: csv_cell(row[field]) for field in fields})


def validate_written_csv(path: Path, fields: list[str], expected_count: int) -> None:
    rows = read_csv(path, fields)
    if len(rows) != expected_count:
        raise RuntimeError(f"{path.name}の再読込件数が想定外です: {len(rows)}")


def build_manifest(input_hashes: dict[str, str], generated_at: str) -> str:
    lines = [
        f"# {PILOT_NAME}",
        "",
        f"- Generated at (JST): {generated_at}",
        "- Scope: THINKR live ready data; no database operation",
        "- Performances: 2021.1023, 2026.0501, 2026.0502",
        "- Event groups: 1",
        "- Performance rows: 3",
        "- Source rows: 3",
        "- Setlist entries: 65",
        "- Linked song_id values: 0",
        "- Markers: 0",
        "- Published performances: 0",
        "- Listed performances: 0",
        "",
        "## Relation",
        "",
        "- Group: ヰ世界情緒 2DAYS LIVE",
        "- 2026.0501: group sort order 1",
        "- 2026.0502: group sort order 2",
        "- 2021.1023: no event group",
        "",
        "## CSV value convention",
        "",
        "- Empty optional cells represent SQL NULL for the future importer.",
        "- streaming_platforms contains [] to represent an empty text array, not NULL.",
        "- performance_key is a local FK-resolution key and is not a public ID.",
        "- source_key identifies the source page even when its value matches performance_key.",
        "",
        "## Inputs",
        "",
    ]
    for path, digest in input_hashes.items():
        lines.append(f"- {path}: SHA-256 {digest}")
    lines.extend(["", "## Validation", "", "- PASS", ""])
    return "\n".join(lines)


def create_output() -> Path:
    if OUTPUT_PATH.exists():
        raise RuntimeError(f"既存ready出力を保護するため上書きしません: {OUTPUT_PATH}")

    input_paths = (EVENTS_RAW_PATH, SETLISTS_RAW_PATH, LIVE_REVIEW_PATH)
    input_hashes_before = {
        str(path.relative_to(REPOSITORY_ROOT)): sha256(path) for path in input_paths
    }
    event_groups, performances, sources, setlists = build_ready_rows()
    validate_ready_rows(event_groups, performances, sources, setlists)

    READY_ROOT.mkdir(parents=True, exist_ok=True)
    temporary_directory = Path(
        tempfile.mkdtemp(prefix=f".{PILOT_NAME}-", dir=READY_ROOT)
    )
    try:
        write_csv_file(
            temporary_directory / "event_groups.csv", EVENT_GROUP_FIELDS, event_groups
        )
        write_csv_file(
            temporary_directory / "performances.csv", PERFORMANCE_FIELDS, performances
        )
        write_csv_file(
            temporary_directory / "performance_sources.csv", SOURCE_FIELDS, sources
        )
        write_csv_file(
            temporary_directory / "setlist_entries.csv", SETLIST_FIELDS, setlists
        )

        validate_written_csv(
            temporary_directory / "event_groups.csv", EVENT_GROUP_FIELDS, 1
        )
        validate_written_csv(
            temporary_directory / "performances.csv", PERFORMANCE_FIELDS, 3
        )
        validate_written_csv(
            temporary_directory / "performance_sources.csv", SOURCE_FIELDS, 3
        )
        validate_written_csv(
            temporary_directory / "setlist_entries.csv", SETLIST_FIELDS, 65
        )

        input_hashes_after = {
            str(path.relative_to(REPOSITORY_ROOT)): sha256(path) for path in input_paths
        }
        if input_hashes_after != input_hashes_before:
            raise RuntimeError("生成中に入力ファイルが変更されたため停止しました。")

        generated_at = datetime.now(ZoneInfo("Asia/Tokyo")).isoformat(
            timespec="seconds"
        )
        manifest = build_manifest(input_hashes_before, generated_at)
        (temporary_directory / "manifest.md").write_text(
            manifest, encoding="utf-8", newline="\n"
        )

        if OUTPUT_PATH.exists():
            raise RuntimeError(
                f"既存ready出力を保護するため上書きしません: {OUTPUT_PATH}"
            )
        os.replace(temporary_directory, OUTPUT_PATH)
        return OUTPUT_PATH
    finally:
        if temporary_directory.exists():
            shutil.rmtree(temporary_directory)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        output_path = create_output()
    except (OSError, RuntimeError, ValueError) as error:
        print(f"ready生成を停止しました: {error}", file=sys.stderr)
        return 1

    print(f"出力先: {output_path}")
    print("event_groups.csv: 1件")
    print("performances.csv: 3件")
    print("performance_sources.csv: 3件")
    print("setlist_entries.csv: 65件")
    print("validation: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

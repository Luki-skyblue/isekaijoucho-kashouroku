"""Create reviewed ready data for new songs and live setlist resolution.

Inputs are the finalized local review workbooks only. The script performs no
network or database operations and refuses to overwrite an existing ready set.
"""

from __future__ import annotations

import csv
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


ROOT: Final = Path(__file__).resolve().parents[2]
WORKING: Final = ROOT / "private-data" / "imports" / "thinkr" / "working"
READY_ROOT: Final = ROOT / "private-data" / "imports" / "thinkr" / "ready"
READY_NAME: Final = "live-song-resolution-001"
OUTPUT: Final = READY_ROOT / READY_NAME
ARTIST_REVIEW: Final = WORKING / "live_setlist_artist_review.xlsx"
SONG_REVIEW: Final = WORKING / "live_setlist_song_review.xlsx"
REGISTRATION_REVIEW: Final = WORKING / "live_setlist_song_registration_review.xlsx"

GROUP_FIELDS: Final = ["local_group_key", "title", "title_kana", "sort_title", "notes"]
SONG_FIELDS: Final = [
    "local_song_key", "creation_kind", "existing_song_group_id", "local_group_key",
    "title", "title_kana", "sort_title", "song_type", "artist_credit",
    "first_date", "first_source", "verification_status", "verification_note",
    "first_status", "first_full_status", "tie_up_status", "album_text_status",
    "original_artist_status", "original_vocal_status", "original_lyricist_status",
    "original_composer_status", "original_arranger_status", "version_name",
    "version_type", "is_primary_version",
]
SETLIST_FIELDS: Final = [
    "setlist_entry_id", "expected_live_performance_id", "expected_sort_order",
    "expected_song_title_raw", "expected_artist_credit_raw", "expected_current_song_id",
    "expected_current_joucho_participation", "target_joucho_participation",
    "target_existing_song_id", "target_local_song_key", "notes_append", "resolution_status",
]


def header(sheet) -> list[str]:
    return [str(cell.value or "") for cell in sheet[1]]


def rows(sheet) -> list[dict[str, object]]:
    fields = header(sheet)
    return [dict(zip(fields, values, strict=True)) for values in sheet.iter_rows(min_row=2, values_only=True)]


def csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def write_csv(path: Path, fields: list[str], data: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in data:
            writer.writerow({field: csv_value(row.get(field)) for field in fields})


def load_inputs() -> tuple[
    list[dict[str, object]], list[dict[str, object]], list[dict[str, object]],
    list[dict[str, object]], list[dict[str, object]], list[dict[str, object]],
]:
    artist_wb = load_workbook(ARTIST_REVIEW, read_only=True, data_only=False)
    song_wb = load_workbook(SONG_REVIEW, read_only=True, data_only=False)
    registration_wb = load_workbook(REGISTRATION_REVIEW, read_only=True, data_only=False)
    try:
        return (
            rows(artist_wb["artist_participation"]),
            rows(artist_wb["_setlist_source"]),
            rows(song_wb["song_id_review"]),
            rows(registration_wb["new_versions"]),
            rows(registration_wb["new_songs"]),
            rows(registration_wb["songs_reference"]),
        )
    finally:
        artist_wb.close()
        song_wb.close()
        registration_wb.close()


def build_ready() -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    artist_rows, source_rows, resolutions, versions, new_songs, songs_reference = load_inputs()
    if len(source_rows) != 499 or len(resolutions) != 254 or len(versions) != 21 or len(new_songs) != 18:
        raise RuntimeError("review入力件数が最終想定と一致しません。")

    participation = {str(row["artist_credit_raw"]): str(row["decision"]) for row in artist_rows}
    if Counter(participation.values()) != Counter({"TRUE": 27, "FALSE": 34}):
        raise RuntimeError("artist participation decisionが27/34種類ではありません。")

    resolution_by_pair = {(str(row["song_title_raw"]), str(row["artist_credit_raw"])): row for row in resolutions}
    if len(resolution_by_pair) != 254:
        raise RuntimeError("resolutionのtitle×artist keyがuniqueではありません。")
    expected_pair_counts = Counter(str(row["resolution_status"]) for row in resolutions)
    if expected_pair_counts != Counter({"LINK_EXISTING": 212, "NEW_VERSION": 21, "NEW_SONG": 18, "SPECIAL_UNRESOLVED": 3}):
        raise RuntimeError(f"最終resolution分類が不正です: {dict(expected_pair_counts)}")

    primary_by_group: dict[int, list[dict[str, object]]] = defaultdict(list)
    for song in songs_reference:
        if song["is_primary_version"] is True:
            primary_by_group[int(song["song_group_id"])].append(song)

    version_keys: dict[tuple[str, str], str] = {}
    song_ready: list[dict[str, object]] = []
    for index, row in enumerate(sorted(versions, key=lambda r: (str(r["song_title_raw"]), str(r["artist_credit_raw"]))), start=1):
        if row["review_decision"] != "ACCEPT":
            raise RuntimeError("NEW_VERSIONに未承認行があります。")
        group_id = int(row["confirmed_song_group_id"])
        primaries = primary_by_group.get(group_id, [])
        if len(primaries) != 1:
            raise RuntimeError(f"song_group #{group_id}のprimary songが一意ではありません。")
        primary = primaries[0]
        local_key = f"new-version-{index:03d}"
        pair = (str(row["song_title_raw"]), str(row["artist_credit_raw"]))
        version_keys[pair] = local_key
        song_ready.append({
            "local_song_key": local_key,
            "creation_kind": "NEW_VERSION",
            "existing_song_group_id": group_id,
            "local_group_key": None,
            "title": row["proposed_title"],
            "title_kana": primary["title_kana"],
            "sort_title": primary["sort_title"],
            "song_type": primary["song_type"],
            "artist_credit": row["proposed_artist_credit"],
            "first_date": None,
            "first_source": None,
            "verification_status": "confirmed",
            "verification_note": None,
            "first_status": "unverified",
            "first_full_status": "unverified",
            "tie_up_status": "unverified",
            "album_text_status": "unverified",
            "original_artist_status": "unverified",
            "original_vocal_status": "unverified",
            "original_lyricist_status": "unverified",
            "original_composer_status": "unverified",
            "original_arranger_status": "unverified",
            "version_name": row["confirmed_version_name"],
            "version_type": row["confirmed_version_type"],
            "is_primary_version": False,
        })

    group_ready: list[dict[str, object]] = []
    new_song_keys: dict[tuple[str, str], str] = {}
    for index, row in enumerate(sorted(new_songs, key=lambda r: (str(r["song_title_raw"]), str(r["artist_credit_raw"]))), start=1):
        if row["review_decision"] != "ACCEPT":
            raise RuntimeError("NEW_SONGに未承認行があります。")
        local_group = f"new-group-{index:03d}"
        local_song = f"new-song-{index:03d}"
        pair = (str(row["song_title_raw"]), str(row["artist_credit_raw"]))
        new_song_keys[pair] = local_song
        group_ready.append({
            "local_group_key": local_group,
            "title": row["proposed_title"],
            "title_kana": None,
            "sort_title": None,
            "notes": None,
        })
        song_ready.append({
            "local_song_key": local_song,
            "creation_kind": "NEW_SONG",
            "existing_song_group_id": None,
            "local_group_key": local_group,
            "title": row["proposed_title"],
            "title_kana": None,
            "sort_title": None,
            "song_type": None,
            "artist_credit": row["proposed_artist_credit"],
            "first_date": None,
            "first_source": None,
            "verification_status": "confirmed",
            "verification_note": None,
            "first_status": "unverified",
            "first_full_status": "unverified",
            "tie_up_status": "unverified",
            "album_text_status": "unverified",
            "original_artist_status": "unverified",
            "original_vocal_status": "unverified",
            "original_lyricist_status": "unverified",
            "original_composer_status": "unverified",
            "original_arranger_status": "unverified",
            "version_name": None,
            "version_type": "standard",
            "is_primary_version": True,
        })

    if len({row["local_song_key"] for row in song_ready}) != 39:
        raise RuntimeError("local_song_keyが39件uniqueではありません。")
    if len({row["local_group_key"] for row in group_ready}) != 18:
        raise RuntimeError("local_group_keyが18件uniqueではありません。")

    setlist_ready: list[dict[str, object]] = []
    for source in source_rows:
        artist = str(source["artist_credit_raw"] or "")
        decision = participation.get(artist)
        if decision not in {"TRUE", "FALSE"}:
            raise RuntimeError(f"artist decisionがありません: {artist}")
        pair = (str(source["song_title_raw"] or ""), artist)
        target_existing = None
        target_local = None
        annotation = None
        if decision == "FALSE":
            status = "NOT_PARTICIPATING"
        else:
            resolution = resolution_by_pair.get(pair)
            if not resolution:
                raise RuntimeError(f"TRUE行にsong resolutionがありません: {pair}")
            status = str(resolution["resolution_status"])
            annotation = resolution.get("annotation_raw")
            if status == "LINK_EXISTING":
                target_existing = int(resolution["song_id"])
            elif status == "NEW_VERSION":
                target_local = version_keys[pair]
            elif status == "NEW_SONG":
                target_local = new_song_keys[pair]
            elif status != "SPECIAL_UNRESOLVED":
                raise RuntimeError(f"未対応resolutionです: {status}")
        setlist_ready.append({
            "setlist_entry_id": int(source["setlist_entry_id"]),
            "expected_live_performance_id": int(source["live_performance_id"]),
            "expected_sort_order": int(source["sort_order"]),
            "expected_song_title_raw": source["song_title_raw"],
            "expected_artist_credit_raw": source["artist_credit_raw"],
            "expected_current_song_id": source["current_song_id"],
            "expected_current_joucho_participation": source["current_joucho_participation"],
            "target_joucho_participation": decision == "TRUE",
            "target_existing_song_id": target_existing,
            "target_local_song_key": target_local,
            "notes_append": annotation,
            "resolution_status": status,
        })

    stats = {
        "ready_name": READY_NAME,
        "new_song_groups": len(group_ready),
        "new_songs": len(song_ready),
        "new_versions": sum(row["creation_kind"] == "NEW_VERSION" for row in song_ready),
        "new_primary_songs": sum(row["creation_kind"] == "NEW_SONG" for row in song_ready),
        "setlist_updates": len(setlist_ready),
        "participation_true": sum(row["target_joucho_participation"] is True for row in setlist_ready),
        "participation_false": sum(row["target_joucho_participation"] is False for row in setlist_ready),
        "song_id_existing_rows": sum(row["target_existing_song_id"] is not None for row in setlist_ready),
        "song_id_local_rows": sum(row["target_local_song_key"] is not None for row in setlist_ready),
        "song_id_null_rows": sum(row["target_existing_song_id"] is None and row["target_local_song_key"] is None for row in setlist_ready),
        "special_unresolved_rows": sum(row["resolution_status"] == "SPECIAL_UNRESOLVED" for row in setlist_ready),
        "annotation_rows": sum(row["notes_append"] is not None for row in setlist_ready),
    }
    expected_stats = {
        "new_song_groups": 18, "new_songs": 39, "new_versions": 21,
        "new_primary_songs": 18, "setlist_updates": 499,
        "participation_true": 378, "participation_false": 121,
        "song_id_existing_rows": 336, "song_id_local_rows": 39,
        "song_id_null_rows": 124, "special_unresolved_rows": 3,
        "annotation_rows": 7,
    }
    for name, expected in expected_stats.items():
        if stats[name] != expected:
            raise RuntimeError(f"{name}が想定外です: {stats[name]} != {expected}")
    return group_ready, song_ready, setlist_ready, stats


def run() -> dict[str, object]:
    if OUTPUT.exists():
        raise RuntimeError(f"既存readyを上書きしません: {OUTPUT}")
    groups, songs, setlists, stats = build_ready()
    OUTPUT.mkdir(parents=True)
    write_csv(OUTPUT / "song_groups.csv", GROUP_FIELDS, groups)
    write_csv(OUTPUT / "songs.csv", SONG_FIELDS, songs)
    write_csv(OUTPUT / "live_setlist_updates.csv", SETLIST_FIELDS, setlists)
    generated_at = datetime.now(ZoneInfo("Asia/Tokyo")).isoformat(timespec="seconds")
    manifest = {**stats, "generated_at": generated_at, "database_writes": 0}
    (OUTPUT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = ["# Live song resolution ready", "", f"- ready: {READY_NAME}", f"- generated_at: {generated_at}"]
    lines.extend(f"- {key}: {value}" for key, value in stats.items() if key != "ready_name")
    lines.append("- database_writes: 0")
    (OUTPUT / "manifest.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return stats


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        stats = run()
        print(json.dumps(stats, ensure_ascii=False, indent=2))
        print(f"ready created: {OUTPUT}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError, KeyError) as error:
        print(f"ready生成を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

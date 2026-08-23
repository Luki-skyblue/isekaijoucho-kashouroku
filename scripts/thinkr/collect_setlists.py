"""Politely collect structured setlists from THINKR Wiki event pages.

The collector is deliberately sequential. It stores only structured setlist
cells, never full page HTML, and persists each result so interrupted runs can
resume without re-fetching completed pages.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import random
import re
import sqlite3
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Final, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, unquote, urljoin, urlparse
from urllib.request import Request, urlopen


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
DATA_ROOT: Final = REPOSITORY_ROOT / "private-data" / "imports" / "thinkr"
EVENTS_PATH: Final = DATA_ROOT / "raw" / "events_raw.csv"
EVENTS_REVIEW_XLSX_PATH: Final = DATA_ROOT / "working" / "events_review.xlsx"
SETLISTS_PATH: Final = DATA_ROOT / "raw" / "setlists_raw.csv"
STATE_DIR: Final = DATA_ROOT / "working" / "setlists"
DATABASE_PATH: Final = STATE_DIR / "state.sqlite3"
PAGES_PATH: Final = STATE_DIR / "pages.csv"
FETCH_LOG_PATH: Final = STATE_DIR / "fetch_log.csv"
LOCK_PATH: Final = STATE_DIR / "collector.lock"

TOP_URL: Final = "https://wikiwiki.jp/thinkr/"
USER_AGENT: Final = (
    "Mozilla/5.0 (compatible; KashourokuSetlistCollector/1.0; "
    "sequential archival research)"
)
MINIMUM_DELAY_SECONDS: Final = 60.0
DEFAULT_DELAY_SECONDS: Final = 60.0
DEFAULT_JITTER_SECONDS: Final = 2.0
MAX_RESPONSE_BYTES: Final = 2_000_000
PAGE_KEY_PATTERN: Final = re.compile(r"^\d{4}\.\d{4}[A-Za-z]*$")
EXPECTED_EVENT_FIELDS: Final = [
    "source_category",
    "year",
    "page_key",
    "event_date_raw",
    "event_title_raw",
    "source_url",
    "import_target",
]
EXPECTED_REVIEW_FIELDS: Final = [*EXPECTED_EVENT_FIELDS, "review_note"]
SUPPORTED_REVIEW_TARGETS: Final = ("TRUE", "?")
EVENT_CATEGORIES: Final = ("箱内", "歌枠・配信", "外部")
SETLIST_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "setlist_table_index",
    "setlist_section_raw",
    "setlist_table_caption_raw",
    "setlist_table_headers_raw",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "setlist_row_cells_raw",
    "source_url",
]
PAGE_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "status",
    "setlist_count",
    "parse_warning_count",
    "http_status",
    "attempt_count",
    "last_attempt_at",
    "next_retry_at",
    "error_message",
    "fetched_at",
    "source_url",
]
LOG_FIELDS: Final = [
    "attempted_at",
    "completed_at",
    "page_key",
    "source_url",
    "outcome",
    "http_status",
    "retry_after_raw",
    "retry_after_parsed_at",
    "date_header_raw",
    "rate_limit_headers_json",
    "restriction_headers_json",
    "local_safety_until",
    "error_message",
    "next_retry_at",
    "next_retry_source",
    "manual_resume_required",
]

NO_HEADERS: Final = {"no", "no.", "番号", "曲順"}
SONG_HEADERS: Final = {"曲名", "楽曲", "楽曲名", "タイトル"}
ARTIST_HEADERS: Final = {"アーティスト", "歌唱", "歌手"}
NOTE_HEADERS: Final = {"備考", "注記", "note", "notes"}
BLOCK_STATUS_CODES: Final = {403, 429, 503}
LOCAL_RESTRICTION_FALLBACK_HOURS: Final = 24
RATE_LIMIT_HEADER_PREFIXES: Final = (
    "x-ratelimit-",
    "ratelimit-",
    "x-rate-limit-",
    "rate-limit-",
)
USEFUL_RESTRICTION_HEADERS: Final = {
    "retry-after",
    "date",
    "server",
    "via",
    "cache-control",
    "expires",
    "age",
    "location",
    "content-type",
    "content-length",
    "connection",
    "www-authenticate",
    "cf-ray",
    "cf-cache-status",
    "x-cache",
    "x-cache-hits",
    "x-served-by",
    "x-timer",
    "x-request-id",
    "request-id",
    "x-correlation-id",
    "x-amzn-requestid",
    "x-envoy-upstream-service-time",
}
BLOCK_MARKERS: Final = (
    "<title>just a moment...</title>",
    "too many requests",
    "access denied",
    "アクセスが制限されています",
    "一時的にアクセスを制限",
)


@dataclass(frozen=True)
class ResponseHeaderSnapshot:
    retry_after_raw: str = ""
    date_header_raw: str = ""
    rate_limit_headers_json: str = "[]"
    restriction_headers_json: str = "[]"


@dataclass(frozen=True)
class RetryDecision:
    retry_after_parsed_at: str
    local_safety_until: str
    next_retry_at: str
    next_retry_source: str


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_time(value: datetime | None = None) -> str:
    return (value or utc_now()).isoformat(timespec="seconds")


def parse_iso_time(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.fromisoformat(value)


def normalized_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source_file:
        for chunk in iter(lambda: source_file.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


class CollectorLock:
    """Use an OS-level lock so two collectors cannot fetch in parallel."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.file = None

    def __enter__(self) -> "CollectorLock":
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.file = self.path.open("a+b")
        self.file.seek(0)
        if self.file.read(1) == b"":
            self.file.seek(0)
            self.file.write(b"0")
            self.file.flush()
        self.file.seek(0)

        try:
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.file.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.file.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except OSError as error:
            self.file.close()
            self.file = None
            raise RuntimeError("別のセットリスト収集処理が実行中です。") from error
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self.file is None:
            return
        try:
            self.file.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.file.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.file.fileno(), fcntl.LOCK_UN)
        finally:
            self.file.close()
            self.file = None


@dataclass(frozen=True)
class Event:
    source_category: str
    year: str
    page_key: str
    event_title_raw: str
    event_date_raw: str
    source_url: str


@dataclass
class HtmlCell:
    tag: str
    attrs: dict[str, str]
    text_parts: list[str] = field(default_factory=list)

    @property
    def text(self) -> str:
        return "".join(self.text_parts).strip()

    def as_raw_dict(self) -> dict[str, str]:
        return {
            "tag": self.tag,
            "text": self.text,
            "colspan": self.attrs.get("colspan", ""),
            "rowspan": self.attrs.get("rowspan", ""),
        }


@dataclass
class HtmlTable:
    section: str
    caption_parts: list[str] = field(default_factory=list)
    rows: list[list[HtmlCell]] = field(default_factory=list)

    @property
    def caption(self) -> str:
        return "".join(self.caption_parts).strip()


@dataclass(frozen=True)
class ParsedSetlistRow:
    table_index: int
    section_raw: str
    caption_raw: str
    headers_raw: str
    setlist_no_raw: str
    song_title_raw: str
    artist_credit_raw: str
    note_raw: str
    row_cells_raw: str


class AvailabilityParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.existing: set[str] = set()
        self.missing: set[str] = set()

    def handle_starttag(
        self, tag: str, attrs_list: list[tuple[str, str | None]]
    ) -> None:
        if tag != "a":
            return
        attrs = {key: value or "" for key, value in attrs_list}
        href = attrs.get("href", "")
        absolute_url = urljoin(TOP_URL, href)
        parsed = urlparse(absolute_url)
        path = unquote(parsed.path)

        prefix = "/thinkr/"
        if path.startswith(prefix):
            candidate = path[len(prefix) :].strip("/")
            if PAGE_KEY_PATTERN.fullmatch(candidate):
                self.existing.add(candidate)

        if path == "/thinkr/::cmd/edit":
            candidates = parse_qs(parsed.query).get("page", [])
            if len(candidates) == 1 and PAGE_KEY_PATTERN.fullmatch(candidates[0]):
                self.missing.add(candidates[0])


class SetlistParser(HTMLParser):
    """Extract only tables inside the main content's setlist section."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[HtmlTable] = []
        self.found_setlist_heading = False
        self._content_depth = 0
        self._heading_level: int | None = None
        self._heading_parts: list[str] = []
        self._setlist_heading_level: int | None = None
        self._inside_setlist = False
        self._section = ""
        self._table_depth = 0
        self._table: HtmlTable | None = None
        self._row: list[HtmlCell] | None = None
        self._cell: HtmlCell | None = None
        self._inside_caption = False

    def handle_starttag(
        self, tag: str, attrs_list: list[tuple[str, str | None]]
    ) -> None:
        attrs = {key: value or "" for key, value in attrs_list}
        if tag == "div":
            if self._content_depth > 0:
                self._content_depth += 1
            elif attrs.get("id") == "content":
                self._content_depth = 1
            return

        if self._content_depth == 0:
            return

        if re.fullmatch(r"h[1-6]", tag):
            self._heading_level = int(tag[1])
            self._heading_parts = []
            return

        if tag == "table" and self._inside_setlist:
            self._table_depth += 1
            if self._table_depth == 1:
                self._table = HtmlTable(section=self._section)
            return

        if self._table_depth != 1 or self._table is None:
            return

        if tag == "caption":
            self._inside_caption = True
        elif tag == "tr":
            self._row = []
        elif tag in {"th", "td"} and self._row is not None:
            self._cell = HtmlCell(tag=tag, attrs=attrs)
        elif tag == "br" and self._cell is not None:
            self._cell.text_parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag == "div" and self._content_depth > 0:
            self._content_depth -= 1
            return

        if self._content_depth == 0:
            return

        if re.fullmatch(r"h[1-6]", tag) and self._heading_level is not None:
            heading = normalized_text("".join(self._heading_parts))
            level = self._heading_level
            if heading == "セットリスト":
                self.found_setlist_heading = True
                self._inside_setlist = True
                self._setlist_heading_level = level
                self._section = ""
            elif self._inside_setlist and self._setlist_heading_level is not None:
                if level <= self._setlist_heading_level:
                    self._inside_setlist = False
                    self._section = ""
                else:
                    self._section = heading
            self._heading_level = None
            self._heading_parts = []
            return

        if tag == "table" and self._table_depth > 0:
            if self._table_depth == 1 and self._table is not None:
                self.tables.append(self._table)
                self._table = None
                self._row = None
                self._cell = None
                self._inside_caption = False
            self._table_depth -= 1
            return

        if self._table_depth != 1 or self._table is None:
            return

        if tag == "caption":
            self._inside_caption = False
        elif tag in {"th", "td"} and self._row is not None and self._cell is not None:
            self._row.append(self._cell)
            self._cell = None
        elif tag == "tr" and self._row is not None:
            self._table.rows.append(self._row)
            self._row = None

    def handle_data(self, data: str) -> None:
        if self._content_depth == 0:
            return
        if self._heading_level is not None:
            self._heading_parts.append(data)
        if self._table_depth == 1 and self._table is not None:
            if self._cell is not None:
                self._cell.text_parts.append(data)
            elif self._inside_caption:
                self._table.caption_parts.append(data)


def read_events() -> list[Event]:
    with EVENTS_PATH.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if reader.fieldnames != EXPECTED_EVENT_FIELDS:
            raise RuntimeError("events_raw.csvの列構成が想定と異なります。")
        rows = list(reader)

    events = [
        Event(
            source_category=row["source_category"],
            year=row["year"],
            page_key=row["page_key"],
            event_title_raw=row["event_title_raw"],
            event_date_raw=row["event_date_raw"],
            source_url=row["source_url"],
        )
        for row in rows
    ]
    if len(events) != len({event.page_key for event in events}):
        raise RuntimeError("events_raw.csvに重複するpage_keyがあります。")
    return events


def read_review_target_page_keys(
    events: list[Event],
    review_target: str,
    workbook_path: Path | None = None,
) -> set[str]:
    """Read exact human selections without modifying or normalizing the workbook."""
    workbook_path = workbook_path or EVENTS_REVIEW_XLSX_PATH
    if review_target not in SUPPORTED_REVIEW_TARGETS:
        raise RuntimeError(
            f"--review-targetは{', '.join(SUPPORTED_REVIEW_TARGETS)}のみ指定できます。"
        )
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "--review-targetにはopenpyxlが必要ですが、利用できません。"
        ) from error

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    except (OSError, ValueError) as error:
        raise RuntimeError(
            f"events_review.xlsxを読み取り専用で開けません: {workbook_path}"
        ) from error

    try:
        if "events_review" not in workbook.sheetnames:
            raise RuntimeError("events_review.xlsxにevents_reviewシートがありません。")
        worksheet = workbook["events_review"]
        row_iterator = worksheet.iter_rows(values_only=True)
        header_row = next(row_iterator, None)
        if list(header_row or ()) != EXPECTED_REVIEW_FIELDS:
            raise RuntimeError("events_review.xlsxの列構成が想定と異なります。")

        page_key_index = EXPECTED_REVIEW_FIELDS.index("page_key")
        import_target_index = EXPECTED_REVIEW_FIELDS.index("import_target")
        reviewed_page_keys: set[str] = set()
        matched_page_keys: set[str] = set()
        allowed_cell_values = {None, "", "TRUE", "FALSE", "?"}
        for row_number, row in enumerate(row_iterator, start=2):
            page_key = row[page_key_index]
            import_target = row[import_target_index]
            if not isinstance(page_key, str):
                raise RuntimeError(
                    f"events_review.xlsxの{row_number}行目のpage_keyが文字列ではありません。"
                )
            if page_key in reviewed_page_keys:
                raise RuntimeError(
                    f"events_review.xlsxに重複するpage_keyがあります: {page_key}"
                )
            if import_target not in allowed_cell_values:
                raise RuntimeError(
                    f"events_review.xlsxの{row_number}行目に未対応のimport_targetがあります。"
                    "値は完全一致のTRUE / FALSE / ? / 空欄だけを使用してください。"
                )
            reviewed_page_keys.add(page_key)
            if import_target == review_target:
                matched_page_keys.add(page_key)
    finally:
        workbook.close()

    raw_page_keys = {event.page_key for event in events}
    missing = raw_page_keys - reviewed_page_keys
    unknown = reviewed_page_keys - raw_page_keys
    if missing or unknown:
        raise RuntimeError(
            "events_review.xlsxとevents_raw.csvのpage_keyが一致しません。"
            f" review側不足={len(missing)}, review側のみ={len(unknown)}"
        )
    return matched_page_keys


def connect_database(read_only: bool = False) -> sqlite3.Connection:
    if read_only:
        connection = sqlite3.connect(
            f"file:{DATABASE_PATH.resolve().as_posix()}?mode=ro", uri=True
        )
        connection.row_factory = sqlite3.Row
        return connection

    STATE_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    connection.executescript(
        """
        PRAGMA journal_mode = WAL;
        CREATE TABLE IF NOT EXISTS meta (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS availability (
          page_key TEXT PRIMARY KEY,
          available INTEGER NOT NULL,
          indexed_at TEXT NOT NULL,
          source_sha256 TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS pages (
          page_key TEXT PRIMARY KEY,
          event_title_raw TEXT NOT NULL,
          event_date_raw TEXT NOT NULL,
          source_url TEXT NOT NULL,
          status TEXT NOT NULL,
          setlist_count INTEGER NOT NULL DEFAULT 0,
          parse_warning_count INTEGER NOT NULL DEFAULT 0,
          http_status INTEGER,
          attempt_count INTEGER NOT NULL DEFAULT 0,
          last_attempt_at TEXT,
          next_retry_at TEXT,
          error_message TEXT,
          fetched_at TEXT
        );
        CREATE TABLE IF NOT EXISTS setlist_rows (
          page_key TEXT NOT NULL,
          row_order INTEGER NOT NULL,
          table_index INTEGER NOT NULL,
          section_raw TEXT NOT NULL,
          caption_raw TEXT NOT NULL,
          headers_raw TEXT NOT NULL,
          setlist_no_raw TEXT NOT NULL,
          song_title_raw TEXT NOT NULL,
          artist_credit_raw TEXT NOT NULL,
          note_raw TEXT NOT NULL,
          row_cells_raw TEXT NOT NULL,
          PRIMARY KEY (page_key, row_order)
        );
        CREATE TABLE IF NOT EXISTS attempts (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          attempted_at TEXT NOT NULL,
          completed_at TEXT,
          page_key TEXT NOT NULL,
          source_url TEXT NOT NULL,
          outcome TEXT NOT NULL,
          http_status INTEGER,
          retry_after_raw TEXT,
          retry_after_parsed_at TEXT,
          date_header_raw TEXT,
          rate_limit_headers_json TEXT NOT NULL DEFAULT '[]',
          restriction_headers_json TEXT NOT NULL DEFAULT '[]',
          local_safety_until TEXT,
          error_message TEXT,
          next_retry_at TEXT,
          next_retry_source TEXT,
          manual_resume_required INTEGER NOT NULL DEFAULT 0
        );
        """
    )
    migrate_database(connection)
    connection.commit()
    return connection


def ensure_column(
    connection: sqlite3.Connection, table: str, column: str, declaration: str
) -> None:
    columns = {
        row["name"] for row in connection.execute(f"PRAGMA table_info({table})")
    }
    if column not in columns:
        connection.execute(f"ALTER TABLE {table} ADD COLUMN {column} {declaration}")


def migrate_database(connection: sqlite3.Connection) -> None:
    """Add restriction logging without deleting or rewriting collected rows."""
    attempt_columns = {
        "retry_after_raw": "TEXT",
        "retry_after_parsed_at": "TEXT",
        "date_header_raw": "TEXT",
        "rate_limit_headers_json": "TEXT NOT NULL DEFAULT '[]'",
        "restriction_headers_json": "TEXT NOT NULL DEFAULT '[]'",
        "local_safety_until": "TEXT",
        "next_retry_source": "TEXT",
        "manual_resume_required": "INTEGER NOT NULL DEFAULT 0",
    }
    for column, declaration in attempt_columns.items():
        ensure_column(connection, "attempts", column, declaration)

    # Historical blocked rows did not retain Retry-After. Their existing deadline
    # can therefore only be identified as the collector's local 24-hour floor.
    connection.execute(
        "UPDATE attempts SET "
        "local_safety_until = COALESCE(local_safety_until, next_retry_at), "
        "next_retry_source = COALESCE(next_retry_source, 'local_safety_fallback'), "
        "manual_resume_required = 1 "
        "WHERE outcome = 'blocked' AND next_retry_at IS NOT NULL"
    )
    existing_blocked_until = get_meta(connection, "blocked_until")
    if existing_blocked_until and get_meta(connection, "blocked_until_source") is None:
        latest = connection.execute(
            "SELECT attempted_at, http_status FROM attempts "
            "WHERE outcome = 'blocked' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        set_meta(connection, "blocked_until_source", "local_safety_fallback")
        set_meta(connection, "blocked_local_safety_until", existing_blocked_until)
        set_meta(connection, "blocked_server_retry_at", "")
        set_meta(connection, "blocked_retry_after_raw", "")
        set_meta(connection, "blocked_date_header_raw", "")
        set_meta(connection, "blocked_rate_limit_headers_json", "[]")
        set_meta(connection, "blocked_restriction_headers_json", "[]")
        set_meta(connection, "restriction_resume_required", "1")
        if latest:
            set_meta(connection, "blocked_attempted_at", latest["attempted_at"])
            set_meta(
                connection,
                "blocked_http_status",
                "" if latest["http_status"] is None else str(latest["http_status"]),
            )
    set_meta(connection, "restriction_log_schema_version", "1")


def get_meta(connection: sqlite3.Connection, key: str) -> str | None:
    row = connection.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else None


def set_meta(connection: sqlite3.Connection, key: str, value: str) -> None:
    connection.execute(
        "INSERT INTO meta(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )


def validate_delay(delay_seconds: float, jitter_seconds: float) -> None:
    if delay_seconds < MINIMUM_DELAY_SECONDS:
        raise RuntimeError(
            f"待機時間は{MINIMUM_DELAY_SECONDS:g}秒未満にできません。"
        )
    if jitter_seconds < 0:
        raise RuntimeError("待機時間の揺らぎは0秒以上にしてください。")


def wait_for_request_slot(
    connection: sqlite3.Connection,
    delay_seconds: float,
    jitter_seconds: float,
) -> None:
    row = connection.execute(
        "SELECT attempted_at FROM attempts ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if row is None:
        return
    previous = parse_iso_time(row["attempted_at"])
    if previous is None:
        return

    target_delay = delay_seconds + random.uniform(0, jitter_seconds)
    elapsed = (utc_now() - previous).total_seconds()
    remaining = target_delay - elapsed
    if remaining > 0:
        print(f"次のリクエストまで {remaining:.1f} 秒待機します。")
        time.sleep(remaining)


def start_attempt(
    connection: sqlite3.Connection, page_key: str, source_url: str
) -> tuple[int, str]:
    attempted_at = iso_time()
    blocked_until = parse_iso_time(get_meta(connection, "blocked_until"))
    early_resume_remaining = int(
        get_meta(connection, "restriction_early_resume_remaining_requests") or "0"
    )
    if blocked_until and blocked_until > utc_now() and early_resume_remaining > 0:
        set_meta(
            connection,
            "restriction_early_resume_remaining_requests",
            str(early_resume_remaining - 1),
        )
        set_meta(connection, "restriction_early_resume_consumed_at", attempted_at)
        set_meta(connection, "restriction_early_resume_page_key", page_key)
    cursor = connection.execute(
        "INSERT INTO attempts(attempted_at, page_key, source_url, outcome) "
        "VALUES(?, ?, ?, 'started')",
        (attempted_at, page_key, source_url),
    )
    connection.commit()
    return int(cursor.lastrowid), attempted_at


def finish_attempt(
    connection: sqlite3.Connection,
    attempt_id: int,
    outcome: str,
    http_status: int | None = None,
    error_message: str | None = None,
    next_retry_at: str | None = None,
    headers: ResponseHeaderSnapshot | None = None,
    retry_decision: RetryDecision | None = None,
    manual_resume_required: bool = False,
) -> None:
    headers = headers or ResponseHeaderSnapshot()
    connection.execute(
        "UPDATE attempts SET completed_at = ?, outcome = ?, http_status = ?, "
        "retry_after_raw = ?, retry_after_parsed_at = ?, date_header_raw = ?, "
        "rate_limit_headers_json = ?, restriction_headers_json = ?, "
        "local_safety_until = ?, error_message = ?, next_retry_at = ?, "
        "next_retry_source = ?, manual_resume_required = ? WHERE id = ?",
        (
            iso_time(),
            outcome,
            http_status,
            headers.retry_after_raw,
            retry_decision.retry_after_parsed_at if retry_decision else "",
            headers.date_header_raw,
            headers.rate_limit_headers_json,
            headers.restriction_headers_json,
            retry_decision.local_safety_until if retry_decision else "",
            error_message,
            next_retry_at,
            retry_decision.next_retry_source if retry_decision else "",
            1 if manual_resume_required else 0,
            attempt_id,
        ),
    )
    connection.commit()


def serialize_header_pairs(pairs: list[tuple[str, str]]) -> str:
    return json.dumps(
        [{"name": name, "value": value} for name, value in pairs],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def capture_restriction_headers(headers: object | None) -> ResponseHeaderSnapshot:
    if headers is None or not hasattr(headers, "items"):
        return ResponseHeaderSnapshot()
    pairs = [(str(name), str(value)) for name, value in headers.items()]
    rate_limit_pairs = [
        (name, value)
        for name, value in pairs
        if name.lower().startswith(RATE_LIMIT_HEADER_PREFIXES)
    ]
    useful_pairs = [
        (name, value)
        for name, value in pairs
        if name.lower() in USEFUL_RESTRICTION_HEADERS
        or name.lower().startswith(RATE_LIMIT_HEADER_PREFIXES)
    ]
    lowered = [(name.lower(), value) for name, value in pairs]
    return ResponseHeaderSnapshot(
        retry_after_raw=next(
            (value for name, value in lowered if name == "retry-after"), ""
        ),
        date_header_raw=next((value for name, value in lowered if name == "date"), ""),
        rate_limit_headers_json=serialize_header_pairs(rate_limit_pairs),
        restriction_headers_json=serialize_header_pairs(useful_pairs),
    )


def request_html(url: str) -> tuple[str, int, str, ResponseHeaderSnapshot]:
    request = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "ja,en;q=0.5",
        },
    )
    with urlopen(request, timeout=30) as response:
        content_type = response.headers.get_content_type()
        if content_type != "text/html":
            raise RuntimeError(f"想定外のContent-Typeです: {content_type}")
        body = response.read(MAX_RESPONSE_BYTES + 1)
        if len(body) > MAX_RESPONSE_BYTES:
            raise RuntimeError("HTMLが上限サイズを超えたため停止します。")
        charset = response.headers.get_content_charset() or "utf-8"
        html = body.decode(charset)
        return (
            html,
            int(response.status),
            response.geturl(),
            capture_restriction_headers(response.headers),
        )


def contains_block_signal(html: str) -> bool:
    lowered = html.lower()
    return any(marker in lowered for marker in BLOCK_MARKERS)


def parse_retry_after(value: str | None, now: datetime) -> datetime | None:
    if not value:
        return None
    try:
        seconds = int(value)
        if seconds < 0:
            return None
        return now + timedelta(seconds=seconds)
    except (ValueError, OverflowError):
        try:
            parsed = parsedate_to_datetime(value)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(timezone.utc)
        except (TypeError, ValueError, OverflowError):
            return None


def retry_decision_from_header(
    value: str | None,
    fallback_hours: int = LOCAL_RESTRICTION_FALLBACK_HOURS,
    now: datetime | None = None,
) -> RetryDecision:
    calculated_at = now or utc_now()
    local_safety_until = calculated_at + timedelta(hours=fallback_hours)
    server_retry_at = parse_retry_after(value, calculated_at)
    if server_retry_at and server_retry_at > local_safety_until:
        next_retry_at = server_retry_at
        source = "server_retry_after"
    else:
        next_retry_at = local_safety_until
        source = "local_safety_fallback"
    return RetryDecision(
        retry_after_parsed_at=iso_time(server_retry_at) if server_retry_at else "",
        local_safety_until=iso_time(local_safety_until),
        next_retry_at=iso_time(next_retry_at),
        next_retry_source=source,
    )


def retry_time_from_header(value: str | None, fallback_hours: int) -> datetime:
    """Compatibility wrapper retaining seconds and HTTP-date support."""
    decision = retry_decision_from_header(value, fallback_hours)
    parsed = parse_iso_time(decision.next_retry_at)
    assert parsed is not None
    return parsed


def ensure_not_blocked(connection: sqlite3.Connection) -> None:
    blocked_until = parse_iso_time(get_meta(connection, "blocked_until"))
    if get_meta(connection, "restriction_resume_required") == "1":
        deadline = get_meta(connection, "blocked_until") or "不明"
        source = get_meta(connection, "blocked_until_source") or "不明"
        raise RuntimeError(
            "アクセス制限後の明示的な再開操作が必要です。"
            f"記録期限={deadline}, 期限種別={source}。"
        )
    early_resume_remaining = int(
        get_meta(connection, "restriction_early_resume_remaining_requests") or "0"
    )
    if blocked_until and blocked_until > utc_now() and early_resume_remaining < 1:
        raise RuntimeError(
            f"アクセス制限の兆候を記録済みです。{iso_time(blocked_until)}まで実行しません。"
        )


def mark_blocked(
    connection: sqlite3.Connection,
    http_status: int,
    attempted_at: str,
    headers: ResponseHeaderSnapshot,
    decision: RetryDecision,
) -> None:
    set_meta(connection, "blocked_until", decision.next_retry_at)
    set_meta(connection, "blocked_until_source", decision.next_retry_source)
    set_meta(connection, "blocked_server_retry_at", decision.retry_after_parsed_at)
    set_meta(connection, "blocked_local_safety_until", decision.local_safety_until)
    set_meta(connection, "blocked_http_status", str(http_status))
    set_meta(connection, "blocked_attempted_at", attempted_at)
    set_meta(connection, "blocked_retry_after_raw", headers.retry_after_raw)
    set_meta(connection, "blocked_date_header_raw", headers.date_header_raw)
    set_meta(
        connection, "blocked_rate_limit_headers_json", headers.rate_limit_headers_json
    )
    set_meta(
        connection,
        "blocked_restriction_headers_json",
        headers.restriction_headers_json,
    )
    set_meta(connection, "restriction_resume_required", "1")
    set_meta(connection, "restriction_early_resume", "0")
    set_meta(connection, "restriction_early_resume_remaining_requests", "0")
    connection.commit()


def prepare_index(
    connection: sqlite3.Connection,
    events: list[Event],
    delay_seconds: float,
    jitter_seconds: float,
    refresh: bool,
) -> None:
    source_sha = file_sha256(EVENTS_PATH)
    cached_sha = get_meta(connection, "availability_source_sha256")
    cached_count_row = connection.execute(
        "SELECT COUNT(*) AS count FROM availability"
    ).fetchone()
    cached_count = int(cached_count_row["count"])
    if not refresh and cached_sha == source_sha and cached_count == len(events):
        print("ページ存在情報は取得済みです。トップページを再取得しません。")
        return

    ensure_not_blocked(connection)
    wait_for_request_slot(connection, delay_seconds, jitter_seconds)
    attempt_id, attempted_at = start_attempt(connection, "__index__", TOP_URL)
    try:
        html, status, final_url, response_headers = request_html(TOP_URL)
    except HTTPError as error:
        blocked = error.code in BLOCK_STATUS_CODES
        headers = capture_restriction_headers(error.headers) if blocked else None
        decision = (
            retry_decision_from_header(headers.retry_after_raw) if headers else None
        )
        next_retry = (
            decision.next_retry_at
            if decision
            else iso_time(retry_time_from_header(None, 6))
        )
        finish_attempt(
            connection,
            attempt_id,
            "blocked" if blocked else "http_error",
            error.code,
            str(error),
            next_retry,
            headers=headers,
            retry_decision=decision,
            manual_resume_required=blocked,
        )
        if blocked and headers and decision:
            mark_blocked(connection, error.code, attempted_at, headers, decision)
        raise RuntimeError(f"トップページ取得時にHTTP {error.code}が返りました。") from error
    except (URLError, TimeoutError, OSError, UnicodeError, RuntimeError) as error:
        next_retry = utc_now() + timedelta(hours=6)
        finish_attempt(
            connection,
            attempt_id,
            "network_error",
            error_message=str(error),
            next_retry_at=iso_time(next_retry),
        )
        raise RuntimeError("トップページの取得に失敗したため停止します。") from error

    if status in BLOCK_STATUS_CODES or contains_block_signal(html):
        decision = retry_decision_from_header(response_headers.retry_after_raw)
        finish_attempt(
            connection,
            attempt_id,
            "blocked",
            status,
            "アクセス制限を示す応答を検出しました。",
            decision.next_retry_at,
            headers=response_headers,
            retry_decision=decision,
            manual_resume_required=True,
        )
        mark_blocked(connection, status, attempted_at, response_headers, decision)
        raise RuntimeError("アクセス制限の兆候を検出したため停止します。")
    if urlparse(final_url).query or "/::" in unquote(urlparse(final_url).path):
        raise RuntimeError("トップページが管理用URLへ転送されたため停止します。")

    availability_parser = AvailabilityParser()
    availability_parser.feed(html)
    event_keys = {event.page_key for event in events}
    ambiguous = event_keys & availability_parser.existing & availability_parser.missing
    unknown = event_keys - availability_parser.existing - availability_parser.missing
    if ambiguous or unknown:
        message = (
            f"ページ存在判定に失敗しました: ambiguous={len(ambiguous)}, "
            f"unknown={len(unknown)}"
        )
        finish_attempt(connection, attempt_id, "parse_error", status, message)
        raise RuntimeError(message)

    indexed_at = iso_time()
    with connection:
        connection.execute("DELETE FROM availability")
        connection.executemany(
            "INSERT INTO availability(page_key, available, indexed_at, source_sha256) "
            "VALUES(?, ?, ?, ?)",
            [
                (
                    event.page_key,
                    1 if event.page_key in availability_parser.existing else 0,
                    indexed_at,
                    source_sha,
                )
                for event in events
            ],
        )
        for event in events:
            if event.page_key in availability_parser.missing:
                connection.execute(
                    "INSERT INTO pages(page_key, event_title_raw, event_date_raw, "
                    "source_url, status) VALUES(?, ?, ?, ?, 'missing') "
                    "ON CONFLICT(page_key) DO UPDATE SET "
                    "event_title_raw = excluded.event_title_raw, "
                    "event_date_raw = excluded.event_date_raw, "
                    "source_url = excluded.source_url, "
                    "status = CASE WHEN pages.status = 'completed' "
                    "THEN pages.status ELSE 'missing' END",
                    (
                        event.page_key,
                        event.event_title_raw,
                        event.event_date_raw,
                        event.source_url,
                    ),
                )
        set_meta(connection, "availability_source_sha256", source_sha)
        set_meta(connection, "availability_indexed_at", indexed_at)

    finish_attempt(connection, attempt_id, "completed", status)
    export_outputs(connection)
    print(
        f"ページ存在情報: 作成済み={len(availability_parser.existing & event_keys)}, "
        f"未作成={len(availability_parser.missing & event_keys)}"
    )


def require_index(connection: sqlite3.Connection, events: list[Event]) -> None:
    source_sha = file_sha256(EVENTS_PATH)
    cached_sha = get_meta(connection, "availability_source_sha256")
    count_row = connection.execute("SELECT COUNT(*) AS count FROM availability").fetchone()
    if cached_sha != source_sha or int(count_row["count"]) != len(events):
        raise RuntimeError(
            "ページ存在情報がありません。先に prepare-index を実行してください。"
        )


def header_index(headers: list[str], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if normalized_text(header).lower() in candidates:
            return index
    return None


def parse_setlists(html: str) -> tuple[list[ParsedSetlistRow], int, bool]:
    parser = SetlistParser()
    parser.feed(html)
    parsed_rows: list[ParsedSetlistRow] = []
    warning_count = 0

    for table_index, table in enumerate(parser.tables, start=1):
        nonempty_rows = [row for row in table.rows if any(cell.text for cell in row)]
        if not nonempty_rows:
            continue

        header_row_index = next(
            (
                index
                for index, row in enumerate(nonempty_rows)
                if any(normalized_text(cell.text) in SONG_HEADERS for cell in row)
            ),
            0,
        )
        headers = [cell.text for cell in nonempty_rows[header_row_index]]
        headers_json = json.dumps(headers, ensure_ascii=False)
        no_index = header_index(headers, NO_HEADERS)
        song_index = header_index(headers, SONG_HEADERS)
        artist_index = header_index(headers, ARTIST_HEADERS)
        note_index = header_index(headers, NOTE_HEADERS)
        current_section = table.section

        for row in nonempty_rows[header_row_index + 1 :]:
            if len(row) == 1 and row[0].attrs.get("colspan", "") not in {"", "1"}:
                current_section = row[0].text
                continue
            values = [cell.text for cell in row]
            if not any(values):
                continue

            def value_at(index: int | None) -> str:
                return values[index] if index is not None and index < len(values) else ""

            if song_index is None:
                warning_count += 1
            parsed_rows.append(
                ParsedSetlistRow(
                    table_index=table_index,
                    section_raw=current_section,
                    caption_raw=table.caption,
                    headers_raw=headers_json,
                    setlist_no_raw=value_at(no_index),
                    song_title_raw=value_at(song_index),
                    artist_credit_raw=value_at(artist_index),
                    note_raw=value_at(note_index),
                    row_cells_raw=json.dumps(
                        [cell.as_raw_dict() for cell in row], ensure_ascii=False
                    ),
                )
            )

    return parsed_rows, warning_count, parser.found_setlist_heading


def save_page_success(
    connection: sqlite3.Connection,
    event: Event,
    rows: list[ParsedSetlistRow],
    warning_count: int,
    http_status: int,
    attempted_at: str,
) -> None:
    previous = connection.execute(
        "SELECT attempt_count FROM pages WHERE page_key = ?", (event.page_key,)
    ).fetchone()
    attempt_count = (int(previous["attempt_count"]) if previous else 0) + 1
    fetched_at = iso_time()
    with connection:
        connection.execute("DELETE FROM setlist_rows WHERE page_key = ?", (event.page_key,))
        connection.executemany(
            "INSERT INTO setlist_rows(page_key, row_order, table_index, section_raw, "
            "caption_raw, headers_raw, setlist_no_raw, song_title_raw, "
            "artist_credit_raw, note_raw, row_cells_raw) "
            "VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            [
                (
                    event.page_key,
                    row_order,
                    row.table_index,
                    row.section_raw,
                    row.caption_raw,
                    row.headers_raw,
                    row.setlist_no_raw,
                    row.song_title_raw,
                    row.artist_credit_raw,
                    row.note_raw,
                    row.row_cells_raw,
                )
                for row_order, row in enumerate(rows, start=1)
            ],
        )
        connection.execute(
            "INSERT INTO pages(page_key, event_title_raw, event_date_raw, source_url, "
            "status, setlist_count, parse_warning_count, http_status, attempt_count, "
            "last_attempt_at, next_retry_at, error_message, fetched_at) "
            "VALUES(?, ?, ?, ?, 'completed', ?, ?, ?, ?, ?, NULL, NULL, ?) "
            "ON CONFLICT(page_key) DO UPDATE SET "
            "event_title_raw = excluded.event_title_raw, "
            "event_date_raw = excluded.event_date_raw, source_url = excluded.source_url, "
            "status = 'completed', setlist_count = excluded.setlist_count, "
            "parse_warning_count = excluded.parse_warning_count, "
            "http_status = excluded.http_status, attempt_count = excluded.attempt_count, "
            "last_attempt_at = excluded.last_attempt_at, next_retry_at = NULL, "
            "error_message = NULL, fetched_at = excluded.fetched_at",
            (
                event.page_key,
                event.event_title_raw,
                event.event_date_raw,
                event.source_url,
                len(rows),
                warning_count,
                http_status,
                attempt_count,
                attempted_at,
                fetched_at,
            ),
        )


def save_page_error(
    connection: sqlite3.Connection,
    event: Event,
    status: str,
    attempted_at: str,
    next_retry_at: str,
    error_message: str,
    http_status: int | None,
) -> None:
    previous = connection.execute(
        "SELECT attempt_count FROM pages WHERE page_key = ?", (event.page_key,)
    ).fetchone()
    attempt_count = (int(previous["attempt_count"]) if previous else 0) + 1
    with connection:
        connection.execute(
            "INSERT INTO pages(page_key, event_title_raw, event_date_raw, source_url, "
            "status, http_status, attempt_count, last_attempt_at, next_retry_at, "
            "error_message) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(page_key) DO UPDATE SET "
            "status = excluded.status, http_status = excluded.http_status, "
            "attempt_count = excluded.attempt_count, "
            "last_attempt_at = excluded.last_attempt_at, "
            "next_retry_at = excluded.next_retry_at, "
            "error_message = excluded.error_message",
            (
                event.page_key,
                event.event_title_raw,
                event.event_date_raw,
                event.source_url,
                status,
                http_status,
                attempt_count,
                attempted_at,
                next_retry_at,
                error_message,
            ),
        )


def handle_page(
    connection: sqlite3.Connection,
    event: Event,
    delay_seconds: float,
    jitter_seconds: float,
) -> bool:
    ensure_not_blocked(connection)
    wait_for_request_slot(connection, delay_seconds, jitter_seconds)
    attempt_id, attempted_at = start_attempt(connection, event.page_key, event.source_url)
    print(f"取得: {event.page_key} {event.event_title_raw}")

    try:
        html, http_status, final_url, response_headers = request_html(event.source_url)
    except HTTPError as error:
        if error.code in {404, 410}:
            next_retry_at = ""
            save_page_error(
                connection,
                event,
                "missing",
                attempted_at,
                next_retry_at,
                f"HTTP {error.code}",
                error.code,
            )
            finish_attempt(connection, attempt_id, "missing", error.code, str(error))
            export_outputs(connection)
            return True

        blocked = error.code in BLOCK_STATUS_CODES
        headers = capture_restriction_headers(error.headers) if blocked else None
        decision = (
            retry_decision_from_header(headers.retry_after_raw) if headers else None
        )
        retry_at = (
            decision.next_retry_at
            if decision
            else iso_time(retry_time_from_header(None, 6))
        )
        message = f"HTTP {error.code}: {error.reason}"
        save_page_error(
            connection,
            event,
            "blocked" if blocked else "error",
            attempted_at,
            retry_at,
            message,
            error.code,
        )
        finish_attempt(
            connection,
            attempt_id,
            "blocked" if blocked else "http_error",
            error.code,
            message,
            retry_at,
            headers=headers,
            retry_decision=decision,
            manual_resume_required=blocked,
        )
        if blocked and headers and decision:
            mark_blocked(connection, error.code, attempted_at, headers, decision)
        export_outputs(connection)
        raise RuntimeError(f"{event.page_key}でHTTPエラーを検出したため停止します。") from error
    except (URLError, TimeoutError, OSError, UnicodeError, RuntimeError) as error:
        retry_at = utc_now() + timedelta(hours=6)
        message = str(error)
        save_page_error(
            connection,
            event,
            "error",
            attempted_at,
            iso_time(retry_at),
            message,
            None,
        )
        finish_attempt(
            connection,
            attempt_id,
            "network_error",
            error_message=message,
            next_retry_at=iso_time(retry_at),
        )
        export_outputs(connection)
        raise RuntimeError(f"{event.page_key}の取得エラーを記録し、停止します。") from error

    if http_status in BLOCK_STATUS_CODES or contains_block_signal(html):
        decision = retry_decision_from_header(response_headers.retry_after_raw)
        message = "アクセス制限を示す応答を検出しました。"
        save_page_error(
            connection,
            event,
            "blocked",
            attempted_at,
            decision.next_retry_at,
            message,
            http_status,
        )
        finish_attempt(
            connection,
            attempt_id,
            "blocked",
            http_status,
            message,
            decision.next_retry_at,
            headers=response_headers,
            retry_decision=decision,
            manual_resume_required=True,
        )
        mark_blocked(connection, http_status, attempted_at, response_headers, decision)
        export_outputs(connection)
        raise RuntimeError("アクセス制限の兆候を検出したため停止します。")

    final = urlparse(final_url)
    if final.query or "/::" in unquote(final.path):
        retry_at = utc_now() + timedelta(hours=24)
        message = f"想定外のURLへ転送されました: {final_url}"
        save_page_error(
            connection,
            event,
            "error",
            attempted_at,
            iso_time(retry_at),
            message,
            http_status,
        )
        finish_attempt(
            connection,
            attempt_id,
            "redirect_error",
            http_status,
            message,
            iso_time(retry_at),
        )
        export_outputs(connection)
        raise RuntimeError(message)

    rows, warning_count, found_heading = parse_setlists(html)
    if '<div id="content">' not in html:
        retry_at = utc_now() + timedelta(hours=24)
        message = "本文領域を確認できず、HTML構造変更の可能性があります。"
        save_page_error(
            connection,
            event,
            "error",
            attempted_at,
            iso_time(retry_at),
            message,
            http_status,
        )
        finish_attempt(
            connection,
            attempt_id,
            "parse_error",
            http_status,
            message,
            iso_time(retry_at),
        )
        export_outputs(connection)
        raise RuntimeError(message)

    save_page_success(
        connection,
        event,
        rows,
        warning_count,
        http_status,
        attempted_at,
    )
    outcome = "completed" if rows else "completed_no_setlist"
    finish_attempt(connection, attempt_id, outcome, http_status)
    export_outputs(connection)
    heading_note = "" if found_heading else "（セットリスト見出しなし）"
    print(f"完了: {event.page_key} セトリ{len(rows)}件{heading_note}")
    return True


def atomic_write_csv(path: Path, fields: list[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    with temporary_path.open("w", encoding="utf-8-sig", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    os.replace(temporary_path, path)


def export_outputs(connection: sqlite3.Connection) -> None:
    setlist_rows = connection.execute(
        "SELECT p.page_key, p.event_title_raw, p.event_date_raw, "
        "r.table_index, r.section_raw, r.caption_raw, r.headers_raw, "
        "r.setlist_no_raw, r.song_title_raw, r.artist_credit_raw, r.note_raw, "
        "r.row_cells_raw, p.source_url "
        "FROM setlist_rows r JOIN pages p ON p.page_key = r.page_key "
        "ORDER BY p.page_key, r.row_order"
    ).fetchall()
    atomic_write_csv(
        SETLISTS_PATH,
        SETLIST_FIELDS,
        (
            {
                "page_key": row["page_key"],
                "event_title_raw": row["event_title_raw"],
                "event_date_raw": row["event_date_raw"],
                "setlist_table_index": row["table_index"],
                "setlist_section_raw": row["section_raw"],
                "setlist_table_caption_raw": row["caption_raw"],
                "setlist_table_headers_raw": row["headers_raw"],
                "setlist_no_raw": row["setlist_no_raw"],
                "song_title_raw": row["song_title_raw"],
                "artist_credit_raw": row["artist_credit_raw"],
                "note_raw": row["note_raw"],
                "setlist_row_cells_raw": row["row_cells_raw"],
                "source_url": row["source_url"],
            }
            for row in setlist_rows
        ),
    )

    page_rows = connection.execute(
        "SELECT page_key, event_title_raw, event_date_raw, status, setlist_count, "
        "parse_warning_count, http_status, attempt_count, last_attempt_at, "
        "next_retry_at, error_message, fetched_at, source_url "
        "FROM pages ORDER BY page_key"
    ).fetchall()
    atomic_write_csv(PAGES_PATH, PAGE_FIELDS, (dict(row) for row in page_rows))

    export_fetch_log(connection)


def export_fetch_log(connection: sqlite3.Connection) -> None:
    """Export attempts only, allowing schema migration without touching raw rows."""

    log_rows = connection.execute(
        "SELECT attempted_at, completed_at, page_key, source_url, outcome, "
        "http_status, retry_after_raw, retry_after_parsed_at, date_header_raw, "
        "rate_limit_headers_json, restriction_headers_json, local_safety_until, "
        "error_message, next_retry_at, next_retry_source, manual_resume_required "
        "FROM attempts ORDER BY id"
    ).fetchall()
    atomic_write_csv(FETCH_LOG_PATH, LOG_FIELDS, (dict(row) for row in log_rows))


@dataclass(frozen=True)
class CollectionPlan:
    events: list[Event]
    skipped_completed: int
    skipped_missing: int
    skipped_cooldown: int
    skipped_excluded: int
    review_target: str | None
    review_match_count: int


def unique_values(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def interleave_categories(events: list[Event], categories: list[str]) -> list[Event]:
    queues = {
        category: [event for event in events if event.source_category == category]
        for category in categories
    }
    interleaved: list[Event] = []
    index = 0
    while any(index < len(queues[category]) for category in categories):
        for category in categories:
            if index < len(queues[category]):
                interleaved.append(queues[category][index])
        index += 1
    return interleaved


def build_collection_plan(
    connection: sqlite3.Connection,
    events: list[Event],
    limit: int,
    page_keys: list[str],
    excluded_page_keys: list[str],
    categories: list[str],
    review_target: str | None = None,
    review_workbook_path: Path | None = None,
) -> CollectionPlan:
    if limit < 1:
        raise RuntimeError("--limitは1以上にしてください。")

    require_index(connection, events)
    event_by_key = {event.page_key: event for event in events}
    selected_categories = unique_values(categories)
    selected_page_keys = unique_values(page_keys)
    excluded = set(unique_values(excluded_page_keys))
    unknown_keys = (set(selected_page_keys) | excluded) - set(event_by_key)
    if unknown_keys:
        raise RuntimeError(f"events_raw.csvにないpage_keyです: {sorted(unknown_keys)}")

    if selected_page_keys:
        candidates = [event_by_key[key] for key in selected_page_keys]
    else:
        candidates = events
    review_page_keys: set[str] | None = None
    if review_target is not None:
        review_page_keys = read_review_target_page_keys(
            events, review_target, review_workbook_path
        )
        candidates = [
            event for event in candidates if event.page_key in review_page_keys
        ]
    if selected_categories:
        candidates = [
            event for event in candidates if event.source_category in selected_categories
        ]
        if len(selected_categories) > 1:
            candidates = interleave_categories(candidates, selected_categories)

    planned: list[Event] = []
    skipped_completed = 0
    skipped_missing = 0
    skipped_cooldown = 0
    skipped_excluded = 0
    for event in candidates:
        if event.page_key in excluded:
            skipped_excluded += 1
            continue
        availability = connection.execute(
            "SELECT available FROM availability WHERE page_key = ?", (event.page_key,)
        ).fetchone()
        if availability is None or not bool(availability["available"]):
            skipped_missing += 1
            continue

        page = connection.execute(
            "SELECT status, next_retry_at FROM pages WHERE page_key = ?",
            (event.page_key,),
        ).fetchone()
        if page and page["status"] == "completed":
            skipped_completed += 1
            continue
        if page and page["status"] == "missing":
            skipped_missing += 1
            continue
        retry_at = parse_iso_time(page["next_retry_at"]) if page else None
        if retry_at and retry_at > utc_now():
            skipped_cooldown += 1
            continue

        planned.append(event)
        if len(planned) >= limit:
            break

    return CollectionPlan(
        events=planned,
        skipped_completed=skipped_completed,
        skipped_missing=skipped_missing,
        skipped_cooldown=skipped_cooldown,
        skipped_excluded=skipped_excluded,
        review_target=review_target,
        review_match_count=len(review_page_keys) if review_page_keys is not None else 0,
    )


def print_collection_plan(plan: CollectionPlan, dry_run: bool) -> None:
    mode = "dry-run（HTTPアクセスなし）" if dry_run else "実行予定"
    if plan.review_target is not None:
        print(
            f"レビュー対象フィルター={plan.review_target}（完全一致 "
            f"{plan.review_match_count}件）"
        )
    print(f"今回のアクセス予定: {len(plan.events)}ページ / {mode}")
    for index, event in enumerate(plan.events, start=1):
        print(
            f"  {index}. [{event.source_category} / {event.year}] "
            f"{event.page_key} {event.event_title_raw}"
        )
    print(
        f"取得済み除外={plan.skipped_completed}, 未作成除外={plan.skipped_missing}, "
        f"再試行期限内除外={plan.skipped_cooldown}, 明示除外={plan.skipped_excluded}"
    )


def run_collection(
    connection: sqlite3.Connection,
    events: list[Event],
    limit: int,
    page_keys: list[str],
    excluded_page_keys: list[str],
    categories: list[str],
    review_target: str | None,
    delay_seconds: float,
    jitter_seconds: float,
    dry_run: bool,
) -> None:
    plan = build_collection_plan(
        connection,
        events,
        limit,
        page_keys,
        excluded_page_keys,
        categories,
        review_target,
    )
    print_collection_plan(plan, dry_run)
    if dry_run:
        return

    ensure_not_blocked(connection)
    requested = 0
    for event in plan.events:
        handle_page(connection, event, delay_seconds, jitter_seconds)
        requested += 1

    print(
        f"今回の個別ページ取得={requested}, 取得済みスキップ={plan.skipped_completed}, "
        f"未作成スキップ={plan.skipped_missing}, "
        f"再試行待ちスキップ={plan.skipped_cooldown}"
    )


def print_status(connection: sqlite3.Connection, total_events: int) -> None:
    rows = connection.execute(
        "SELECT status, COUNT(*) AS count FROM pages GROUP BY status ORDER BY status"
    ).fetchall()
    counts = {row["status"]: int(row["count"]) for row in rows}
    completed = counts.get("completed", 0)
    missing = counts.get("missing", 0)
    error_count = counts.get("error", 0) + counts.get("blocked", 0)
    pending = total_events - completed - missing - error_count
    setlist_count_row = connection.execute(
        "SELECT COALESCE(SUM(setlist_count), 0) AS count FROM pages WHERE status = 'completed'"
    ).fetchone()
    zero_count_row = connection.execute(
        "SELECT COUNT(*) AS count FROM pages "
        "WHERE status = 'completed' AND setlist_count = 0"
    ).fetchone()
    print(f"入力イベント={total_events}")
    print(
        f"取得済み={completed}, セトリ0件={int(zero_count_row['count'])}, "
        f"未作成={missing}, エラー/制限={error_count}, 未処理={pending}"
    )
    print(f"セットリスト行={int(setlist_count_row['count'])}")
    print(f"ページ存在情報取得日時={get_meta(connection, 'availability_indexed_at') or '未取得'}")
    blocked_until = get_meta(connection, "blocked_until")
    if blocked_until:
        source = get_meta(connection, "blocked_until_source") or "未分類"
        print(f"アクセス停止期限={blocked_until}（{source}）")
        print(
            "直近の制限応答="
            f"HTTP {get_meta(connection, 'blocked_http_status') or '不明'}, "
            f"attempted_at={get_meta(connection, 'blocked_attempted_at') or '不明'}"
        )
        print(
            "Retry-After="
            f"{get_meta(connection, 'blocked_retry_after_raw') or '未記録'}, "
            "server_retry_at="
            f"{get_meta(connection, 'blocked_server_retry_at') or '未記録'}, "
            "local_safety_until="
            f"{get_meta(connection, 'blocked_local_safety_until') or '未記録'}"
        )
    if get_meta(connection, "restriction_resume_required") == "1":
        print("明示的な再開操作=必要")


def resume_after_block(
    connection: sqlite3.Connection,
    confirmed: bool,
    override_local_safety: bool = False,
) -> None:
    """Locally acknowledge a restriction after its recorded deadline has passed."""
    if not confirmed:
        raise RuntimeError("再開には --confirm が必要です。")
    if get_meta(connection, "restriction_resume_required") != "1":
        print("明示的な再開待ちは記録されていません。")
        return
    blocked_until = parse_iso_time(get_meta(connection, "blocked_until"))
    if blocked_until and blocked_until > utc_now():
        source = get_meta(connection, "blocked_until_source")
        if not override_local_safety:
            raise RuntimeError(
                f"記録された安全期限 {iso_time(blocked_until)} より前には再開できません。"
            )
        if source != "local_safety_fallback":
            raise RuntimeError(
                "サーバー指定期限は--override-local-safetyで上書きできません。"
            )
        set_meta(connection, "restriction_early_resume", "1")
        set_meta(connection, "restriction_early_resume_remaining_requests", "1")
        set_meta(connection, "restriction_early_resume_reason", "human_confirmed_probe")
        set_meta(
            connection, "restriction_early_resume_original_until", iso_time(blocked_until)
        )
    set_meta(connection, "restriction_resume_required", "0")
    set_meta(connection, "restriction_resumed_at", iso_time())
    connection.commit()
    print("アクセス制限後の明示的な再開をローカル状態に記録しました。")


def authorize_local_probe(
    connection: sqlite3.Connection, confirmed: bool, max_requests: int
) -> None:
    """Grant an audited, bounded probe inside a local-only safety window."""
    if not confirmed:
        raise RuntimeError("限定試験の許可には --confirm が必要です。")
    if max_requests < 1 or max_requests > 5:
        raise RuntimeError("限定試験のHTTPリクエスト上限は1〜5件です。")
    blocked_until = parse_iso_time(get_meta(connection, "blocked_until"))
    if blocked_until is None or blocked_until <= utc_now():
        raise RuntimeError("現在有効なlocal safety期限がないため、限定許可は不要です。")
    if get_meta(connection, "blocked_until_source") != "local_safety_fallback":
        raise RuntimeError("サーバー指定期限に対する限定試験は許可できません。")
    if get_meta(connection, "restriction_resume_required") == "1":
        raise RuntimeError("先に人間の明示的なresume操作が必要です。")
    if int(get_meta(connection, "restriction_early_resume_remaining_requests") or "0"):
        raise RuntimeError("未消費の限定試験許可が既にあります。")

    authorized_at = iso_time()
    set_meta(connection, "restriction_early_resume", "1")
    set_meta(connection, "restriction_early_resume_remaining_requests", str(max_requests))
    set_meta(connection, "restriction_early_resume_reason", "human_confirmed_probe")
    set_meta(connection, "restriction_early_resume_authorized_at", authorized_at)
    set_meta(connection, "restriction_early_resume_authorized_requests", str(max_requests))
    set_meta(
        connection, "restriction_early_resume_original_until", iso_time(blocked_until)
    )
    connection.commit()
    print(f"local safety期限内の限定試験を{max_requests}リクエスト許可しました。")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare = subparsers.add_parser(
        "prepare-index", help="トップ1ページから未作成ページを判定する"
    )
    prepare.add_argument("--refresh", action="store_true")
    prepare.add_argument("--delay-seconds", type=float, default=DEFAULT_DELAY_SECONDS)
    prepare.add_argument("--jitter-seconds", type=float, default=DEFAULT_JITTER_SECONDS)

    run = subparsers.add_parser("run", help="個別ページを逐次取得する")
    run.add_argument(
        "--limit",
        "--max-pages",
        dest="limit",
        type=int,
        required=True,
        help="今回の最大新規取得ページ数（--max-pagesは互換用）",
    )
    run.add_argument(
        "--category",
        action="append",
        choices=EVENT_CATEGORIES,
        default=[],
        help="対象分類。複数指定時は指定順に交互処理する",
    )
    run.add_argument("--page-key", action="append", default=[])
    run.add_argument("--exclude-page-key", action="append", default=[])
    run.add_argument(
        "--review-target",
        choices=SUPPORTED_REVIEW_TARGETS,
        help=(
            "events_review.xlsxのimport_target完全一致で候補を絞る。"
            "未指定時はevents_raw.csvの全件を母集団にする"
        ),
    )
    run.add_argument("--delay-seconds", type=float, default=DEFAULT_DELAY_SECONDS)
    run.add_argument("--jitter-seconds", type=float, default=DEFAULT_JITTER_SECONDS)
    run.add_argument(
        "--dry-run",
        action="store_true",
        help="対象一覧だけを表示し、HTTPアクセスも状態更新も行わない",
    )

    subparsers.add_parser("status", help="ネットワークアクセスなしで進捗を表示する")
    subparsers.add_parser("export", help="ネットワークアクセスなしでCSVを再出力する")
    subparsers.add_parser(
        "migrate-state", help="HTTPアクセスなしで制限ログ用の状態形式へ移行する"
    )
    resume = subparsers.add_parser(
        "resume-after-block",
        help="記録期限後に人間の判断で取得再開を明示する（HTTPアクセスなし）",
    )
    resume.add_argument("--confirm", action="store_true")
    resume.add_argument(
        "--override-local-safety",
        action="store_true",
        help=(
            "人間の明示判断でlocal_safety_fallbackだけを期限前に解除し、"
            "監査情報を残す。サーバー指定期限には使用不可"
        ),
    )
    probe = subparsers.add_parser(
        "authorize-local-probe",
        help="local safety期限内に最大5件の監査付き限定試験を許可する（HTTPなし）",
    )
    probe.add_argument("--confirm", action="store_true")
    probe.add_argument("--max-requests", type=int, required=True)
    return parser


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")

    args = build_parser().parse_args()
    try:
        events = read_events()
        with CollectorLock(LOCK_PATH):
            read_only = args.command == "status" or (
                args.command == "run" and args.dry_run
            )
            connection = connect_database(read_only=read_only)
            try:
                if args.command == "prepare-index":
                    validate_delay(args.delay_seconds, args.jitter_seconds)
                    prepare_index(
                        connection,
                        events,
                        args.delay_seconds,
                        args.jitter_seconds,
                        args.refresh,
                    )
                elif args.command == "run":
                    validate_delay(args.delay_seconds, args.jitter_seconds)
                    run_collection(
                        connection,
                        events,
                        args.limit,
                        args.page_key,
                        args.exclude_page_key,
                        args.category,
                        args.review_target,
                        args.delay_seconds,
                        args.jitter_seconds,
                        args.dry_run,
                    )
                elif args.command == "status":
                    print_status(connection, len(events))
                elif args.command == "export":
                    export_outputs(connection)
                    print("CSVを再出力しました。")
                elif args.command == "migrate-state":
                    export_fetch_log(connection)
                    print("制限ログ用の状態形式へ移行し、fetch_log.csvを更新しました。")
                elif args.command == "resume-after-block":
                    resume_after_block(
                        connection, args.confirm, args.override_local_safety
                    )
                elif args.command == "authorize-local-probe":
                    authorize_local_probe(
                        connection, args.confirm, args.max_requests
                    )
            finally:
                connection.close()
    except KeyboardInterrupt:
        print("中断しました。完了済みページは保存されています。", file=sys.stderr)
        return 130
    except RuntimeError as error:
        print(f"停止: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

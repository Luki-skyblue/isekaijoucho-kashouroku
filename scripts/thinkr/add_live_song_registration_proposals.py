"""Append final registration proposals to the live song review workbook.

The source rows and all human-input cells are preserved. This script only adds
proposal columns to NEW_VERSION and NEW_SONG sheets, after a separate read-only
comparison has confirmed that the embedded songs_reference matches Supabase.
It never connects to or writes to Supabase.
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
WORKING_ROOT: Final = (
    REPOSITORY_ROOT / "private-data" / "imports" / "thinkr" / "working"
)
WORKBOOK_PATH: Final = WORKING_ROOT / "live_setlist_song_registration_review.xlsx"
BACKUP_ROOT: Final = WORKING_ROOT / "backups"

EXPECTED_SONG_REFERENCE_COUNT: Final = 369
EXPECTED_VERSION_COUNT: Final = 24
EXPECTED_NEW_SONG_COUNT: Final = 16
EXPECTED_SAFE_COUNT: Final = 37
EXPECTED_HUMAN_COUNT: Final = 3

VERSION_PROPOSAL_HEADERS: Final = [
    "proposed_title",
    "proposed_artist_credit",
    "proposed_song_group_id",
    "proposed_version_type",
    "proposed_version_name",
    "registration_reason",
    "confidence",
    "approval_class",
    "remaining_human_question",
]
NEW_SONG_PROPOSAL_HEADERS: Final = [
    "proposed_title",
    "proposed_artist_credit",
    "proposed_create_song_group",
    "proposed_version_type",
    "proposed_version_name",
    "registration_reason",
    "confidence",
    "approval_class",
    "remaining_human_question",
]

MISSING_GROUP_TITLES: Final = {"ロマンチック願望", "自由に捕らわれる"}
MULTI_PERFORMANCE_TITLE: Final = "生きていく光は"
SOLO_PAIR: Final = ("変身", "ヰ世界情緒")

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
PROPOSAL_FILL: Final = PatternFill("solid", fgColor="E9F0FA")
SAFE_FILL: Final = PatternFill("solid", fgColor="E8F3EB")
HUMAN_FILL: Final = PatternFill("solid", fgColor="FFF0D9")
THIN_BORDER: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def headers(sheet) -> list[str]:
    return [str(cell.value or "") for cell in sheet[1]]


def rows(sheet) -> list[dict[str, object]]:
    fields = headers(sheet)
    return [
        dict(zip(fields, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]


def snapshot(workbook) -> dict[str, tuple[tuple[object, ...], ...]]:
    return {
        sheet.title: tuple(
            tuple(cell.value for cell in row) for row in sheet.iter_rows()
        )
        for sheet in workbook.worksheets
    }


def proposal_for_version(row: dict[str, object]) -> dict[str, object]:
    title = str(row["song_title"])
    artist = str(row["artist_credit"])
    group_id = row["song_group_id_candidate"]

    if title in MISSING_GROUP_TITLES:
        return {
            "proposed_title": title,
            "proposed_artist_credit": artist,
            "proposed_song_group_id": None,
            "proposed_version_type": None,
            "proposed_version_name": None,
            "registration_reason": (
                "Supabaseのsongs/song_groupsに同名・保守的表記差候補がなく、"
                "既存groupへ属するNEW_VERSIONとしては登録できない。"
            ),
            "confidence": "HIGH（候補なしの判定）",
            "approval_class": "NEEDS_HUMAN",
            "remaining_human_question": (
                "新規song_group＋standard songへ再分類するか、先に未登録の基準versionを登録するか。"
            ),
        }

    if title == MULTI_PERFORMANCE_TITLE:
        return {
            "proposed_title": title,
            "proposed_artist_credit": artist,
            "proposed_song_group_id": group_id,
            "proposed_version_type": "live",
            "proposed_version_name": f"{artist} live ver.",
            "registration_reason": (
                "既存groupは一意だが、同じ歌唱者構成が複数公演に現れるため"
                "既存の公演名ベースlive命名をそのまま適用できない。"
            ),
            "confidence": "MEDIUM",
            "approval_class": "NEEDS_HUMAN",
            "remaining_human_question": (
                "version_typeをliveとするかsoloとするか、version_nameを歌唱名義ベースにするか。"
            ),
        }

    version_type = str(row["version_type_candidate"])
    version_name = str(row["version_name_candidate"])
    reason = (
        "既存songと歌唱者構成が異なり、1公演に固有の歌唱version。"
        "既存DBのlive＋『公演名 ver.』形式に一致。"
    )
    if (title, artist) == SOLO_PAIR:
        version_type = "solo"
        version_name = "solo ver."
        reason = (
            "V.W.P標準版に対するヰ世界情緒solo。既存DBのsolo / solo ver.命名に一致。"
        )

    return {
        "proposed_title": title,
        "proposed_artist_credit": artist,
        "proposed_song_group_id": group_id,
        "proposed_version_type": version_type,
        "proposed_version_name": version_name,
        "registration_reason": reason,
        "confidence": "HIGH",
        "approval_class": "SAFE_TO_APPROVE",
        "remaining_human_question": None,
    }


def proposal_for_new_song(row: dict[str, object]) -> dict[str, object]:
    if row["existing_title_check"] != "NO_EXACT_OR_NORMALIZED_MATCH":
        raise RuntimeError(f"NEW_SONGの重複確認が未完了です: {row['title']}")
    return {
        "proposed_title": row["title"],
        "proposed_artist_credit": row["artist_credit"],
        "proposed_create_song_group": True,
        "proposed_version_type": "standard",
        "proposed_version_name": None,
        "registration_reason": (
            "Supabase実データと一致確認済みのsongs_referenceに完全一致・"
            "保守的表記差一致がなく、既存の初回登録規則（新規group＋standard）を適用可能。"
        ),
        "confidence": "HIGH",
        "approval_class": "SAFE_TO_APPROVE",
        "remaining_human_question": None,
    }


def append_proposals(sheet, proposal_headers: list[str], proposals: list[dict[str, object]]) -> None:
    existing_headers = headers(sheet)
    if any(name in existing_headers for name in proposal_headers):
        raise RuntimeError(f"{sheet.title}には既に提案列があります。上書きしません。")
    if len(proposals) != sheet.max_row - 1:
        raise RuntimeError(f"{sheet.title}の行数と提案件数が一致しません。")

    start_column = sheet.max_column + 1
    for offset, name in enumerate(proposal_headers):
        cell = sheet.cell(1, start_column + offset, name)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.border = THIN_BORDER
        cell.comment = Comment(
            "既存DB規則に基づく提案値。人間入力欄ではなく、DB未反映です。", "Codex"
        )
    for row_number, proposal in enumerate(proposals, start=2):
        for offset, name in enumerate(proposal_headers):
            cell = sheet.cell(row_number, start_column + offset, proposal[name])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            cell.fill = PROPOSAL_FILL
            if name not in {
                "proposed_song_group_id",
                "proposed_create_song_group",
            }:
                cell.number_format = "@"
        approval_cell = sheet.cell(
            row_number, start_column + proposal_headers.index("approval_class")
        )
        approval_cell.fill = (
            SAFE_FILL
            if proposal["approval_class"] == "SAFE_TO_APPROVE"
            else HUMAN_FILL
        )

    width_by_name = {
        "proposed_title": 38,
        "proposed_artist_credit": 38,
        "proposed_song_group_id": 22,
        "proposed_create_song_group": 26,
        "proposed_version_type": 22,
        "proposed_version_name": 44,
        "registration_reason": 72,
        "confidence": 24,
        "approval_class": 24,
        "remaining_human_question": 72,
    }
    for offset, name in enumerate(proposal_headers):
        sheet.column_dimensions[get_column_letter(start_column + offset)].width = width_by_name[name]
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    )


def backup_path() -> Path:
    timestamp = datetime.now(ZoneInfo("Asia/Tokyo")).strftime("%Y%m%d-%H%M%S")
    candidate = BACKUP_ROOT / (
        f"live_setlist_song_registration_review.before-proposals-{timestamp}.xlsx"
    )
    suffix = 1
    while candidate.exists():
        candidate = BACKUP_ROOT / (
            "live_setlist_song_registration_review.before-proposals-"
            f"{timestamp}-{suffix}.xlsx"
        )
        suffix += 1
    return candidate


def validate_preserved(
    saved_path: Path, original: dict[str, tuple[tuple[object, ...], ...]]
) -> None:
    workbook = load_workbook(saved_path, read_only=False, data_only=False)
    try:
        for sheet_name, old_rows in original.items():
            sheet = workbook[sheet_name]
            for row_number, old_row in enumerate(old_rows, start=1):
                current = tuple(
                    sheet.cell(row_number, column).value
                    for column in range(1, len(old_row) + 1)
                )
                if current != old_row:
                    raise RuntimeError(
                        f"{sheet_name}の既存セルが意図せず変化しました: row {row_number}"
                    )

        versions = rows(workbook["new_versions"])
        new_songs = rows(workbook["new_songs"])
        combined = versions + new_songs
        safe = sum(row["approval_class"] == "SAFE_TO_APPROVE" for row in combined)
        human = sum(row["approval_class"] == "NEEDS_HUMAN" for row in combined)
        if safe != EXPECTED_SAFE_COUNT or human != EXPECTED_HUMAN_COUNT:
            raise RuntimeError(f"分類件数が想定外です: SAFE={safe}, HUMAN={human}")
        if any(
            row["proposed_song_group_id"] is None
            for row in versions
            if row["approval_class"] == "SAFE_TO_APPROVE"
        ):
            raise RuntimeError("SAFE_TO_APPROVEのNEW_VERSIONにgroup候補NULLがあります。")
        if any(
            row["proposed_version_type"] != "standard"
            or row["proposed_version_name"] is not None
            or row["proposed_create_song_group"] is not True
            for row in new_songs
        ):
            raise RuntimeError("NEW_SONGの標準初回登録案が一致しません。")
    finally:
        workbook.close()


def run() -> Path:
    if not WORKBOOK_PATH.exists():
        raise RuntimeError(f"登録reviewがありません: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    temporary_handle, temporary_name = tempfile.mkstemp(
        prefix=".live-song-proposals-", suffix=".xlsx", dir=WORKING_ROOT
    )
    os.close(temporary_handle)
    temporary = Path(temporary_name)
    try:
        original = snapshot(workbook)
        songs = rows(workbook["songs_reference"])
        if len(songs) != EXPECTED_SONG_REFERENCE_COUNT:
            raise RuntimeError("songs_referenceが369件ではありません。")
        versions = rows(workbook["new_versions"])
        new_songs = rows(workbook["new_songs"])
        if len(versions) != EXPECTED_VERSION_COUNT or len(new_songs) != EXPECTED_NEW_SONG_COUNT:
            raise RuntimeError("review対象件数が24件 / 16件ではありません。")

        group_ids = {int(song["song_group_id"]) for song in songs}
        for row in versions:
            group_id = row["song_group_id_candidate"]
            if group_id is not None and int(group_id) not in group_ids:
                raise RuntimeError(f"存在しないsong_group_id候補です: {group_id}")

        version_proposals = [proposal_for_version(row) for row in versions]
        song_proposals = [proposal_for_new_song(row) for row in new_songs]
        append_proposals(
            workbook["new_versions"], VERSION_PROPOSAL_HEADERS, version_proposals
        )
        append_proposals(
            workbook["new_songs"], NEW_SONG_PROPOSAL_HEADERS, song_proposals
        )

        workbook.properties.description = (
            "Final proposals appended from Supabase-verified songs/song_groups rules. "
            "No DB writes. Existing cells and human inputs preserved."
        )
        workbook.save(temporary)
        workbook.close()
        validate_preserved(temporary, original)

        BACKUP_ROOT.mkdir(parents=True, exist_ok=True)
        backup = backup_path()
        shutil.copy2(WORKBOOK_PATH, backup)
        os.replace(temporary, WORKBOOK_PATH)
        return backup
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        temporary.unlink(missing_ok=True)


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        backup = run()
        print(f"review updated: {WORKBOOK_PATH}")
        print(f"backup created: {backup}")
        print(f"SAFE_TO_APPROVE: {EXPECTED_SAFE_COUNT}")
        print(f"NEEDS_HUMAN: {EXPECTED_HUMAN_COUNT}")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError, KeyError) as error:
        print(f"最終提案の追記を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

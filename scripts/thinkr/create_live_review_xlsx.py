"""Create a local-only workbook for reviewing collected THINKR live data."""

from __future__ import annotations

import csv
import os
import sys
import warnings
from collections import Counter
from pathlib import Path
from typing import Final, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


warnings.filterwarnings(
    "ignore", message="Data Validation extension is not supported.*"
)


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
DATA_ROOT: Final = REPOSITORY_ROOT / "private-data" / "imports" / "thinkr"
EVENTS_RAW_PATH: Final = DATA_ROOT / "raw" / "events_raw.csv"
EVENTS_REVIEW_XLSX_PATH: Final = DATA_ROOT / "working" / "events_review.xlsx"
SETLISTS_RAW_PATH: Final = DATA_ROOT / "raw" / "setlists_raw.csv"
PAGES_PATH: Final = DATA_ROOT / "working" / "setlists" / "pages.csv"
OUTPUT_PATH: Final = DATA_ROOT / "working" / "live_review.xlsx"

EVENT_RAW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "source_url",
    "import_target",
]
EVENT_REVIEW_FIELDS: Final = [
    "title",
    "title_kana",
    "artist_credit",
    "event_type",
    "performance_name",
    "performance_date",
    "venue",
    "official_url",
    "review_note",
]
EVENT_FIELDS: Final = [*EVENT_RAW_FIELDS, *EVENT_REVIEW_FIELDS]

SETLIST_RAW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "import_target",
    "setlist_table_index",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "source_url",
]
SETLIST_REVIEW_FIELDS: Final = ["song_id", "joucho_participation", "review_note"]
SETLIST_FIELDS: Final = [*SETLIST_RAW_FIELDS, *SETLIST_REVIEW_FIELDS]

EXPECTED_EVENTS_RAW_FIELDS: Final = [
    "source_category",
    "year",
    "page_key",
    "event_date_raw",
    "event_title_raw",
    "source_url",
    "import_target",
]
EXPECTED_EVENTS_REVIEW_FIELDS: Final = [
    *EXPECTED_EVENTS_RAW_FIELDS,
    "review_note",
]
EXPECTED_SETLISTS_RAW_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "setlist_table_index",
    "setlist_section_raw",
    "setlist_table_caption_raw",
    "setlist_table_headers_raw",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "setlist_row_cells_raw",
    "source_url",
]
EXPECTED_PAGES_FIELDS: Final = [
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "status",
    "setlist_count",
    "parse_warning_count",
    "http_status",
    "attempt_count",
    "last_attempt_at",
    "next_retry_at",
    "error_message",
    "fetched_at",
    "source_url",
]
ALLOWED_IMPORT_TARGETS: Final = {None, "", "TRUE", "FALSE", "?"}

EVENT_WIDTHS: Final = {
    "A": 16,
    "B": 15,
    "C": 60,
    "D": 18,
    "E": 52,
    "F": 15,
    "G": 42,
    "H": 28,
    "I": 28,
    "J": 20,
    "K": 32,
    "L": 18,
    "M": 26,
    "N": 48,
    "O": 42,
}
SETLIST_WIDTHS: Final = {
    "A": 16,
    "B": 15,
    "C": 52,
    "D": 18,
    "E": 15,
    "F": 18,
    "G": 18,
    "H": 42,
    "I": 36,
    "J": 42,
    "K": 52,
    "L": 16,
    "M": 24,
    "N": 42,
}


def read_csv(path: Path, expected_fields: list[str]) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if reader.fieldnames != expected_fields:
            raise RuntimeError(f"{path.name}の列構成が想定と異なります。")
        return list(reader)


def read_import_targets() -> dict[str, str]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore", message="Data Validation extension is not supported.*"
        )
        workbook = load_workbook(
            EVENTS_REVIEW_XLSX_PATH, read_only=True, data_only=False
        )
    try:
        if "events_review" not in workbook.sheetnames:
            raise RuntimeError("events_review.xlsxにevents_reviewシートがありません。")
        worksheet = workbook["events_review"]
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows, ()))
        if headers != EXPECTED_EVENTS_REVIEW_FIELDS:
            raise RuntimeError("events_review.xlsxの列構成が想定と異なります。")
        page_key_index = headers.index("page_key")
        import_target_index = headers.index("import_target")
        targets: dict[str, str] = {}
        for row_number, row in enumerate(rows, start=2):
            page_key = row[page_key_index]
            target = row[import_target_index]
            if not isinstance(page_key, str):
                raise RuntimeError(
                    f"events_review.xlsxの{row_number}行目のpage_keyが文字列ではありません。"
                )
            if page_key in targets:
                raise RuntimeError(
                    f"events_review.xlsxに重複するpage_keyがあります: {page_key}"
                )
            if target not in ALLOWED_IMPORT_TARGETS:
                raise RuntimeError(
                    f"events_review.xlsxの{row_number}行目に未対応のimport_targetがあります。"
                )
            targets[page_key] = "" if target is None else target
        return targets
    finally:
        workbook.close()


def build_rows() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    events_raw = read_csv(EVENTS_RAW_PATH, EXPECTED_EVENTS_RAW_FIELDS)
    setlists_raw = read_csv(SETLISTS_RAW_PATH, EXPECTED_SETLISTS_RAW_FIELDS)
    pages = read_csv(PAGES_PATH, EXPECTED_PAGES_FIELDS)
    import_targets = read_import_targets()

    events_by_key = {row["page_key"]: row for row in events_raw}
    if len(events_by_key) != len(events_raw):
        raise RuntimeError("events_raw.csvに重複するpage_keyがあります。")
    raw_keys = set(events_by_key)
    if set(import_targets) != raw_keys:
        raise RuntimeError("events_review.xlsxとevents_raw.csvのpage_keyが一致しません。")

    completed_keys = {
        row["page_key"] for row in pages if row["status"] == "completed"
    }
    if len(completed_keys) != 35:
        raise RuntimeError(
            f"取得済みページは35件を想定していますが、{len(completed_keys)}件です。"
        )
    if not completed_keys <= raw_keys:
        raise RuntimeError("pages.csvにevents_raw.csv未登録のpage_keyがあります。")

    event_rows: list[dict[str, str]] = []
    for raw in events_raw:
        page_key = raw["page_key"]
        if page_key not in completed_keys:
            continue
        event_rows.append(
            {
                "source_category": raw["source_category"],
                "page_key": page_key,
                "event_title_raw": raw["event_title_raw"],
                "event_date_raw": raw["event_date_raw"],
                "source_url": raw["source_url"],
                "import_target": import_targets[page_key],
                **{field: "" for field in EVENT_REVIEW_FIELDS},
            }
        )

    setlist_rows: list[dict[str, str]] = []
    for raw in setlists_raw:
        page_key = raw["page_key"]
        if page_key not in completed_keys:
            raise RuntimeError(
                f"setlists_raw.csvに取得済みでないpage_keyがあります: {page_key}"
            )
        event = events_by_key[page_key]
        setlist_rows.append(
            {
                "source_category": event["source_category"],
                "page_key": page_key,
                "event_title_raw": raw["event_title_raw"],
                "event_date_raw": raw["event_date_raw"],
                "import_target": import_targets[page_key],
                "setlist_table_index": raw["setlist_table_index"],
                "setlist_no_raw": raw["setlist_no_raw"],
                "song_title_raw": raw["song_title_raw"],
                "artist_credit_raw": raw["artist_credit_raw"],
                "note_raw": raw["note_raw"],
                "source_url": raw["source_url"],
                **{field: "" for field in SETLIST_REVIEW_FIELDS},
            }
        )

    if len(event_rows) != 35 or len(setlist_rows) != 571:
        raise RuntimeError(
            f"出力件数が想定外です: events={len(event_rows)}, "
            f"setlists={len(setlist_rows)}"
        )
    return event_rows, setlist_rows


def populate_sheet(
    worksheet,
    fields: list[str],
    raw_fields: list[str],
    rows: Iterable[dict[str, str]],
    widths: dict[str, int],
    wide_fields: set[str],
) -> int:
    raw_fill = PatternFill(fill_type="solid", fgColor="E7E5E1")
    review_fill = PatternFill(fill_type="solid", fgColor="DCEAF3")
    review_body_fill = PatternFill(fill_type="solid", fgColor="F3F8FB")
    raw_count = len(raw_fields)

    for column_index, field in enumerate(fields, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=field)
        cell.data_type = "s"
        cell.number_format = "@"
        cell.font = Font(bold=True)
        cell.fill = raw_fill if column_index <= raw_count else review_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    row_count = 0
    for row_index, row in enumerate(rows, start=2):
        row_count += 1
        for column_index, field in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row[field])
            cell.data_type = "s"
            cell.number_format = "@"
            cell.alignment = Alignment(
                vertical="top", wrap_text=field in wide_fields
            )
            if column_index > raw_count:
                cell.fill = review_body_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(fields)).column_letter}{row_count + 1}"
    worksheet.row_dimensions[1].height = 30
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width
    return row_count


def write_workbook(
    event_rows: list[dict[str, str]], setlist_rows: list[dict[str, str]]
) -> None:
    if OUTPUT_PATH.exists():
        raise RuntimeError(
            f"既存の人間入力を保護するため上書きしません: {OUTPUT_PATH}"
        )
    workbook = Workbook()
    events_sheet = workbook.active
    events_sheet.title = "events"
    populate_sheet(
        events_sheet,
        EVENT_FIELDS,
        EVENT_RAW_FIELDS,
        event_rows,
        EVENT_WIDTHS,
        {"event_title_raw", "source_url", "official_url", "review_note"},
    )

    setlists_sheet = workbook.create_sheet("setlists")
    populate_sheet(
        setlists_sheet,
        SETLIST_FIELDS,
        SETLIST_RAW_FIELDS,
        setlist_rows,
        SETLIST_WIDTHS,
        {
            "event_title_raw",
            "song_title_raw",
            "artist_credit_raw",
            "note_raw",
            "source_url",
            "review_note",
        },
    )

    validation_sheet = workbook.create_sheet("_validation")
    for row_index, value in enumerate(("TRUE", "FALSE", "?"), start=1):
        cell = validation_sheet.cell(row=row_index, column=1, value=value)
        cell.data_type = "s"
        cell.number_format = "@"
    validation_sheet.sheet_state = "hidden"
    participation_validation = DataValidation(
        type="list",
        formula1="'_validation'!$A$1:$A$3",
        allow_blank=True,
    )
    participation_validation.error = "TRUE、FALSE、? のいずれかを選択してください。"
    participation_validation.errorTitle = "入力値を確認してください"
    participation_validation.prompt = "TRUE、FALSE、? から選択できます。"
    participation_validation.promptTitle = "ヰ世界情緒の歌唱参加"
    participation_validation.showErrorMessage = True
    participation_validation.showInputMessage = True
    setlists_sheet.add_data_validation(participation_validation)
    participation_validation.add(f"M2:M{len(setlist_rows) + 1}")

    temporary_path = OUTPUT_PATH.with_name(f"{OUTPUT_PATH.stem}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        if OUTPUT_PATH.exists():
            raise RuntimeError(
                f"既存の人間入力を保護するため上書きしません: {OUTPUT_PATH}"
            )
        os.replace(temporary_path, OUTPUT_PATH)
    finally:
        workbook.close()
        if temporary_path.exists():
            temporary_path.unlink()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        if OUTPUT_PATH.exists():
            raise RuntimeError(
                f"既存の人間入力を保護するため上書きしません: {OUTPUT_PATH}"
            )
        event_rows, setlist_rows = build_rows()
        write_workbook(event_rows, setlist_rows)
    except (OSError, RuntimeError) as error:
        print(f"Excel作成を停止しました: {error}", file=sys.stderr)
        return 1

    targets = Counter(row["import_target"] or "空欄" for row in event_rows)
    print(f"出力先: {OUTPUT_PATH}")
    print(f"events: {len(event_rows)}件")
    print(
        "import_target: "
        f"TRUE={targets['TRUE']}, ?={targets['?']}, "
        f"FALSE={targets['FALSE']}, 空欄={targets['空欄']}"
    )
    print(f"setlists: {len(setlist_rows)}件")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

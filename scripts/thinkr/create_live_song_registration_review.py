"""Create the final local review workbook for live song registration.

This script is deliberately local-only. It reads the existing setlist song
resolution workbook and creates a new workbook without modifying the source,
Supabase, songs, or live setlist rows.
"""

from __future__ import annotations

import hashlib
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
WORKING_ROOT: Final = (
    REPOSITORY_ROOT / "private-data" / "imports" / "thinkr" / "working"
)
SOURCE_PATH: Final = WORKING_ROOT / "live_setlist_song_review.xlsx"
OUTPUT_PATH: Final = WORKING_ROOT / "live_setlist_song_registration_review.xlsx"

EXPECTED_COUNTS: Final = {
    "LINK_EXISTING": 211,
    "NEW_VERSION": 24,
    "NEW_SONG": 16,
    "SPECIAL_UNRESOLVED": 3,
}

LAGTRAIN_PAIR: Final = ("ラグトレイン", "ヰ世界情緒")
LAGTRAIN_STANDARD_ID: Final = 126

HEADER_FILL: Final = PatternFill("solid", fgColor="E9E7E2")
REFERENCE_FILL: Final = PatternFill("solid", fgColor="F4F2ED")
CANDIDATE_FILL: Final = PatternFill("solid", fgColor="FFF0D9")
HUMAN_FILL: Final = PatternFill("solid", fgColor="E8F3EB")
SPECIAL_FILL: Final = PatternFill("solid", fgColor="F8E2E2")
THIN_BORDER: Final = Border(bottom=Side(style="hair", color="CFCBC2"))


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sheet_rows(sheet) -> list[dict[str, object]]:
    fields = [str(cell.value or "") for cell in sheet[1]]
    return [
        dict(zip(fields, values, strict=True))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]


def comparison_key(value: object) -> str:
    """Conservative duplicate-review key; never used to auto-link a song."""

    normalized = unicodedata.normalize("NFKC", str(value or "")).casefold()
    return "".join(character for character in normalized if character.isalnum())


def apply_lagtrain_decision(
    resolutions: list[dict[str, object]], songs_by_id: dict[int, dict[str, object]]
) -> None:
    matches = [
        row
        for row in resolutions
        if (row["song_title_raw"], row["artist_credit_raw"]) == LAGTRAIN_PAIR
    ]
    if len(matches) != 1:
        raise RuntimeError("ラグトレインのreview行を一意に特定できません。")
    target = songs_by_id.get(LAGTRAIN_STANDARD_ID)
    if not target:
        raise RuntimeError("ラグトレイン通常版のsongs参照行がありません。")
    if (
        target["title"] != LAGTRAIN_PAIR[0]
        or target["artist_credit"] != LAGTRAIN_PAIR[1]
        or target["version_type"] != "standard"
    ):
        raise RuntimeError("ラグトレイン通常版のsongs参照内容が想定と異なります。")

    row = matches[0]
    row.update(
        {
            "resolution_status": "LINK_EXISTING",
            "song_id": target["id"],
            "reference_song_id": target["id"],
            "reference_song_title": target["title"],
            "reference_artist_credit": target["artist_credit"],
            "reference_version_type": target["version_type"],
            "reference_version_name": target["version_name"],
            "song_group_id": target["song_group_id"],
            "decision_note": "人間判断によりCANDY LIVEで歌唱した通常版へリンク確定。",
        }
    )


def base_title(row: dict[str, object]) -> str:
    raw = str(row["song_title_raw"])
    if row.get("annotation_raw") == "アーカイブ限定":
        return raw.replace("(アーカイブ限定)", "").strip()
    return str(row.get("reference_song_title") or raw)


def version_name_candidate(row: dict[str, object]) -> str | None:
    performances = [
        value.strip()
        for value in str(row.get("example_performances") or "").split(" / ")
        if value.strip()
    ]
    if len(performances) == 1:
        return f"{performances[0]} ver."
    if len(performances) > 1:
        return f"{row['artist_credit_raw']} live ver."
    return None


def prepare_version_rows(
    rows: list[dict[str, object]], songs: list[dict[str, object]]
) -> list[dict[str, object]]:
    by_title: dict[str, list[dict[str, object]]] = {}
    for song in songs:
        by_title.setdefault(str(song["title"]), []).append(song)

    result: list[dict[str, object]] = []
    for row in rows:
        title = base_title(row)
        group_id = row.get("song_group_id")
        if group_id is None:
            candidate_groups = {
                song["song_group_id"]
                for song in by_title.get(title, [])
                if song.get("song_group_id") is not None
            }
            if len(candidate_groups) == 1:
                group_id = candidate_groups.pop()

        candidate_name = version_name_candidate(row)
        notes = [
            "歌唱者構成が既存songと異なるため新version候補。",
            "feat./×/with/&や名義順だけの差をversion理由にはしていません。",
        ]
        if row.get("annotation_raw"):
            notes.append(
                f"{row['annotation_raw']}はversion名に含めず、setlist行annotationとして保持。"
            )
        if group_id is None:
            notes.append("対応するsong_group_idを既存参照から一意に特定できません。")
        if len(str(row.get("example_performances") or "").split(" / ")) > 1:
            notes.append(
                "複数公演で同じ歌唱者構成を確認。公演名ではなく歌唱名義ベースのversion_name候補。"
            )

        result.append(
            {
                "song_title": title,
                "artist_credit": row["artist_credit_raw"],
                "song_group_id_candidate": group_id,
                "version_type_candidate": "live",
                "version_name_candidate": candidate_name,
                "song_title_raw": row["song_title_raw"],
                "artist_credit_raw": row["artist_credit_raw"],
                "reference_song_id": row.get("reference_song_id"),
                "reference_artist_credit": row.get("reference_artist_credit"),
                "occurrence_count": row["occurrence_count"],
                "example_performances": row["example_performances"],
                "annotation_raw": row.get("annotation_raw"),
                "review_decision": None,
                "confirmed_song_group_id": None,
                "confirmed_version_type": None,
                "confirmed_version_name": None,
                "created_song_id": None,
                "note": " ".join(notes),
            }
        )
    return sorted(result, key=lambda row: (str(row["song_title"]), str(row["artist_credit"])))


def prepare_new_song_rows(
    rows: list[dict[str, object]], songs: list[dict[str, object]]
) -> list[dict[str, object]]:
    songs_by_key: dict[str, list[dict[str, object]]] = {}
    for song in songs:
        songs_by_key.setdefault(comparison_key(song["title"]), []).append(song)

    result: list[dict[str, object]] = []
    for row in rows:
        normalized_matches = songs_by_key.get(comparison_key(row["song_title_raw"]), [])
        if normalized_matches:
            raise RuntimeError(
                "NEW_SONGに既存titleの表記差候補があります: "
                f"{row['song_title_raw']}"
            )
        result.append(
            {
                "title": row["song_title_raw"],
                "artist_credit": row["artist_credit_raw"],
                "occurrence_count": row["occurrence_count"],
                "example_performances": row["example_performances"],
                "song_title_raw": row["song_title_raw"],
                "artist_credit_raw": row["artist_credit_raw"],
                "existing_title_check": "NO_EXACT_OR_NORMALIZED_MATCH",
                "song_group_creation_candidate": True,
                "review_decision": None,
                "created_song_group_id": None,
                "created_song_id": None,
                "note": "ローカルsongs_referenceで完全一致・NFKC/大小文字/空白/記号差一致なし。新規song_group＋song候補。",
            }
        )
    return sorted(result, key=lambda row: (str(row["title"]), str(row["artist_credit"])))


def prepare_special_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    result = []
    for row in rows:
        result.append(
            {
                "song_title_raw": row["song_title_raw"],
                "artist_credit_raw": row["artist_credit_raw"],
                "occurrence_count": row["occurrence_count"],
                "example_performances": row["example_performances"],
                "song_id": None,
                "resolution_status": "SPECIAL_UNRESOLVED",
                "note": row.get("decision_note")
                or "composite/mashupモデル確定まで通常song_idを割り当てない。",
            }
        )
    return sorted(result, key=lambda row: (str(row["song_title_raw"]), str(row["artist_credit_raw"])))


def prepare_link_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(
        [
            {
                "song_title_raw": row["song_title_raw"],
                "artist_credit_raw": row["artist_credit_raw"],
                "occurrence_count": row["occurrence_count"],
                "example_performances": row["example_performances"],
                "song_id": row["song_id"],
                "reference_song_title": row["reference_song_title"],
                "reference_artist_credit": row["reference_artist_credit"],
                "reference_version_type": row["reference_version_type"],
                "reference_version_name": row["reference_version_name"],
                "song_group_id": row["song_group_id"],
                "annotation_raw": row["annotation_raw"],
                "note": row["decision_note"],
            }
            for row in rows
        ],
        key=lambda row: (str(row["song_title_raw"]), str(row["artist_credit_raw"])),
    )


def append_table(
    workbook: Workbook,
    name: str,
    headers: list[str],
    rows: list[dict[str, object]],
    widths: dict[str, float],
    *,
    human_columns: set[str] | None = None,
    candidate_columns: set[str] | None = None,
    special: bool = False,
) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(field) for field in headers])

    human_columns = human_columns or set()
    candidate_columns = candidate_columns or set()
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="34312D")
        cell.border = THIN_BORDER
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            cell.number_format = "@" if cell.column != 3 else "General"
        for header_name in human_columns:
            row[headers.index(header_name)].fill = HUMAN_FILL
        for header_name in candidate_columns:
            row[headers.index(header_name)].fill = CANDIDATE_FILL
        if special:
            for cell in row:
                cell.fill = SPECIAL_FILL

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(rows) + 1}"

    if "review_decision" in headers:
        column = headers.index("review_decision") + 1
        letter = chr(64 + column)
        validation = DataValidation(
            type="list", formula1='"ACCEPT,REVISE,REJECT"', allow_blank=True
        )
        sheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{len(rows) + 1}")
        sheet.cell(1, column).comment = Comment(
            "人間確認用。候補は未確定のため初期値は空欄です。", "Codex"
        )


def copy_songs_reference(workbook: Workbook, rows: list[dict[str, object]]) -> None:
    headers = list(rows[0])
    append_table(
        workbook,
        "songs_reference",
        headers,
        rows,
        {
            "A": 10,
            "B": 38,
            "C": 32,
            "D": 32,
            "E": 36,
            "F": 16,
            "G": 28,
            "H": 18,
            "I": 16,
            "J": 16,
            "K": 16,
        },
    )


def validate_output(path: Path) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        expected = {
            "summary": 4,
            "link_existing": 211,
            "new_versions": 24,
            "new_songs": 16,
            "special_unresolved": 3,
            "songs_reference": 369,
        }
        for sheet_name, count in expected.items():
            if workbook[sheet_name].max_row - 1 != count:
                raise RuntimeError(f"{sheet_name}の件数が想定外です。")
        link_headers = [cell.value for cell in workbook["link_existing"][1]]
        song_id_column = link_headers.index("song_id") + 1
        if any(
            workbook["link_existing"].cell(row, song_id_column).value is None
            for row in range(2, workbook["link_existing"].max_row + 1)
        ):
            raise RuntimeError("LINK_EXISTINGにsong_id NULLがあります。")
        special_headers = [cell.value for cell in workbook["special_unresolved"][1]]
        special_song_id = special_headers.index("song_id") + 1
        if any(
            workbook["special_unresolved"].cell(row, special_song_id).value is not None
            for row in range(2, workbook["special_unresolved"].max_row + 1)
        ):
            raise RuntimeError("SPECIAL_UNRESOLVEDにsong_idが入っています。")
    finally:
        workbook.close()


def run() -> None:
    if not SOURCE_PATH.exists():
        raise RuntimeError(f"入力workbookがありません: {SOURCE_PATH}")
    if OUTPUT_PATH.exists():
        raise RuntimeError(f"既存の最終reviewを上書きしません: {OUTPUT_PATH}")

    source_hash = file_sha256(SOURCE_PATH)
    source = load_workbook(SOURCE_PATH, read_only=True, data_only=False)
    try:
        resolutions = sheet_rows(source["song_id_review"])
        songs = sheet_rows(source["songs_reference"])
    finally:
        source.close()

    songs_by_id = {int(song["id"]): song for song in songs}
    apply_lagtrain_decision(resolutions, songs_by_id)
    counts = Counter(str(row["resolution_status"]) for row in resolutions)
    if dict(counts) != EXPECTED_COUNTS:
        raise RuntimeError(f"最終status件数が想定外です: {dict(counts)}")

    links = prepare_link_rows(
        [row for row in resolutions if row["resolution_status"] == "LINK_EXISTING"]
    )
    versions = prepare_version_rows(
        [row for row in resolutions if row["resolution_status"] == "NEW_VERSION"],
        songs,
    )
    new_songs = prepare_new_song_rows(
        [row for row in resolutions if row["resolution_status"] == "NEW_SONG"],
        songs,
    )
    special = prepare_special_rows(
        [
            row
            for row in resolutions
            if row["resolution_status"] == "SPECIAL_UNRESOLVED"
        ]
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("summary")
    summary.append(["status", "pair_count", "human_review_required", "note"])
    summary_rows = [
        ("LINK_EXISTING", 211, 0, "既存song_id対応を維持。DB未反映。"),
        ("NEW_VERSION", 24, 24, "version候補。人間確定待ち。"),
        ("NEW_SONG", 16, 16, "song_group＋song作成候補。人間確定待ち。"),
        ("SPECIAL_UNRESOLVED", 3, 0, "composite/mashupモデル待ち。"),
    ]
    for row in summary_rows:
        summary.append(row)
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = "A1:D5"
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["C"].width = 24
    summary.column_dimensions["D"].width = 60

    append_table(
        workbook,
        "link_existing",
        list(links[0]),
        links,
        {"A": 42, "B": 36, "C": 14, "D": 58, "E": 12, "F": 38, "G": 36,
         "H": 18, "I": 24, "J": 16, "K": 20, "L": 58},
        candidate_columns={"song_id"},
    )
    append_table(
        workbook,
        "new_versions",
        list(versions[0]),
        versions,
        {"A": 38, "B": 36, "C": 20, "D": 20, "E": 42, "F": 42, "G": 36,
         "H": 18, "I": 36, "J": 14, "K": 58, "L": 20, "M": 18, "N": 20,
         "O": 20, "P": 42, "Q": 16, "R": 72},
        human_columns={"review_decision", "confirmed_song_group_id", "confirmed_version_type",
                       "confirmed_version_name", "created_song_id"},
        candidate_columns={"song_group_id_candidate", "version_type_candidate",
                           "version_name_candidate"},
    )
    append_table(
        workbook,
        "new_songs",
        list(new_songs[0]),
        new_songs,
        {"A": 42, "B": 36, "C": 14, "D": 58, "E": 42, "F": 36, "G": 34,
         "H": 24, "I": 18, "J": 22, "K": 18, "L": 70},
        human_columns={"review_decision", "created_song_group_id", "created_song_id"},
        candidate_columns={"song_group_creation_candidate"},
    )
    append_table(
        workbook,
        "special_unresolved",
        list(special[0]),
        special,
        {"A": 46, "B": 38, "C": 14, "D": 60, "E": 12, "F": 28, "G": 76},
        special=True,
    )
    copy_songs_reference(workbook, songs)

    now = datetime.now(ZoneInfo("Asia/Tokyo"))
    workbook.properties.title = "Live setlist song registration final review"
    workbook.properties.description = (
        "Local review only. No DB changes. Candidates require human confirmation."
    )
    workbook.properties.created = now.replace(tzinfo=None)
    workbook.save(OUTPUT_PATH)
    workbook.close()

    validate_output(OUTPUT_PATH)
    if file_sha256(SOURCE_PATH) != source_hash:
        raise RuntimeError("入力workbookが意図せず変更されました。")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        run()
        print(f"review created: {OUTPUT_PATH}")
        for status, count in EXPECTED_COUNTS.items():
            print(f"{status}: {count}")
        print("human review candidates: 40")
        print("validation: PASS")
        return 0
    except (OSError, RuntimeError, ValueError, KeyError) as error:
        print(f"最終song review作成を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

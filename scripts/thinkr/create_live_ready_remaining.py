"""Create a validated ready batch for reviewed, non-pilot TRUE live events.

This is a local-only transform. It reads the collected raw CSVs and the human
review workbook, never contacts Supabase or WIKIWIKI, and refuses to overwrite
an existing ready directory.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import sys
import tempfile
import warnings
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Final
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore", message="Data Validation extension is not supported.*"
)

REPOSITORY_ROOT: Final = Path(__file__).resolve().parents[2]
DATA_ROOT: Final = REPOSITORY_ROOT / "private-data" / "imports" / "thinkr"
EVENTS_RAW_PATH: Final = DATA_ROOT / "raw" / "events_raw.csv"
SETLISTS_RAW_PATH: Final = DATA_ROOT / "raw" / "setlists_raw.csv"
LIVE_REVIEW_PATH: Final = DATA_ROOT / "working" / "live_review.xlsx"
READY_ROOT: Final = DATA_ROOT / "ready"

BATCH_NAME: Final = "true-remaining-001"
OUTPUT_PATH: Final = READY_ROOT / BATCH_NAME
PILOT_PAGE_KEYS: Final = {"2021.1023", "2026.0501", "2026.0502"}
EXPECTED_PERFORMANCE_COUNT: Final = 26

EVENT_REVIEW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "source_url",
    "import_target",
    "title",
    "title_kana",
    "artist_credit",
    "event_type",
    "performance_name",
    "performance_date",
    "venue",
    "official_url",
    "review_note",
]
SETLIST_REVIEW_FIELDS: Final = [
    "source_category",
    "page_key",
    "event_title_raw",
    "event_date_raw",
    "import_target",
    "setlist_table_index",
    "setlist_no_raw",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "source_url",
    "song_id",
    "joucho_participation",
    "review_note",
]
EVENT_GROUP_FIELDS: Final = [
    "local_group_key",
    "title",
    "title_kana",
    "sort_title",
    "notes",
]
PERFORMANCE_FIELDS: Final = [
    "performance_key",
    "local_group_key",
    "group_sort_order",
    "title",
    "title_kana",
    "sort_title",
    "artist_credit",
    "performance_date",
    "format_label",
    "image_url",
    "venue",
    "streaming_platforms",
    "notes",
    "published_at",
    "is_listed",
]
SOURCE_FIELDS: Final = [
    "performance_key",
    "source_type",
    "source_url",
    "source_key",
    "event_title_raw",
    "event_date_raw",
    "source_category_raw",
    "notes",
]
LINK_FIELDS: Final = [
    "performance_key",
    "link_type",
    "label",
    "url",
    "published_date",
    "notes",
    "sort_order",
]
SETLIST_FIELDS: Final = [
    "performance_key",
    "sort_order",
    "entry_type",
    "setlist_no_raw",
    "song_id",
    "song_title_raw",
    "artist_credit_raw",
    "note_raw",
    "marker_label",
    "notes",
    "joucho_participation",
]


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def required(value: object, label: str) -> str:
    result = text(value)
    if not result.strip():
        raise RuntimeError(f"{label}が空です。")
    return result


def is_true(value: object) -> bool:
    return value is True or text(value).strip() == "TRUE"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise RuntimeError(f"CSV headerがありません: {path.name}")
        return [dict(row) for row in reader]


def read_sheet(sheet_name: str, expected_fields: list[str]) -> list[dict[str, str]]:
    workbook = load_workbook(LIVE_REVIEW_PATH, read_only=True, data_only=False)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = [text(value) for value in next(rows)]
        if header != expected_fields:
            raise RuntimeError(f"{sheet_name}シートの列構成が想定外です。")
        return [
            {field: text(value) for field, value in zip(header, row, strict=True)}
            for row in rows
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()


def parse_song_id(value: str, page_key: str) -> int | None:
    if value == "":
        return None
    try:
        parsed = int(value)
    except ValueError as error:
        raise RuntimeError(f"{page_key}: song_idが整数ではありません。") from error
    if parsed <= 0:
        raise RuntimeError(f"{page_key}: song_idは正の整数である必要があります。")
    return parsed


def parse_participation(value: str, page_key: str) -> bool | None:
    if value == "":
        return None
    if value == "TRUE":
        return True
    if value == "FALSE":
        return False
    raise RuntimeError(
        f"{page_key}: joucho_participation={value!r} はDB booleanへ確定できません。"
    )


def build_rows() -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    dict[str, int],
]:
    raw_events = read_csv(EVENTS_RAW_PATH)
    raw_setlists = read_csv(SETLISTS_RAW_PATH)
    review_events = read_sheet("events", EVENT_REVIEW_FIELDS)
    review_setlists = read_sheet("setlists", SETLIST_REVIEW_FIELDS)

    raw_event_by_key = {row["page_key"]: row for row in raw_events}
    review_event_by_key = {row["page_key"]: row for row in review_events}
    target_keys = sorted(
        key
        for key, row in review_event_by_key.items()
        if is_true(row["import_target"]) and key not in PILOT_PAGE_KEYS
    )
    if len(target_keys) != EXPECTED_PERFORMANCE_COUNT:
        raise RuntimeError(
            f"pilot除外後のTRUE件数が{EXPECTED_PERFORMANCE_COUNT}件ではありません: "
            f"{len(target_keys)}"
        )

    raw_setlists_by_key: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in raw_setlists:
        if row["page_key"] in target_keys:
            raw_setlists_by_key[row["page_key"]].append(row)
    review_setlists_by_key: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in review_setlists:
        if row["page_key"] in target_keys:
            review_setlists_by_key[row["page_key"]].append(row)

    performances: list[dict[str, object]] = []
    sources: list[dict[str, object]] = []
    links: list[dict[str, object]] = []
    setlists: list[dict[str, object]] = []
    expected_counts: dict[str, int] = {}

    for key in target_keys:
        raw_event = raw_event_by_key.get(key)
        review = review_event_by_key[key]
        if raw_event is None:
            raise RuntimeError(f"{key}: events_raw.csvに存在しません。")
        for raw_field, review_field in (
            ("source_category", "source_category"),
            ("event_title_raw", "event_title_raw"),
            ("event_date_raw", "event_date_raw"),
            ("source_url", "source_url"),
        ):
            if raw_event[raw_field] != review[review_field]:
                raise RuntimeError(f"{key}: event raw値がreviewと一致しません: {raw_field}")

        title = required(review["title"], f"{key} title")
        # live_performances.artist_credit is nullable. A deliberately blank
        # reviewed value must remain NULL rather than being inferred from raw.
        artist_credit = review["artist_credit"] or None
        performance_date = required(
            review["performance_date"], f"{key} performance_date"
        )
        try:
            if date.fromisoformat(performance_date).isoformat() != performance_date:
                raise ValueError
        except ValueError as error:
            raise RuntimeError(f"{key}: performance_dateがISO dateではありません。") from error
        performance_name = review["performance_name"]
        if performance_name and performance_name != title:
            raise RuntimeError(
                f"{key}: performance_nameを格納するDB列がなくtitleとも異なります。"
            )

        performances.append(
            {
                "performance_key": key,
                "local_group_key": None,
                "group_sort_order": None,
                "title": title,
                "title_kana": review["title_kana"] or None,
                "sort_title": None,
                "artist_credit": artist_credit,
                "performance_date": performance_date,
                "format_label": review["event_type"] or None,
                "image_url": None,
                "venue": review["venue"] or None,
                "streaming_platforms": [],
                "notes": review["review_note"] or None,
                "published_at": None,
                "is_listed": False,
            }
        )
        sources.append(
            {
                "performance_key": key,
                "source_type": "thinkr_wiki",
                "source_url": raw_event["source_url"],
                "source_key": key,
                "event_title_raw": raw_event["event_title_raw"],
                "event_date_raw": raw_event["event_date_raw"],
                "source_category_raw": raw_event["source_category"],
                "notes": None,
            }
        )
        if review["official_url"]:
            links.append(
                {
                    "performance_key": key,
                    "link_type": "official",
                    "label": None,
                    "url": review["official_url"],
                    "published_date": None,
                    "notes": None,
                    "sort_order": 100,
                }
            )

        raw_rows = raw_setlists_by_key[key]
        review_rows = review_setlists_by_key[key]
        if len(raw_rows) != len(review_rows):
            raise RuntimeError(
                f"{key}: raw/review setlist件数が一致しません "
                f"({len(raw_rows)} != {len(review_rows)})"
            )
        expected_counts[key] = len(raw_rows)
        for index, (raw_row, review_row) in enumerate(
            zip(raw_rows, review_rows, strict=True), start=1
        ):
            for field in (
                "page_key",
                "event_title_raw",
                "event_date_raw",
                "setlist_table_index",
                "setlist_no_raw",
                "song_title_raw",
                "artist_credit_raw",
                "note_raw",
                "source_url",
            ):
                if raw_row[field] != review_row[field]:
                    raise RuntimeError(
                        f"{key} row {index}: raw/reviewが一致しません: {field}"
                    )
            if not is_true(review_row["import_target"]):
                raise RuntimeError(f"{key} row {index}: import_targetがTRUEではありません。")
            setlists.append(
                {
                    "performance_key": key,
                    "sort_order": index * 100,
                    "entry_type": "song",
                    "setlist_no_raw": raw_row["setlist_no_raw"] or None,
                    "song_id": parse_song_id(review_row["song_id"], key),
                    "song_title_raw": required(
                        raw_row["song_title_raw"], f"{key} row {index} song_title_raw"
                    ),
                    "artist_credit_raw": raw_row["artist_credit_raw"] or None,
                    "note_raw": raw_row["note_raw"] or None,
                    "marker_label": None,
                    "notes": review_row["review_note"] or None,
                    "joucho_participation": parse_participation(
                        review_row["joucho_participation"], key
                    ),
                }
            )

    return [], performances, sources, links, setlists, expected_counts


def validate_rows(
    event_groups: list[dict[str, object]],
    performances: list[dict[str, object]],
    sources: list[dict[str, object]],
    links: list[dict[str, object]],
    setlists: list[dict[str, object]],
    expected_counts: dict[str, int],
) -> None:
    if event_groups:
        raise RuntimeError("このbatchでは新規event groupを作成しません。")
    if len(performances) != EXPECTED_PERFORMANCE_COUNT:
        raise RuntimeError("performances件数が26件ではありません。")
    keys = [str(row["performance_key"]) for row in performances]
    if len(set(keys)) != len(keys):
        raise RuntimeError("performance_keyがuniqueではありません。")
    for row in performances:
        if not row["title"]:
            raise RuntimeError("titleが空のperformanceがあります。")
        if row["local_group_key"] is not None or row["group_sort_order"] is not None:
            raise RuntimeError("このbatchにparent relationが混入しています。")
        if row["streaming_platforms"] != []:
            raise RuntimeError("streaming_platformsが空配列ではありません。")
        if row["published_at"] is not None or row["is_listed"] is not False:
            raise RuntimeError("公開状態が非公開初期値ではありません。")

    if Counter(str(row["performance_key"]) for row in sources) != Counter(keys):
        raise RuntimeError("各performanceにsourceが1件ずつありません。")
    for row in sources:
        key = str(row["performance_key"])
        if row["source_type"] != "thinkr_wiki" or row["source_key"] != key:
            raise RuntimeError(f"{key}: source type/keyが不正です。")
        if not row["source_url"]:
            raise RuntimeError(f"{key}: source_urlが空です。")

    for row in links:
        if row["performance_key"] not in keys or row["link_type"] != "official":
            raise RuntimeError("performance linkの参照または種別が不正です。")
        if not row["url"]:
            raise RuntimeError("performance link URLが空です。")

    actual_counts = Counter(str(row["performance_key"]) for row in setlists)
    if actual_counts != Counter(expected_counts):
        raise RuntimeError("setlist件数がraw集計と一致しません。")
    for key in keys:
        rows = [row for row in setlists if row["performance_key"] == key]
        expected_orders = list(range(100, len(rows) * 100 + 1, 100))
        orders = [int(row["sort_order"]) for row in rows]
        if orders != expected_orders or len(set(orders)) != len(orders):
            raise RuntimeError(f"{key}: sort_orderまたはraw行順が不正です。")
    if any(row["entry_type"] != "song" for row in setlists):
        raise RuntimeError("markerまたは未知のentry typeが混入しています。")
    if any(row["marker_label"] is not None for row in setlists):
        raise RuntimeError("marker_labelが混入しています。")


def csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def write_csv(
    path: Path, fields: list[str], rows: list[dict[str, object]]
) -> None:
    with path.open("x", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(
            {field: csv_value(row.get(field)) for field in fields} for row in rows
        )


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def manifest_text(
    performances: list[dict[str, object]],
    sources: list[dict[str, object]],
    links: list[dict[str, object]],
    setlists: list[dict[str, object]],
) -> str:
    generated_at = datetime.now(ZoneInfo("Asia/Tokyo")).isoformat(timespec="seconds")
    linked = sum(row["song_id"] is not None for row in setlists)
    participation = sum(row["joucho_participation"] is not None for row in setlists)
    lines = [
        f"# {BATCH_NAME}",
        "",
        f"- Generated at (JST): {generated_at}",
        "- Input scope: import_target=TRUE excluding pilot-001 page keys",
        "- Event groups: 0",
        f"- Performance rows: {len(performances)}",
        f"- Source rows: {len(sources)}",
        f"- Performance link rows: {len(links)}",
        f"- Setlist entries: {len(setlists)}",
        f"- Linked song_id values: {linked}",
        f"- Reviewed joucho_participation values: {participation}",
        "- Markers: 0",
        "- Published performances: 0",
        "- Listed performances: 0",
        "",
        "## Performance keys",
        "",
    ]
    lines.extend(f"- {row['performance_key']}: {row['title']}" for row in performances)
    lines.extend(
        [
            "",
            "## Input fingerprints (SHA-256)",
            "",
            f"- events_raw.csv: `{file_hash(EVENTS_RAW_PATH)}`",
            f"- setlists_raw.csv: `{file_hash(SETLISTS_RAW_PATH)}`",
            f"- live_review.xlsx: `{file_hash(LIVE_REVIEW_PATH)}`",
            "",
            "CSV empty cells represent NULL. `streaming_platforms` contains the JSON "
            "literal `[]`, which represents a non-NULL empty PostgreSQL text array.",
            "",
        ]
    )
    return "\n".join(lines)


def create_output() -> None:
    if OUTPUT_PATH.exists():
        raise RuntimeError(f"既存readyを上書きしません: {OUTPUT_PATH}")
    READY_ROOT.mkdir(parents=True, exist_ok=True)
    event_groups, performances, sources, links, setlists, counts = build_rows()
    validate_rows(event_groups, performances, sources, links, setlists, counts)

    temporary = Path(tempfile.mkdtemp(prefix=f".{BATCH_NAME}-", dir=READY_ROOT))
    try:
        write_csv(temporary / "event_groups.csv", EVENT_GROUP_FIELDS, event_groups)
        write_csv(temporary / "performances.csv", PERFORMANCE_FIELDS, performances)
        write_csv(temporary / "performance_sources.csv", SOURCE_FIELDS, sources)
        write_csv(temporary / "performance_links.csv", LINK_FIELDS, links)
        write_csv(temporary / "setlist_entries.csv", SETLIST_FIELDS, setlists)
        (temporary / "manifest.md").write_text(
            manifest_text(performances, sources, links, setlists),
            encoding="utf-8",
            newline="\n",
        )
        os.replace(temporary, OUTPUT_PATH)
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise

    print(f"ready created: {OUTPUT_PATH}")
    print(f"performances: {len(performances)}")
    print(f"sources: {len(sources)}")
    print(f"performance links: {len(links)}")
    print(f"setlist entries: {len(setlists)}")
    print(f"song_id non-null: {sum(row['song_id'] is not None for row in setlists)}")
    print(
        "joucho_participation non-null: "
        f"{sum(row['joucho_participation'] is not None for row in setlists)}"
    )
    print("validation: PASS")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    try:
        create_output()
        return 0
    except (OSError, RuntimeError, ValueError) as error:
        print(f"ready生成を停止しました: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
